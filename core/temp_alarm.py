"""Temp alarm exclusion against Power alarm coverage windows."""

import json
import math
import re
import zipfile
from dataclasses import replace
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from alarm_app.core.backup_time import fmt_td
    from alarm_app.core.duration import duration_to_secs
    from alarm_app.data.alarm_store import AlarmQuery, query_alarms
except ImportError:
    from core.backup_time import fmt_td
    from core.duration import duration_to_secs
    from data.alarm_store import AlarmQuery, query_alarms


TEMP_SUMMARY_BASE_COLUMNS = [
    "##",
    "Site Name",
    "Site Code",
    "Area",
    "Contractor",
    "No. Of HT Alarms",
    "HT Duration",
    "Batteries Types",
    "Batteries Status",
    "Week No.",
]

TEMP_RAW_HT_COLUMNS = [
    "Alarm Source",
    "Site Name",
    "Last Occurred On",
    "Cleared On",
    "Duration(hh:mm:ss)",
    "Alarm Name",
    "Clearance Status",
    "Cleared By",
    "Alarm Reporting Type",
    "Week",
    "Area",
]

TEMP_RAW_POWER_COLUMNS = [
    "Alarm Source",
    "Site Name",
    "Support",
    "Day",
    "Last Occurred On",
    "Cleared On",
    "Duration(hh:mm:ss)",
    "Alarm Name",
    "Clearance Status",
    "Cleared By",
    "Alarm Reporting Type",
    "SUM",
]

TEMP_STUDY_COLUMNS = [
    "Alarm Source",
    "Site Name",
    "Support",
    "Day",
    "Last Occurred On",
    "Cleared On",
    "Duration(hh:mm:ss)",
    "Alarm Name",
    "Clearance Status",
    "Cleared By",
    "Alarm Reporting Type",
    "HT SUM IFS",
    "Powr SUM IFS",
    "Diff",
    "Meet",
]

TEMP_MEET_COLUMNS = [
    "Site Name",
    "Alarm Source",
    "Last Occurred On",
    "Cleared On",
    "Duration(hh:mm:ss)",
    "Alarm Name",
    "Clearance Status",
    "Cleared By",
    "Alarm Reporting Type",
]

HT_MEET_THRESHOLD_SECONDS = 7 * 60 * 60
DEFAULT_HT_HISTORY_START_WEEK = "W40-22"


def _normalize_margin_minutes(margin_minutes: int | None) -> int:
    try:
        value = int(margin_minutes or 0)
    except (TypeError, ValueError):
        value = 0
    return max(0, value)


def _safe_margin_delta(margin_minutes: int) -> pd.Timedelta | None:
    if margin_minutes <= 0:
        return pd.Timedelta(0)
    try:
        return pd.Timedelta(minutes=margin_minutes)
    except (OverflowError, ValueError):
        return None


def _fmt_table_datetime(value) -> str:
    if pd.isna(value):
        return ""
    return pd.Timestamp(value).strftime("%Y-%m-%d %H:%M:%S")


def _fmt_delay(start, end) -> str:
    if pd.isna(start) or pd.isna(end) or end < start:
        return ""
    return _fmt_seconds_to_hhmmss((end - start).total_seconds())


def compute_temp_alarm_matches(df: pd.DataFrame, margin_minutes: int = 60):
    """Find Temp alarms not covered by a same-site Power window plus Y margin."""
    if df.empty or "alarm_category" not in df.columns:
        return pd.DataFrame(), "No data loaded."

    margin_minutes = _normalize_margin_minutes(margin_minutes)
    need = [
        "site_id", "alarm_source", "alarm_name", "occurred_on", "cleared_on",
        "duration", "alarm_category", "network_type", "vendor", "clearance_status",
        "site_name", "site_code", "area", "contractor", "battery_type",
        "battery_status", "battery_brand",
    ]
    sub = df[[c for c in need if c in df.columns]].copy()
    sub = sub.dropna(subset=["site_id", "occurred_on"])
    sub["site_id"] = sub["site_id"].astype(str).str.strip()
    sub["occurred_on"] = pd.to_datetime(sub["occurred_on"], errors="coerce", format="mixed")
    if "cleared_on" not in sub.columns:
        sub["cleared_on"] = pd.NaT
    sub["cleared_on"] = pd.to_datetime(sub["cleared_on"], errors="coerce", format="mixed")
    sub = sub.dropna(subset=["occurred_on"])

    pwr = sub[sub["alarm_category"] == "Power"].copy()
    tmp = sub[sub["alarm_category"] == "Temp"].copy()
    if tmp.empty:
        return pd.DataFrame(), "No Temp alarms found in loaded data."

    valid_power = pwr.dropna(subset=["occurred_on"]).copy()
    valid_power["_active_power"] = valid_power["cleared_on"].isna()
    valid_power["_power_cleared_raw"] = valid_power["cleared_on"]
    valid_power["cleared_on"] = valid_power["cleared_on"].fillna(pd.Timestamp.max)
    valid_power = valid_power[valid_power["cleared_on"] >= valid_power["occurred_on"]]
    valid_power = valid_power.sort_values(["site_id", "occurred_on"]).reset_index(drop=True)
    power_by_site: dict[str, tuple[pd.Series, pd.Series, pd.DataFrame]] = {}
    for site_id, group in valid_power.groupby("site_id", sort=False):
        site_power_rows = group.reset_index(drop=True)
        coverage_end = site_power_rows["cleared_on"].reset_index(drop=True)
        if margin_minutes:
            active_power = site_power_rows["_active_power"].reset_index(drop=True)
            finite_power = ~active_power
            coverage_end = coverage_end.copy()
            margin = _safe_margin_delta(margin_minutes)
            if margin is None:
                coverage_end.loc[finite_power] = pd.Timestamp.max
            else:
                safe_to_add_margin = finite_power & (coverage_end <= pd.Timestamp.max - margin)
                coverage_end.loc[safe_to_add_margin] = coverage_end.loc[safe_to_add_margin] + margin
                overflow_margin = finite_power & ~safe_to_add_margin
                if overflow_margin.any():
                    coverage_end.loc[overflow_margin] = pd.Timestamp.max
        coverage_end = coverage_end.cummax().reset_index(drop=True)
        power_by_site[str(site_id)] = (site_power_rows["occurred_on"].reset_index(drop=True), coverage_end, site_power_rows)

    uncovered_parts: list[pd.DataFrame] = []
    context_columns = [
        "_power_time", "_power_cleared", "_x_duration",
        "_temp_delay_after_power", "_temp_delay_after_power_clearance",
        "_match_window",
    ]
    for site_id, group in tmp.sort_values(["site_id", "occurred_on"]).groupby("site_id", sort=False):
        site_id = str(site_id)
        if not site_id:
            continue
        site_power = power_by_site.get(site_id)
        if site_power is None:
            uncovered = group.copy()
            for column in context_columns:
                uncovered[column] = ""
            uncovered["_match_window"] = "No same-site Power alarm before Temp"
            uncovered_parts.append(uncovered)
            continue
        starts, coverage_end, site_power_rows = site_power
        temp_times = group["occurred_on"]
        indexes = starts.searchsorted(temp_times, side="right") - 1
        covered = pd.Series(False, index=group.index)
        has_prior_power = indexes >= 0
        if has_prior_power.any():
            covered.loc[has_prior_power] = coverage_end.iloc[indexes[has_prior_power]].to_numpy() >= temp_times.loc[has_prior_power].to_numpy()
        uncovered = group.loc[~covered].copy()
        if not uncovered.empty:
            for column in context_columns:
                uncovered[column] = ""
            uncovered["_match_window"] = "No same-site Power alarm before Temp"
            prior_positions = pd.Series(indexes, index=group.index).loc[uncovered.index]
            for temp_idx, prior_pos in prior_positions.items():
                if prior_pos < 0:
                    continue
                power = site_power_rows.iloc[int(prior_pos)]
                temp_time = uncovered.at[temp_idx, "occurred_on"]
                power_time = power.get("occurred_on")
                power_cleared = power.get("_power_cleared_raw")
                uncovered.at[temp_idx, "_power_time"] = _fmt_table_datetime(power_time)
                uncovered.at[temp_idx, "_power_cleared"] = _fmt_table_datetime(power_cleared)
                uncovered.at[temp_idx, "_x_duration"] = _temp_duration(power, power_time, power_cleared)
                uncovered.at[temp_idx, "_temp_delay_after_power"] = _fmt_delay(power_time, temp_time)
                uncovered.at[temp_idx, "_temp_delay_after_power_clearance"] = _fmt_delay(power_cleared, temp_time)
                uncovered.at[temp_idx, "_match_window"] = "Outside Power coverage"
            uncovered_parts.append(uncovered)

    if not uncovered_parts:
        return pd.DataFrame(), "No uncovered Temp alarms found outside Power windows."
    uncovered = pd.concat(uncovered_parts, ignore_index=True).sort_values(["site_id", "occurred_on"]).reset_index(drop=True)
    rows = pd.DataFrame({
        "site_id": uncovered["site_id"],
        "network_type": uncovered["network_type"] if "network_type" in uncovered.columns else "",
        "vendor": uncovered["vendor"] if "vendor" in uncovered.columns else "",
        "power_time": uncovered["_power_time"],
        "power_cleared": uncovered["_power_cleared"],
        "x_duration": uncovered["_x_duration"],
        "y_margin": f"{margin_minutes} min",
        "temp_time": uncovered["occurred_on"].dt.strftime("%Y-%m-%d %H:%M:%S").fillna(""),
        "temp_cleared": uncovered["cleared_on"].dt.strftime("%Y-%m-%d %H:%M:%S").fillna(""),
        "temp_delay_after_power": uncovered["_temp_delay_after_power"],
        "temp_delay_after_power_clearance": uncovered["_temp_delay_after_power_clearance"],
        "temp_clear_duration": _temp_duration_series(uncovered),
        "temp_alarm_name": uncovered["alarm_name"] if "alarm_name" in uncovered.columns else "",
        "temp_alarm_source": uncovered["alarm_source"] if "alarm_source" in uncovered.columns else "",
        "temp_clearance_status": uncovered["clearance_status"] if "clearance_status" in uncovered.columns else "",
        "match_window": uncovered["_match_window"],
        "site_name": _metadata_series(uncovered, "site_name"),
        "site_code": _metadata_series(uncovered, "site_code", fallback_col="site_id"),
        "area": _metadata_series(uncovered, "area"),
        "contractor": _metadata_series(uncovered, "contractor"),
        "battery_type": _battery_type_series(uncovered),
        "battery_status": _metadata_series(uncovered, "battery_status"),
    })
    return rows.reset_index(drop=True), ""


def compute_temp_alarm_matches_for_query(
    alarm_query: AlarmQuery | None = None,
    margin_minutes: int = 60,
    result_filter_query: AlarmQuery | None = None,
    include_full_temp_source: bool = False,
):
    """Load a targeted alarm subset from DuckDB, then run temp coverage analysis."""
    query = alarm_query or AlarmQuery()
    if result_filter_query is not None:
        temp_query = replace(
            result_filter_query,
            category="Temp",
            limit=None,
            offset=0,
            sort_by=None,
            sort_desc=False,
        )
        temp_df = query_alarms(temp_query)
        if temp_df.empty:
            return pd.DataFrame(), "No Temp alarms found in selected data.", temp_df
        site_ids = sorted({str(v).strip() for v in temp_df.get("site_id", pd.Series(dtype=object)).dropna() if str(v).strip()})
        selected_site_ids = {_normalize_site_identifier(value) for value in site_ids}
        site_scope_keys = site_ids if len(site_ids) <= 500 else None
        power_query = replace(
            query,
            site_text="",
            category="Power",
            vendor="All",
            network_type="All",
            min_duration_secs=None,
            date_from=None,
            date_to=_source_query_date_to(result_filter_query),
            manual_days=None,
            site_scope_keys=site_scope_keys,
            allowed_values={},
            column_filters={},
            col_filters={},
            limit=None,
            offset=0,
            sort_by=None,
            sort_desc=False,
        )
        power_df = query_alarms(power_query)
        if include_full_temp_source:
            temp_source_query = replace(
                query,
                category="Temp",
                site_scope_keys=site_scope_keys,
                limit=None,
                offset=0,
                sort_by=None,
                sort_desc=False,
            )
            source_temp_df = query_alarms(temp_source_query)
            if selected_site_ids and "site_id" in source_temp_df.columns:
                source_site_ids = source_temp_df["site_id"].map(_normalize_site_identifier)
                source_temp_df = source_temp_df[source_site_ids.isin(selected_site_ids)]
        else:
            source_temp_df = temp_df
        df = pd.concat([power_df, source_temp_df], ignore_index=True)
        analysis_df = pd.concat([power_df, temp_df], ignore_index=True)
    else:
        df = query_alarms(query)
        analysis_df = df
    result, err = compute_temp_alarm_matches(analysis_df, margin_minutes=margin_minutes)
    return result, err, df


def _source_query_date_to(query: AlarmQuery) -> object:
    if query.date_to is not None:
        return query.date_to
    if query.manual_days is None:
        return None
    days = [pd.Timestamp(day).date() for day in query.manual_days if not pd.isna(pd.Timestamp(day))]
    return max(days) if days else None


def filter_temp_matches_to_query(matches: pd.DataFrame, alarm_query: AlarmQuery | None) -> pd.DataFrame:
    """Limit widened-query matches back to the user's original Temp alarm date scope."""
    if matches.empty or alarm_query is None or "temp_time" not in matches.columns:
        return matches
    times = pd.to_datetime(matches["temp_time"], errors="coerce", format="mixed")
    mask = pd.Series(True, index=matches.index)
    if alarm_query.date_from is not None:
        start = pd.Timestamp(alarm_query.date_from).normalize()
        mask &= times >= start
    if alarm_query.date_to is not None:
        end = pd.Timestamp(alarm_query.date_to).normalize() + pd.Timedelta(days=1)
        mask &= times < end
    if alarm_query.manual_days is not None:
        days = {
            pd.Timestamp(day).date()
            for day in alarm_query.manual_days
            if not pd.isna(pd.Timestamp(day))
        }
        if days:
            mask &= times.dt.date.isin(days)
    return matches[mask].reset_index(drop=True)


def filter_temp_matches_to_selected_temps(matches: pd.DataFrame, selected_temp: pd.DataFrame) -> pd.DataFrame:
    """Keep only matches whose Temp row exists in the already-filtered Temp set."""
    if matches.empty or selected_temp.empty or "site_id" not in matches.columns or "temp_time" not in matches.columns:
        return matches
    if "site_id" not in selected_temp.columns or "occurred_on" not in selected_temp.columns:
        return matches
    selected_key_cols = ["site_id", "occurred_on", "cleared_on", "alarm_name", "alarm_source"]
    match_key_cols = ["site_id", "temp_time", "temp_cleared", "temp_alarm_name", "temp_alarm_source"]
    if not all(col in selected_temp.columns for col in selected_key_cols) or not all(col in matches.columns for col in match_key_cols):
        selected_key_cols = ["site_id", "occurred_on"]
        match_key_cols = ["site_id", "temp_time"]
    selected_keys = {
        _temp_match_key(row, selected_key_cols)
        for row in selected_temp[selected_key_cols].dropna(subset=["site_id", "occurred_on"]).itertuples(index=False, name=None)
    }
    if not selected_keys:
        return matches.iloc[0:0].copy()
    keys = [_temp_match_key(row, match_key_cols) for row in matches[match_key_cols].itertuples(index=False, name=None)]
    mask = pd.Series([key in selected_keys for key in keys], index=matches.index)
    return matches[mask].reset_index(drop=True)


def _temp_match_key(row: tuple, columns: list[str]) -> tuple[str, ...]:
    values = []
    for col, value in zip(columns, row):
        if col in {"occurred_on", "cleared_on", "temp_time", "temp_cleared"}:
            timestamp = pd.to_datetime(value, errors="coerce")
            if pd.isna(timestamp):
                values.append("")
            else:
                values.append(pd.Timestamp(timestamp).strftime("%Y-%m-%d %H:%M:%S"))
        else:
            values.append(str(value or "").strip())
    return tuple(values)


def _metadata_series(df: pd.DataFrame, column: str, fallback_col: str | None = None):
    if column in df.columns:
        values = df[column]
    elif fallback_col and fallback_col in df.columns:
        values = df[fallback_col]
    else:
        return ""
    return values.fillna("").astype(str).str.strip()


def _metadata_from_aliases(df: pd.DataFrame, columns: list[str]):
    for column in columns:
        if column in df.columns:
            return _metadata_series(df, column)
    normalized = {_normalize_column_name(column): column for column in df.columns}
    for column in columns:
        actual = normalized.get(_normalize_column_name(column))
        if actual:
            return _metadata_series(df, actual)
    return ""


def _metadata_from_aliases_or_default(df: pd.DataFrame, columns: list[str], default: str):
    values = _metadata_from_aliases(df, columns)
    if isinstance(values, str):
        return default if not values else values
    normalized = values.fillna("").astype(str).str.strip()
    return normalized.where(normalized != "", default)


def _normalize_column_name(value) -> str:
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


def _support_series(df: pd.DataFrame):
    values = _metadata_from_aliases(df, ["support", "Support"])
    if isinstance(values, str):
        return values
    return values


def _day_series(df: pd.DataFrame):
    values = _metadata_from_aliases(df, ["day", "Day"])
    if not isinstance(values, str):
        non_empty = values.fillna("").astype(str).str.strip()
        if (non_empty != "").any():
            return non_empty
    if "occurred_on" not in df.columns:
        return ""
    occurred = pd.to_datetime(df["occurred_on"], errors="coerce", format="mixed")
    return occurred.dt.day_name().fillna("")


def _battery_type_series(df: pd.DataFrame):
    if "battery_type" in df.columns:
        return _metadata_series(df, "battery_type")
    return _metadata_series(df, "battery_brand")


def ht_export_week_from_date(value) -> dict[str, object]:
    """Return Reference Workbook-compatible HT export week metadata for a date."""
    timestamp = pd.Timestamp(value)
    week_label = _week_label_from_timestamp(timestamp)
    start, end = ht_export_week_range(week_label)
    return {
        "week_label": week_label,
        "short_week_label": _short_week_label(week_label),
        "start": start,
        "end": end,
        "filename": ht_export_filename(week_label),
    }


def ht_export_week_range(week_label: str) -> tuple[pd.Timestamp, pd.Timestamp]:
    """Return Sunday-inclusive to next-Sunday-exclusive range for a Wnn-yy label."""
    parsed = _parse_week_label(week_label)
    if parsed is None:
        raise ValueError(f"Invalid HT export week label: {week_label!r}")
    year, week = parsed
    first_day = date(year, 1, 1)
    max_week = int(pd.Timestamp(date(year, 12, 31)).strftime("%U")) + 1
    if week > max_week:
        raise ValueError(f"Could not resolve HT export week range: {week_label!r}")
    if week == 1:
        start = pd.Timestamp(first_day).normalize()
        return start, start + pd.Timedelta(days=7)
    days_until_first_sunday = (6 - first_day.weekday()) % 7
    first_sunday = pd.Timestamp(first_day + timedelta(days=days_until_first_sunday)).normalize()
    start = first_sunday + pd.Timedelta(weeks=week - 2)
    return start, start + pd.Timedelta(days=7)


def ht_export_filename(week_label: str) -> str:
    """Return Reference Workbook-compatible HT export filename."""
    parsed = _parse_week_label(week_label)
    if parsed is None:
        raise ValueError(f"Invalid HT export week label: {week_label!r}")
    year, week = parsed
    return f"{year}-HT-Alarms-W{week:02d}.xlsx"


def compute_ht_meet_rows(source_df: pd.DataFrame | None, week_label: str | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build HT Study and Meet rows using the Reference Workbook daily Meet rule."""
    study, meet, _ = _compute_ht_meet_frames(source_df, week_label=week_label)
    return study, meet


def _first_metadata(df: pd.DataFrame, column: str, fallback_col: str | None = None) -> str:
    value = _first_text(df, column)
    if value:
        return value
    if fallback_col:
        return _first_text(df, fallback_col)
    return ""


def build_temp_alarm_summary(
    matches: pd.DataFrame,
    week_label: str | None = None,
    rolling_week_label: str | None = None,
) -> pd.DataFrame:
    """Build a W27-style weekly summary for Reference Workbook Meet rows."""
    week = rolling_week_label or week_label or _week_label_from_summary_rows(matches)
    week_columns = _week_history_columns(week) if week else []
    columns = TEMP_SUMMARY_BASE_COLUMNS + week_columns
    if matches.empty:
        return pd.DataFrame(columns=columns)

    data = matches.copy()
    time_values = _summary_time_values(data)
    data["_week_label"] = time_values.apply(_week_label_from_timestamp)
    data = data[data["_week_label"] != ""]
    if week_label:
        data = data[data["_week_label"] == week_label]
    if data.empty:
        return pd.DataFrame(columns=columns)

    site_key = _summary_site_key(data)
    data["_site_key"] = site_key.map(_normalize_site_identifier).where(site_key != "", _metadata_series(data, "Site Name"))

    records = []
    group_columns = ["_site_key"] if week_label else ["_week_label", "_site_key"]
    for _, group in data.groupby(group_columns, sort=True):
        current_week = str(group["_week_label"].iloc[0])
        duration_secs = _summary_duration_seconds(group)
        records.append({
            "Site Name": _first_metadata(group, "site_name", fallback_col="Site Name") or _first_metadata(group, "Site Name"),
            "Site Code": _first_metadata(group, "site_code", fallback_col="site_id") or _first_metadata(group, "site_id"),
            "Area": _first_metadata(group, "area"),
            "Contractor": _first_metadata(group, "contractor"),
            "No. Of HT Alarms": len(group),
            "HT Duration": _fmt_hours_minutes(duration_secs),
            "Batteries Types": _first_metadata(group, "battery_type", fallback_col="battery_brand"),
            "Batteries Status": _first_metadata(group, "battery_status"),
            "Week No.": current_week,
            **{col: (col if col == current_week else "") for col in week_columns},
        })
    summary = pd.DataFrame(records).sort_values(["Week No.", "Site Code"]).reset_index(drop=True)
    summary.insert(0, "##", range(1, len(summary) + 1))
    return summary[columns]


def export_temp_alarm_workbook(
    matches: pd.DataFrame,
    path: str | Path,
    week_label: str | None = None,
    source_df: pd.DataFrame | None = None,
    margin_minutes: int = 60,
    site_metadata_df: pd.DataFrame | None = None,
    historical_start_week: str | None = DEFAULT_HT_HISTORY_START_WEEK,
    return_warnings: bool = False,
) -> dict[str, object] | None:
    """Export reference-style HT Alarm Workbook."""
    export_week = week_label or _week_label_from_source(source_df) or _week_label_from_matches(matches)
    missing_metadata = pd.DataFrame()
    if source_df is not None and site_metadata_df is not None:
        source_df, missing_metadata = enrich_source_with_site_metadata(source_df, site_metadata_df)
    scoped_matches = _filter_matches_to_week(matches, export_week) if export_week else matches
    scoped_source = _filter_source_to_week(source_df, export_week) if export_week and source_df is not None else source_df
    short_week = _short_week_label(export_week or "Summary")
    ht_sheet = f"{short_week} AUTIN HT"
    power_sheet = f"{short_week} AUTIN Power"
    ht_raw = _build_temp_raw_sheet(scoped_source, "Temp") if scoped_source is not None else _empty_frame(TEMP_RAW_HT_COLUMNS)
    power_raw = _build_temp_raw_sheet(scoped_source, "Power", sheet_name=power_sheet) if scoped_source is not None else _empty_frame(TEMP_RAW_POWER_COLUMNS)
    if scoped_source is not None:
        study, meet, meet_source = _compute_ht_meet_frames(scoped_source, week_label=export_week, ht_sheet=ht_sheet, power_sheet=power_sheet)
        history_source = _filter_source_from_week(source_df, historical_start_week)
        _, _, consolidated_source = _compute_ht_meet_frames(history_source, week_label=None, ht_sheet=ht_sheet, power_sheet=power_sheet)
    else:
        study = _empty_frame(TEMP_STUDY_COLUMNS)
        meet = _empty_frame(TEMP_MEET_COLUMNS)
        meet_source = scoped_matches
        consolidated_source = matches
    summary = build_temp_alarm_summary(meet_source, week_label=export_week)
    consolidated = build_temp_alarm_summary(consolidated_source, week_label=None, rolling_week_label=export_week)
    path = Path(path)
    sheets = [
        (ht_sheet, ht_raw),
        (power_sheet, power_raw),
        (f"{short_week} AUTIN HT Study", study),
        ("Meet", meet),
        (short_week, summary),
        ("Consolidated", consolidated),
    ]
    if not missing_metadata.empty:
        sheets.append(("Missing Metadata", missing_metadata))
    _write_temp_alarm_workbook(path, sheets)
    if return_warnings:
        missing_ids = missing_metadata["Site ID"].dropna().astype(str).tolist() if not missing_metadata.empty else []
        return {"missing_metadata_site_ids": missing_ids, "missing_metadata_count": len(missing_ids)}
    return None


def enrich_source_with_site_metadata(source_df: pd.DataFrame, site_metadata_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Enrich alarm rows by exact normalized Site ID, with alarm-source token fallback only."""
    if source_df is None or source_df.empty or site_metadata_df is None or site_metadata_df.empty:
        return source_df, pd.DataFrame(columns=["Site ID", "Alarm Source", "Reason"])
    metadata = _metadata_lookup(site_metadata_df)
    if not metadata:
        return source_df, pd.DataFrame(columns=["Site ID", "Alarm Source", "Reason"])
    enriched = source_df.copy()
    missing_rows = []
    for idx, row in enriched.iterrows():
        raw_site_id = row.get("site_id") or row.get("site_code")
        site_id = _normalize_site_identifier(raw_site_id)
        meta = metadata.get(site_id) if site_id else None
        if meta is None:
            fallback = _site_id_from_alarm_source(row.get("alarm_source"), metadata.keys())
            if fallback:
                site_id = fallback
                meta = metadata.get(fallback)
        if meta is None:
            candidate = site_id or _normalize_site_identifier(row.get("alarm_source"))
            if candidate:
                missing_rows.append({"Site ID": candidate, "Alarm Source": row.get("alarm_source") or "", "Reason": "No Site Metadata match"})
            continue
        enriched.at[idx, "site_id"] = site_id
        enriched.at[idx, "site_code"] = site_id
        for target, keys in {
            "site_name": ["site_name", "sitename", "name"],
            "area": ["orange_area", "area", "orangearea"],
            "contractor": ["subcontractor", "contractor"],
            "battery_type": ["battery_type", "batterytype", "battery_brand", "batterybrand"],
            "battery_status": ["backup_status", "backupstatus", "battery_status", "batterystatus"],
        }.items():
            value = _first_meta_value(meta, keys)
            if value != "":
                enriched.at[idx, target] = value
    missing = pd.DataFrame(missing_rows).drop_duplicates().reset_index(drop=True) if missing_rows else pd.DataFrame(columns=["Site ID", "Alarm Source", "Reason"])
    return enriched, missing[["Site ID", "Alarm Source", "Reason"]]


def _metadata_lookup(site_metadata_df: pd.DataFrame) -> dict[str, dict[str, object]]:
    lookup: dict[str, dict[str, object]] = {}
    for _, row in site_metadata_df.iterrows():
        raw = row.to_dict()
        if raw.get("raw_data_json"):
            try:
                parsed = json.loads(str(raw.get("raw_data_json") or "{}"))
                if isinstance(parsed, dict):
                    raw.update(parsed)
            except (TypeError, ValueError):
                pass
        normalized = {_normalize_metadata_key(key): value for key, value in raw.items()}
        site_id = _normalize_site_identifier(raw.get("site_id") or raw.get("Code") or raw.get("code") or normalized.get("code"))
        if site_id:
            lookup[site_id] = {**raw, **normalized}
    return lookup


def _normalize_site_identifier(value) -> str:
    text = str(value or "").strip().upper()
    return "".join(ch for ch in text if ch.isalnum())


def _normalize_metadata_key(value) -> str:
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum() or ch == "_")


def _site_id_from_alarm_source(value, known_site_ids) -> str:
    known = set(known_site_ids)
    for token in re.split(r"[^A-Za-z0-9]+", str(value or "").upper()):
        normalized = _normalize_site_identifier(token)
        if normalized in known:
            return normalized
    return ""


def _first_meta_value(meta: dict[str, object], keys: list[str]) -> str:
    for key in keys:
        for candidate in {key, _normalize_metadata_key(key)}:
            value = meta.get(candidate)
            if value not in (None, "") and not pd.isna(value):
                return str(value).strip()
    return ""


def _write_temp_alarm_workbook(path: Path, sheets: list[tuple[str, pd.DataFrame]]) -> None:
    wb = Workbook(write_only=True)
    styles = _temp_workbook_styles()
    dimensions = []
    for title, frame in sheets:
        _append_temp_alarm_sheet(wb, title, frame, styles)
        dimensions.append(f"A1:{get_column_letter(max(len(frame.columns), 1))}{len(frame) + 1}")
    wb.save(path)
    _patch_xlsx_worksheet_dimensions(path, dimensions)


def _patch_xlsx_worksheet_dimensions(path: Path, dimensions: list[str]) -> None:
    tmp_path = path.with_name(f".{path.name}.tmp")
    dimension_pattern = re.compile(rb"<dimension\s+ref=\"[^\"]*\"\s*/>")
    try:
        with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                    try:
                        sheet_index = int(Path(item.filename).stem.replace("sheet", "")) - 1
                    except ValueError:
                        sheet_index = -1
                    if 0 <= sheet_index < len(dimensions):
                        dimension = f'<dimension ref="{dimensions[sheet_index]}"/>'.encode("ascii")
                        data = dimension_pattern.sub(b"", data)
                        sheet_pr_end = data.find(b"</sheetPr>")
                        if sheet_pr_end >= 0:
                            insert_at = sheet_pr_end + len(b"</sheetPr>")
                        else:
                            worksheet_start = data.find(b"<worksheet")
                            insert_at = data.find(b">", worksheet_start) + 1 if worksheet_start >= 0 else 0
                        if insert_at > 0:
                            data = data[:insert_at] + dimension + data[insert_at:]
                zout.writestr(item, data)
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def _temp_workbook_styles() -> dict[str, object]:
    border = Border(
        left=Side(style="thin", color="FF000000"),
        right=Side(style="thin", color="FF000000"),
        top=Side(style="thin", color="FF000000"),
        bottom=Side(style="thin", color="FF000000"),
    )
    return {
        "blue_fill": PatternFill("solid", fgColor="4F81BD"),
        "gold_fill": PatternFill("solid", fgColor="FFC000"),
        "green_fill": PatternFill("solid", fgColor="92D050"),
        "yellow_fill": PatternFill("solid", fgColor="FFFF00"),
        "header_font": Font(bold=True),
        "body_font": Font(size=11),
        "border": border,
        "header_alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "body_alignment": Alignment(vertical="center"),
    }


def _append_temp_alarm_sheet(wb: Workbook, title: str, frame: pd.DataFrame, styles: dict[str, object]) -> None:
    ws = wb.create_sheet(title)
    headers = list(frame.columns)
    display_headers = _temp_display_headers(title, headers)
    _apply_temp_sheet_layout(ws, title, len(headers), len(frame))
    ws.append([
        _temp_write_cell(
            ws,
            header,
            styles,
            fill=_temp_header_fill(title, idx, header, styles),
            bold=True,
            num_fmt=_temp_number_format(title, idx, header, is_header=True),
            header=True,
        )
        for idx, header in enumerate(display_headers, start=1)
    ])
    for row in frame.itertuples(index=False, name=None):
        ws.append([
            _temp_write_cell(
                ws,
                value,
                styles,
                fill=_temp_data_fill(title, idx, display_headers[idx - 1], styles),
                num_fmt=_temp_number_format(title, idx, display_headers[idx - 1], is_header=False),
            )
            for idx, value in enumerate(row, start=1)
        ])


def _temp_write_cell(
    ws,
    value,
    styles: dict[str, object],
    fill=None,
    bold: bool = False,
    num_fmt: str | None = None,
    header: bool = False,
):
    cell = WriteOnlyCell(ws, value=value)
    cell.border = styles["border"]
    cell.alignment = styles["header_alignment"] if header else styles["body_alignment"]
    cell.font = styles["header_font"] if bold else styles["body_font"]
    if fill is not None:
        cell.fill = fill
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _temp_display_headers(title: str, headers: list[str]) -> list[str]:
    display = list(headers)
    if title.endswith("AUTIN HT") and len(display) >= 5:
        display[4] = "Duration\n(hh:mm:ss)"
    if "HT Study" in title and len(display) >= 7:
        display[6] = "Duration\n(hh:mm:ss)"
    return display


def _temp_header_fill(title: str, col_idx: int, header: str, styles: dict[str, object]):
    if header in {"Support", "Day"}:
        return styles["yellow_fill"]
    if "HT Study" in title and 12 <= col_idx <= 15:
        return styles["green_fill"]
    if title == "Meet" or "AUTIN" in title:
        return styles["gold_fill"]
    return styles["blue_fill"]


def _temp_data_fill(title: str, col_idx: int, header: str, styles: dict[str, object]):
    if ("HT Study" in title or "AUTIN Power" in title) and header in {"Support", "Day"}:
        return styles["yellow_fill"]
    if "HT Study" in title and 12 <= col_idx <= 15:
        return styles["green_fill"]
    return None


def _temp_number_format(title: str, col_idx: int, header: str, is_header: bool) -> str | None:
    if header in {"Last Occurred On", "Cleared On", "HT Alarm", "HT Cleared", "Power Alarm", "Power Cleared"}:
        return "m/d/yy h:mm"
    if title.endswith("AUTIN HT") and col_idx == 5 and not is_header:
        return None
    if header.strip() in {
        "Duration(hh:mm:ss)",
        "Duration\n(hh:mm:ss)",
        "HT Duration",
        "Power Duration",
        "HT SUM IFS",
        "Powr SUM IFS",
        "HT SUM",
        "Power SUM",
        "Diff",
        "SUM",
    }:
        return "[hh]:mm"
    return None


def _apply_temp_sheet_layout(ws, title: str, column_count: int, row_count: int) -> None:
    if title == "Meet" or "AUTIN" in title:
        widths = {
            1: 47.43,
            2: 14.71,
            3: 14.71,
            4: 11.14,
            5: 19.43,
            6: 16.57,
            7: 14.57,
            8: 26.71,
            9: 15.57,
            10: 30.57,
            11: 14.43,
            12: 14.57,
            14: 9.14,
            15: 14.43,
        }
        for col in range(1, column_count + 1):
            ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 18)
        if "HT Study" in title:
            ws.freeze_panes = "A3722" if row_count + 1 >= 3722 else "A2"
    else:
        widths = [6, 34, 12, 12, 14, 16, 12, 20, 20, 12]
        for index in range(1, column_count + 1):
            width = widths[index - 1] if index <= len(widths) else 12
            ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[1].height = 45
    ws.auto_filter.ref = f"A1:{get_column_letter(max(column_count, 1))}{row_count + 1}"


def _empty_frame(columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=columns)


def _filter_matches_to_week(matches: pd.DataFrame, week_label: str | None) -> pd.DataFrame:
    if not week_label or matches.empty or "temp_time" not in matches.columns:
        return matches
    temp_weeks = pd.to_datetime(matches["temp_time"], errors="coerce").apply(_week_label_from_timestamp)
    return matches[temp_weeks == week_label].copy().reset_index(drop=True)


def _filter_source_to_week(source_df: pd.DataFrame | None, week_label: str | None) -> pd.DataFrame | None:
    if source_df is None or not week_label or source_df.empty or "occurred_on" not in source_df.columns:
        return source_df
    occurred = pd.to_datetime(source_df["occurred_on"], errors="coerce", format="mixed")
    source_weeks = occurred.apply(_week_label_from_timestamp)
    return source_df[source_weeks == week_label].copy().reset_index(drop=True)


def _filter_source_from_week(source_df: pd.DataFrame | None, start_week_label: str | None) -> pd.DataFrame | None:
    if source_df is None or source_df.empty or not start_week_label or "occurred_on" not in source_df.columns:
        return source_df
    try:
        start, _ = ht_export_week_range(start_week_label)
    except ValueError:
        return source_df
    occurred = pd.to_datetime(source_df["occurred_on"], errors="coerce", format="mixed")
    return source_df[occurred >= start].copy().reset_index(drop=True)


def _short_week_label(week_label: str) -> str:
    text = str(week_label or "").strip()
    return text.split("-", 1)[0] if "-" in text else text


def _build_temp_raw_sheet(
    source_df: pd.DataFrame | None,
    category: str,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    if source_df is None or source_df.empty or "alarm_category" not in source_df.columns:
        return _empty_frame(TEMP_RAW_HT_COLUMNS if category == "Temp" else TEMP_RAW_POWER_COLUMNS)
    data = source_df[source_df["alarm_category"] == category].copy().reset_index(drop=True)
    if data.empty:
        return _empty_frame(TEMP_RAW_HT_COLUMNS if category == "Temp" else TEMP_RAW_POWER_COLUMNS)
    data["occurred_on"] = pd.to_datetime(data.get("occurred_on"), errors="coerce", format="mixed")
    data["cleared_on"] = pd.to_datetime(data.get("cleared_on"), errors="coerce", format="mixed")
    if category == "Temp":
        week_numbers = data["occurred_on"].apply(_excel_week_number)
        rows = pd.DataFrame({
            "Alarm Source": _metadata_series(data, "alarm_source"),
            "Site Name": _metadata_series(data, "site_name", fallback_col="site_id"),
            "Last Occurred On": data["occurred_on"].apply(_fmt_dt),
            "Cleared On": data["cleared_on"].apply(_fmt_dt),
            "Duration(hh:mm:ss)": _temp_duration_series(data),
            "Alarm Name": _metadata_series(data, "alarm_name"),
            "Clearance Status": _metadata_series(data, "clearance_status"),
            "Cleared By": _metadata_from_aliases_or_default(data, ["cleared_by", "Cleared By"], "EMSReport"),
            "Alarm Reporting Type": _metadata_from_aliases_or_default(
                data, ["alarm_reporting_type", "Alarm Reporting Type"], "Real Time"
            ),
            "Week": week_numbers,
            "Area": _metadata_series(data, "area"),
        })
        return rows[TEMP_RAW_HT_COLUMNS]
    rows = pd.DataFrame({
        "Alarm Source": _metadata_series(data, "alarm_source"),
        "Site Name": _metadata_series(data, "site_name", fallback_col="site_id"),
        "Support": _support_series(data),
        "Day": _day_series(data),
        "Last Occurred On": data["occurred_on"].apply(_fmt_dt),
        "Cleared On": data["cleared_on"].apply(_fmt_dt),
        "Duration(hh:mm:ss)": _duration_series(data),
        "Alarm Name": _metadata_series(data, "alarm_name"),
        "Clearance Status": _metadata_series(data, "clearance_status"),
        "Cleared By": _metadata_from_aliases_or_default(data, ["cleared_by", "Cleared By"], "EMSReport"),
        "Alarm Reporting Type": _metadata_from_aliases_or_default(
            data, ["alarm_reporting_type", "Alarm Reporting Type"], "Real Time"
        ),
        "SUM": "",
    })
    for idx in range(len(rows)):
        row_number = idx + 2
        rows.at[rows.index[idx], "Support"] = f'=B{row_number}&" "&D{row_number}'
        rows.at[rows.index[idx], "Day"] = f"=DAY(E{row_number})"
        rows.at[rows.index[idx], "SUM"] = f"=SUMIFS(G:G,B:B,B{row_number},D:D,D{row_number})"
    return rows[TEMP_RAW_POWER_COLUMNS]


def _build_temp_study_sheet(source_df: pd.DataFrame | None, ht_sheet: str, power_sheet: str) -> pd.DataFrame:
    study, _, _ = _compute_ht_meet_frames(source_df, ht_sheet=ht_sheet, power_sheet=power_sheet)
    return study


def _build_temp_meet_sheet(
    source_df: pd.DataFrame | None,
    week_label: str | None = None,
    margin_minutes: int = 60,
) -> pd.DataFrame:
    _, meet, _ = _compute_ht_meet_frames(source_df, week_label=week_label)
    return meet


def _compute_ht_meet_frames(
    source_df: pd.DataFrame | None,
    week_label: str | None = None,
    ht_sheet: str = "HT",
    power_sheet: str = "Power",
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if source_df is None or source_df.empty or "alarm_category" not in source_df.columns:
        return _empty_frame(TEMP_STUDY_COLUMNS), _empty_frame(TEMP_MEET_COLUMNS), pd.DataFrame()
    sub = source_df.copy()
    sub["occurred_on"] = pd.to_datetime(sub.get("occurred_on"), errors="coerce", format="mixed")
    sub["cleared_on"] = pd.to_datetime(sub.get("cleared_on"), errors="coerce", format="mixed")
    sub = sub.dropna(subset=["occurred_on"])
    if week_label:
        sub = _filter_source_to_week(sub, week_label)
    if sub is None or sub.empty:
        return _empty_frame(TEMP_STUDY_COLUMNS), _empty_frame(TEMP_MEET_COLUMNS), pd.DataFrame()
    if "site_id" not in sub.columns:
        sub["site_id"] = _metadata_series(sub, "site_code")
    sub["_site_key"] = sub["site_id"].fillna("").astype(str).map(_normalize_site_identifier)
    sub["_day_key"] = sub["occurred_on"].dt.normalize()
    sub["_duration_secs"] = _duration_seconds_for_rows(sub)

    temp = sub[sub["alarm_category"] == "Temp"].copy().reset_index(drop=True)
    if temp.empty:
        return _empty_frame(TEMP_STUDY_COLUMNS), _empty_frame(TEMP_MEET_COLUMNS), pd.DataFrame()
    power = sub[sub["alarm_category"] == "Power"].copy()
    temp_totals = temp.groupby(["_site_key", "_day_key"], dropna=False)["_duration_secs"].sum().to_dict()
    power_totals = power.groupby(["_site_key", "_day_key"], dropna=False)["_duration_secs"].sum().to_dict() if not power.empty else {}
    temp["_ht_daily_secs"] = [float(temp_totals.get((row["_site_key"], row["_day_key"]), 0.0) or 0.0) for _, row in temp.iterrows()]
    power_values = []
    meet_values = []
    for _, row in temp.iterrows():
        key = (row["_site_key"], row["_day_key"])
        if key not in power_totals:
            power_values.append(math.nan)
            meet_values.append(True)
            continue
        power_secs = float(power_totals.get(key, 0.0) or 0.0)
        power_values.append(power_secs)
        meet_values.append((float(row["_ht_daily_secs"]) - power_secs) > HT_MEET_THRESHOLD_SECONDS)
    temp["_power_daily_secs"] = power_values
    temp["_meet"] = meet_values
    temp["_week_label"] = temp["occurred_on"].apply(_week_label_from_timestamp)

    study = pd.DataFrame({
        "Alarm Source": _metadata_series(temp, "alarm_source"),
        "Site Name": _metadata_series(temp, "site_name", fallback_col="site_id"),
        "Support": "",
        "Day": "",
        "Last Occurred On": temp["occurred_on"].apply(_fmt_dt),
        "Cleared On": temp["cleared_on"].apply(_fmt_dt),
        "Duration(hh:mm:ss)": _duration_series(temp),
        "Alarm Name": _metadata_series(temp, "alarm_name"),
        "Clearance Status": _metadata_series(temp, "clearance_status"),
        "Cleared By": _metadata_from_aliases_or_default(temp, ["cleared_by", "Cleared By"], "EMSReport"),
        "Alarm Reporting Type": _metadata_from_aliases_or_default(temp, ["alarm_reporting_type", "Alarm Reporting Type"], "Real Time"),
        "HT SUM IFS": "",
        "Powr SUM IFS": "",
        "Diff": "",
        "Meet": ["Yes" if value else "" for value in temp["_meet"]],
    })
    for idx in range(len(study)):
        row_number = idx + 2
        study.at[idx, "Support"] = f'=B{row_number}&" "&D{row_number}'
        study.at[idx, "Day"] = f"=DAY(E{row_number})"
        study.at[idx, "HT SUM IFS"] = f"=SUMIFS(G:G,B:B,B{row_number},D:D,D{row_number})"
        study.at[idx, "Powr SUM IFS"] = f"=SUMIFS('{power_sheet}'!$G:$G,'{power_sheet}'!$B:$B,$B{row_number},'{power_sheet}'!$D:$D,$D{row_number})"
        study.at[idx, "Diff"] = f"=L{row_number}-M{row_number}"

    meet_source = temp[temp["_meet"]].copy().reset_index(drop=True)
    meet = _meet_sheet_from_temp_rows(meet_source)
    return study[TEMP_STUDY_COLUMNS], meet[TEMP_MEET_COLUMNS], meet_source


def _duration_seconds_for_rows(df: pd.DataFrame) -> pd.Series:
    if df.empty:
        return pd.Series(dtype=float)
    if "duration" in df.columns:
        duration = pd.to_timedelta(df["duration"], errors="coerce").dt.total_seconds()
    else:
        duration = pd.Series(float("nan"), index=df.index)
    fallback = (df["cleared_on"] - df["occurred_on"]).dt.total_seconds()
    seconds = duration.where(duration > 0, fallback.where(fallback >= 0))
    return seconds.fillna(0.0)


def _meet_sheet_from_temp_rows(temp: pd.DataFrame) -> pd.DataFrame:
    if temp.empty:
        return _empty_frame(TEMP_MEET_COLUMNS)
    rows = pd.DataFrame({
        "Site Name": _metadata_series(temp, "site_name", fallback_col="site_id"),
        "Alarm Source": _metadata_series(temp, "alarm_source"),
        "Last Occurred On": temp["occurred_on"].apply(_fmt_dt),
        "Cleared On": temp["cleared_on"].apply(_fmt_dt),
        "Duration(hh:mm:ss)": _temp_duration_series(temp),
        "Alarm Name": _metadata_series(temp, "alarm_name"),
        "Clearance Status": _metadata_series(temp, "clearance_status"),
        "Cleared By": _metadata_from_aliases_or_default(temp, ["cleared_by", "Cleared By"], "EMSReport"),
        "Alarm Reporting Type": _metadata_from_aliases_or_default(temp, ["alarm_reporting_type", "Alarm Reporting Type"], "Real Time"),
    })
    return rows[TEMP_MEET_COLUMNS]


def _covered_temp_rows(source_df: pd.DataFrame, margin_minutes: int = 60) -> tuple[pd.DataFrame, pd.DataFrame]:
    margin_minutes = _normalize_margin_minutes(margin_minutes)
    sub = source_df.copy()
    sub["occurred_on"] = pd.to_datetime(sub.get("occurred_on"), errors="coerce", format="mixed")
    sub["cleared_on"] = pd.to_datetime(sub.get("cleared_on"), errors="coerce", format="mixed")
    sub = sub.dropna(subset=["site_id", "occurred_on"])
    sub["site_id"] = sub["site_id"].astype(str).str.strip()
    power = sub[sub["alarm_category"] == "Power"].dropna(subset=["occurred_on", "cleared_on"]).copy()
    power = power[power["cleared_on"] >= power["occurred_on"]].sort_values(["site_id", "occurred_on"])
    temp = sub[sub["alarm_category"] == "Temp"].sort_values(["site_id", "occurred_on"]).copy()
    power_by_site = {
        site_id: group.reset_index(drop=True)
        for site_id, group in power.groupby("site_id", sort=False)
    }
    matched_temp_indexes: list[int] = []
    matched_power_rows: list[pd.Series] = []
    for site_id, temps in temp.groupby("site_id", sort=False):
        site_power = power_by_site.get(site_id)
        if site_power is None or site_power.empty:
            continue
        starts = site_power["occurred_on"].reset_index(drop=True)
        coverage_end = site_power["cleared_on"].cummax().reset_index(drop=True)
        if margin_minutes:
            margin = _safe_margin_delta(margin_minutes)
            if margin is None:
                coverage_end = pd.Series(pd.Timestamp.max, index=coverage_end.index)
            else:
                safe_to_add_margin = coverage_end <= pd.Timestamp.max - margin
                coverage_end = coverage_end.copy()
                coverage_end.loc[safe_to_add_margin] = coverage_end.loc[safe_to_add_margin] + margin
                if (~safe_to_add_margin).any():
                    coverage_end.loc[~safe_to_add_margin] = pd.Timestamp.max
        for temp_idx, temp_row in temps.iterrows():
            pos = starts.searchsorted(temp_row["occurred_on"], side="right") - 1
            if pos < 0:
                continue
            if coverage_end.iloc[pos] >= temp_row["occurred_on"]:
                matched_temp_indexes.append(temp_idx)
                matched_power_rows.append(site_power.iloc[pos])
    if not matched_temp_indexes:
        return temp.iloc[0:0].copy(), power.iloc[0:0].copy()
    covered = temp.loc[matched_temp_indexes].copy()
    matched_power = pd.DataFrame(matched_power_rows, index=matched_temp_indexes)
    return covered, matched_power


def _summary_time_values(data: pd.DataFrame) -> pd.Series:
    if "occurred_on" in data.columns:
        return pd.to_datetime(data["occurred_on"], errors="coerce", format="mixed")
    if "Last Occurred On" in data.columns:
        return pd.to_datetime(data["Last Occurred On"], errors="coerce", format="mixed")
    if "temp_time" in data.columns:
        return pd.to_datetime(data["temp_time"], errors="coerce", format="mixed")
    return pd.Series(pd.NaT, index=data.index)


def _summary_site_key(data: pd.DataFrame) -> pd.Series:
    for column in ("site_id", "site_code", "Site Code", "Site Name"):
        if column in data.columns:
            return data[column].fillna("").astype(str).str.strip()
    return pd.Series("", index=data.index)


def _summary_duration_seconds(data: pd.DataFrame) -> float:
    if "duration" in data.columns and "occurred_on" in data.columns:
        return float(_duration_seconds_for_rows(data).sum())
    if "Duration(hh:mm:ss)" in data.columns:
        return float(sum(duration_to_secs(value) for value in data["Duration(hh:mm:ss)"]))
    if "temp_clear_duration" in data.columns:
        return float(sum(duration_to_secs(value) for value in data["temp_clear_duration"]))
    return 0.0


def _week_label_from_summary_rows(data: pd.DataFrame) -> str:
    if data.empty:
        return ""
    values = _summary_time_values(data).dropna()
    if values.empty:
        return ""
    return _week_label_from_timestamp(values.max())


def _week_label_from_source(source_df: pd.DataFrame | None) -> str:
    if source_df is None or source_df.empty or "occurred_on" not in source_df.columns:
        return ""
    values = pd.to_datetime(source_df["occurred_on"], errors="coerce", format="mixed").dropna()
    if values.empty:
        return ""
    return _week_label_from_timestamp(values.max())


def _duration_series(df: pd.DataFrame) -> list[float | str]:
    if df.empty:
        return []
    if "duration" in df.columns:
        duration = pd.to_timedelta(df["duration"], errors="coerce")
        seconds = duration.dt.total_seconds()
    else:
        seconds = pd.Series(float("nan"), index=df.index)
    fallback = (df["cleared_on"] - df["occurred_on"]).dt.total_seconds()
    seconds = seconds.where(seconds > 0, fallback.where(fallback >= 0))
    return [float(value) / 86400 if pd.notna(value) and value > 0 else "" for value in seconds]


def _fmt_dt(value) -> str:
    if pd.isna(value):
        return ""
    timestamp = pd.Timestamp(value)
    return f"{timestamp.month}/{timestamp.day}/{str(timestamp.year)[-2:]} {timestamp:%H:%M}"


def _temp_duration(temp: pd.Series, temp_time, temp_cleared) -> str:
    seconds = duration_to_secs(temp.get("duration"))
    if seconds > 0:
        return fmt_td(pd.Timedelta(seconds=seconds))
    if pd.notna(temp_time) and pd.notna(temp_cleared) and temp_cleared >= temp_time:
        return fmt_td(temp_cleared - temp_time)
    return ""


def _temp_duration_series(temps: pd.DataFrame) -> list[str]:
    if "duration" in temps.columns:
        duration = pd.to_timedelta(temps["duration"], errors="coerce")
        seconds = duration.dt.total_seconds()
    else:
        seconds = pd.Series(float("nan"), index=temps.index)
    fallback = (temps["cleared_on"] - temps["occurred_on"]).dt.total_seconds()
    seconds = seconds.where(seconds > 0, fallback.where(fallback >= 0))
    return [_fmt_seconds_to_hhmmss(value) for value in seconds]


def _fmt_seconds_to_hhmmss(seconds: float) -> str:
    if pd.isna(seconds) or seconds <= 0:
        return ""
    total = int(seconds)
    hours, remainder = divmod(total, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def _fmt_hours_minutes(seconds: float) -> str:
    total_minutes = int(seconds // 60)
    hours, minutes = divmod(total_minutes, 60)
    return f"{hours:02d}:{minutes:02d}"


def _week_history_columns(week_label: str, count: int = 8) -> list[str]:
    parsed = _parse_week_label(week_label)
    if parsed is None:
        return [week_label]
    try:
        current, _ = ht_export_week_range(week_label)
    except ValueError:
        return [week_label]
    labels = []
    for offset in range(count):
        labels.append(_week_label_from_timestamp(current - pd.Timedelta(weeks=offset)))
    return labels


def _parse_week_label(week_label: str) -> tuple[int, int] | None:
    text = str(week_label or "").strip().upper()
    if not text.startswith("W") or "-" not in text:
        return None
    week_text, year_text = text[1:].split("-", 1)
    try:
        week = int(week_text)
        year = int(year_text)
    except ValueError:
        return None
    year += 2000 if year < 100 else 0
    if week < 1 or week > 54:
        return None
    return year, week


def _week_label_from_matches(matches: pd.DataFrame) -> str:
    if matches.empty or "temp_time" not in matches.columns:
        return ""
    first = pd.to_datetime(matches["temp_time"], errors="coerce").dropna()
    if first.empty:
        return ""
    return _week_label_from_timestamp(first.max())


def _week_label_from_timestamp(value) -> str:
    if pd.isna(value):
        return ""
    timestamp = pd.Timestamp(value)
    week = _excel_week_number(timestamp)
    if week == "":
        return ""
    return f"W{int(week):02d}-{str(int(timestamp.year))[-2:]}"


def _excel_week_number(value) -> int | str:
    if pd.isna(value):
        return ""
    timestamp = pd.Timestamp(value)
    return int(timestamp.strftime("%U")) + 1


def _first_text(df: pd.DataFrame, column: str) -> str:
    if column not in df.columns:
        return ""
    values = df[column].dropna().astype(str).str.strip()
    values = values[values != ""]
    return values.iloc[0] if not values.empty else ""
