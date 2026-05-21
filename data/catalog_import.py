"""Import Network Summary and BDT Summary workbooks into catalog stores."""

from __future__ import annotations

import hashlib
import json
import logging
import re
from pathlib import Path
from typing import Any

import pandas as pd

_log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _normalize_site_id(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    if not text:
        return ""
    return "".join(ch for ch in text if ch.isalnum())


def _normalize_header(name: str) -> str:
    """Strip, lowercase, replace runs of non-alnum with underscore."""
    replaced = re.sub(r"[^a-z0-9]+", "_", name.strip().lower())
    return replaced.strip("_")


def _build_header_map(raw_headers: list[str]) -> dict[str, str]:
    """Return {original_header: normalized_header} for every raw header."""
    mapping: dict[str, str] = {}
    for h in raw_headers:
        key = str(h).strip()
        if key:
            mapping[key] = _normalize_header(key)
    return mapping


def _row_to_dict(
    row: pd.Series,
    headers: list[str],
) -> dict[str, Any]:
    """Convert a row (value per original header) to a plain dict."""
    result: dict[str, Any] = {}
    for i, h in enumerate(headers):
        val = row.iloc[i] if i < len(row) else None
        if isinstance(val, float) and pd.isna(val):
            val = None
        result[h] = val
    return result


def _normalized_row_data(row_dict: dict[str, Any], header_map: dict[str, str]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for original, normalized_key in header_map.items():
        normalized[normalized_key] = row_dict.get(original)
    return normalized


def _compute_content_hash(raw_data: dict[str, Any]) -> str:
    """Deterministic hash of row data for dedup."""
    canonical = json.dumps(raw_data, sort_keys=True, default=str)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# Network Summary (Site Metadata)
# ---------------------------------------------------------------------------


def import_network_summary_db_sheet(workbook_path: str | Path) -> int:
    """Import the ``DB`` sheet from a Network Summary workbook.

    1. Open *workbook_path* with openpyxl, locate ``DB`` sheet.
    2. Validate ``Code`` column exists and has non-empty rows.
    3. Preserve every column; normalize headers.
    4. Replace SQLite and DuckDB site_metadata_catalog all-or-nothing.

    Returns number of sites imported.
    Raises ``ValueError`` when ``Code`` is missing or no valid rows.
    Raises any underlying DB errors (DuckDB or SQLite) as-is so callers
    can detect partial-failure.
    """
    import openpyxl

    path = Path(workbook_path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows_list = _read_network_summary_rows(wb, path)
    finally:
        wb.close()

    # -- write SQLite --
    try:
        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.repos.catalog_repo import replace_all_site_metadata
    except ImportError:
        from db.engine import get_shared_session
        from db.repos.catalog_repo import replace_all_site_metadata

    # -- write DuckDB --
    try:
        from alarm_app.data.catalog_store import read_site_metadata, replace_site_metadata
    except ImportError:
        from data.catalog_store import read_site_metadata, replace_site_metadata

    session = get_shared_session()
    duckdb_snapshot = read_site_metadata()
    duckdb_written = False
    try:
        count_sql = replace_all_site_metadata(session, rows_list)
        df = pd.DataFrame(rows_list)
        replace_site_metadata(df)
        duckdb_written = True
        session.commit()
    except Exception:
        session.rollback()
        if duckdb_written:
            _restore_duckdb_snapshot(replace_site_metadata, duckdb_snapshot, "site metadata")
        raise
    finally:
        session.close()

    _log.info("Network summary import complete: %d sites", count_sql)
    return count_sql


def _read_network_summary_rows(wb, path: Path) -> list[dict[str, Any]]:

    # locate DB sheet
    sheet = None
    for candidate in ["DB", "db", "Db"]:
        if candidate in wb.sheetnames:
            sheet = wb[candidate]
            break
    if sheet is None:
        raise ValueError(
            f"No 'DB' sheet found in {path.name}. "
            f"Available sheets: {wb.sheetnames}"
        )

    # read all rows (read_only mode gives cells, not row generators, so
    # we iterate explicitly)
    all_rows: list[list[Any]] = []
    for row_cells in sheet.iter_rows(values_only=True):
        all_rows.append(list(row_cells))

    if len(all_rows) < 2:
        raise ValueError("DB sheet is empty or has no header row")

    # header row
    raw_headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    # find index of "Code"
    code_indices = [i for i, h in enumerate(raw_headers) if h.upper() == "CODE"]
    if not code_indices:
        raise ValueError(
            f"DB sheet missing 'Code' column. Columns found: {raw_headers}"
        )
    code_idx = code_indices[0]

    header_map = _build_header_map(raw_headers)

    # data rows
    rows_list: list[dict[str, Any]] = []
    for raw_row in all_rows[1:]:
        # extract Code value
        code_val = raw_row[code_idx] if code_idx < len(raw_row) else None
        site_id = _normalize_site_id(code_val)
        if not site_id:
            continue  # skip blank Code rows

        original_data = _row_to_dict(pd.Series(raw_row, index=raw_headers), raw_headers)
        normalized_data = _normalized_row_data(original_data, header_map)
        normalized_data["site_id"] = site_id
        rows_list.append({
            **normalized_data,
            "original_headers_json": json.dumps(header_map, ensure_ascii=False),
            "raw_data_json": json.dumps(normalized_data, default=str, ensure_ascii=False),
        })

    if not rows_list:
        raise ValueError("No rows with non-empty Code found in DB sheet")
    return rows_list


# ---------------------------------------------------------------------------
# BDT Summary
# ---------------------------------------------------------------------------


def _extract_week(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _extract_test_date(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return None
        return ts.strftime("%Y-%m-%d")
    except Exception:
        return None


def _extract_test_year(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int) and not isinstance(value, bool):
        return value if 1000 <= value <= 9999 else None
    if isinstance(value, float) and value.is_integer():
        year = int(value)
        return year if 1000 <= year <= 9999 else None
    try:
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            # try bare int
            try:
                year = int(float(str(value)))
            except (ValueError, TypeError, OverflowError):
                return None
            return year if 1000 <= year <= 9999 else None
        return ts.year
    except (ValueError, TypeError, OverflowError):
        return None


def _restore_duckdb_snapshot(restore_func, snapshot: pd.DataFrame, label: str) -> None:
    try:
        restore_func(snapshot)
    except Exception:
        _log.exception("Failed to restore DuckDB %s snapshot after import rollback", label)


def import_bdt_summary_workbook(workbook_path: str | Path) -> dict[str, int]:
    """Import all sheets from a BDT Summary Workbook.

    Every sheet is treated as BDT summary data. Sheet names are
    preserved as *reporting_period* keys.  Rows are merged by reporting
    period: periods present in the workbook replace existing rows;
    periods not present are left untouched.

    Returns ``{reporting_period: row_count}`` dict.
    """
    import openpyxl

    path = Path(workbook_path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        all_rows, period_counts = _read_bdt_summary_rows(wb, path)
    finally:
        wb.close()

    if not all_rows and not period_counts:
        _log.warning("BDT import: no rows found in any sheet of %s", path.name)
        return period_counts

    # -- write SQLite: merge per period --
    try:
        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.repos.catalog_repo import merge_bdt_period
    except ImportError:
        from db.engine import get_shared_session
        from db.repos.catalog_repo import merge_bdt_period

    reporting_periods = list(period_counts.keys())
    # -- write DuckDB --
    try:
        from alarm_app.data.catalog_store import merge_bdt_summary, read_bdt_summary, replace_bdt_summary
    except ImportError:
        from data.catalog_store import merge_bdt_summary, read_bdt_summary, replace_bdt_summary

    session = get_shared_session()
    duckdb_snapshot = read_bdt_summary()
    duckdb_written = False
    try:
        for period in reporting_periods:
            period_rows_list = [r for r in all_rows if r["reporting_period"] == period]
            merge_bdt_period(session, period, period_rows_list)
        df = pd.DataFrame(all_rows)
        merge_bdt_summary(df, reporting_periods)
        duckdb_written = True
        session.commit()
    except Exception:
        session.rollback()
        if duckdb_written:
            _restore_duckdb_snapshot(replace_bdt_summary, duckdb_snapshot, "BDT summary")
        raise
    finally:
        session.close()

    _log.info("BDT summary import complete: periods=%s", period_counts)
    return period_counts


def _read_bdt_summary_rows(wb, path: Path) -> tuple[list[dict[str, Any]], dict[str, int]]:

    all_rows: list[dict[str, Any]] = []
    period_counts: dict[str, int] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows: list[list[Any]] = []
        for row_cells in ws.iter_rows(values_only=True):
            raw_rows.append(list(row_cells))

        if len(raw_rows) < 2:
            period_counts[sheet_name] = 0
            continue  # skip empty or header-only sheets

        raw_headers = [str(h).strip() if h is not None else "" for h in raw_rows[0]]
        header_map = _build_header_map(raw_headers)

        # locate possible key columns
        site_col_idx: int | None = None
        week_col_idx: int | None = None
        date_col_idx: int | None = None
        year_col_idx: int | None = None

        for i, h in enumerate(raw_headers):
            hl = h.lower()
            if hl in ("site id", "site code", "site", "code", "site_id", "site_code"):
                if site_col_idx is None:
                    site_col_idx = i
            if hl in ("week", "week number", "week_no"):
                if week_col_idx is None:
                    week_col_idx = i
            if hl in ("test date", "test_date", "date", "discharge date"):
                if date_col_idx is None:
                    date_col_idx = i
            if hl in ("test year", "test_year", "year"):
                if year_col_idx is None:
                    year_col_idx = i

        period_rows = 0
        for raw_row in raw_rows[1:]:
            # site_id
            site_val = (
                raw_row[site_col_idx]
                if site_col_idx is not None and site_col_idx < len(raw_row)
                else None
            )
            site_id = _normalize_site_id(site_val)
            if not site_id:
                continue  # skip rows without site

            week_val = (
                raw_row[week_col_idx]
                if week_col_idx is not None and week_col_idx < len(raw_row)
                else None
            )
            date_val = (
                raw_row[date_col_idx]
                if date_col_idx is not None and date_col_idx < len(raw_row)
                else None
            )
            year_val = (
                raw_row[year_col_idx]
                if year_col_idx is not None and year_col_idx < len(raw_row)
                else None
            )

            week = _extract_week(week_val)
            test_date = _extract_test_date(date_val)
            test_year = _extract_test_year(year_val)

            original_data = _row_to_dict(pd.Series(raw_row, index=raw_headers), raw_headers)
            raw_data = _normalized_row_data(original_data, header_map)
            # add extracted fields
            raw_data["site_id"] = site_id
            raw_data["reporting_period"] = sheet_name
            raw_data["week"] = week
            raw_data["test_date"] = test_date
            raw_data["test_year"] = test_year

            content_hash = _compute_content_hash(raw_data)

            all_rows.append(
                {
                    **raw_data,
                    "site_id": site_id,
                    "reporting_period": sheet_name,
                    "week": week,
                    "test_date": test_date,
                    "test_year": test_year,
                    "content_hash": content_hash,
                    "original_headers_json": json.dumps(
                        header_map, ensure_ascii=False
                    ),
                    "raw_data_json": json.dumps(
                        raw_data, default=str, ensure_ascii=False
                    ),
                }
            )
            period_rows += 1

        period_counts[sheet_name] = period_rows
    return all_rows, period_counts
