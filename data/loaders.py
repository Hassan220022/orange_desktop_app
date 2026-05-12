"""
File discovery, CSV/XLSX parsing, and alarm-data normalisation.

Pure data-layer module with no Qt dependencies.

Optimisations:
 - Encoding detection tries utf-8-sig first (covers BOM and plain UTF-8)
   before falling back to latin-1.  cp1252 is skipped because latin-1 is
   a strict superset for all byte values.
 - Column renaming uses a single dict-comprehension instead of two passes.
 - datetime conversion uses format= hint when possible.
"""

import os

import numpy as np
import pandas as pd

try:
    from alarm_app.constants import ALL_INTERNAL_COLS, SCHEMA_1_MAP, SCHEMA_2_MAP
except ImportError:
    from constants import ALL_INTERNAL_COLS, SCHEMA_1_MAP, SCHEMA_2_MAP

_EXTS = frozenset((".csv", ".xlsx", ".xls"))
_ENCODINGS = ("utf-8-sig", "latin-1")          # utf-8-sig covers plain utf-8 too


# ─────────────────────────────────────────────────────────────────
# Header validation (shared by discovery + parsing)
# ─────────────────────────────────────────────────────────────────
_SCHEMA_KEYS_1 = frozenset(SCHEMA_1_MAP.keys())
_SCHEMA_KEYS_2 = frozenset(SCHEMA_2_MAP.keys())
_MIN_MATCH = 3  # minimum columns that must match to qualify as alarm data


def _is_alarm_header(columns: list[str]) -> bool:
    """Quick check whether column headers look like alarm data."""
    col_set = {str(c).strip() for c in columns}
    return (len(col_set & _SCHEMA_KEYS_1) >= _MIN_MATCH
            or len(col_set & _SCHEMA_KEYS_2) >= _MIN_MATCH)


def _quick_header_check(path: str, ext: str) -> bool:
    """Read only the header row and verify it matches an alarm schema.

    Uses calamine (fast Rust reader) for xlsx/xls, falling back to
    openpyxl if calamine is unavailable or fails.  Returns False for
    files that aren't alarm data (BDT files, random spreadsheets, etc.).
    """
    try:
        if ext == ".csv":
            for enc in _ENCODINGS:
                try:
                    df = pd.read_csv(path, encoding=enc, nrows=0)
                    return _is_alarm_header(df.columns.tolist())
                except Exception:
                    continue
            return False
        # xlsx / xls — try calamine first (Rust-based, ~3x faster)
        try:
            df = pd.read_excel(path, engine="calamine", nrows=0)
            return _is_alarm_header(df.columns.tolist())
        except Exception:
            pass
        # Fallback to openpyxl read_only for header-only access
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb.active
            if ws is None:
                return False
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            return _is_alarm_header(headers)
        finally:
            wb.close()
    except Exception:
        return False


# Filename patterns that strongly indicate alarm data — skip header check
_ALARM_NAME_HINTS = ("alarm", "power_alarm", "down_alarm")


# External weekly-summary workbook support (R11 checklist matching).
_SUMMARY_FILE_EXTS = frozenset((".xlsx", ".xlsm", ".xls", ".xlsb", ".ods"))
_SUMMARY_CANONICAL_KEYS = (
    "Short Code",
    "PLVD Value",
    "Rectifier Brand",
    "# of Modules",
    "Battery Brand",
    "Battery Volt",
    "No of String",
    "No of Batteries",
    "Start Volt",
    "Start Amp",
    "End Volt",
    "End Amp",
    "Discharge time( Mins)",
    "Test Date",
)
_SUMMARY_KEY_ALIASES = {
    "Short Code": ("Short Code", "Site Code"),
    "PLVD Value": ("PLVD Value", "PLD Value"),
    "Rectifier Brand": ("Rectifier Brand",),
    "# of Modules": ("# of Modules", "Number of Modules", "No of Modules"),
    "Battery Brand": ("Battery Brand",),
    "Battery Volt": ("Battery Volt", "Battery Voltage"),
    "No of String": ("No of String", "No of Strings", "Number of Strings"),
    "No of Batteries": ("No of Batteries", "No of Batteries ", "Number of Batteries"),
    "Start Volt": ("Start Volt", "Start Voltage"),
    "Start Amp": ("Start Amp",),
    "End Volt": ("End Volt", "End Voltage"),
    "End Amp": ("End Amp",),
    "Discharge time( Mins)": ("Discharge time( Mins)", "Discharge Time (mins)", "Discharge Time (min)"),
    "Test Date": ("Test Date",),
}


def _normalize_summary_key(key) -> str:
    return "".join(ch for ch in str(key or "").strip().lower() if ch.isalnum())


_SUMMARY_ALIAS_TO_CANONICAL = {
    _normalize_summary_key(alias): canonical
    for canonical, aliases in _SUMMARY_KEY_ALIASES.items()
    for alias in aliases
}


def _summary_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            return str(int(value))
    if isinstance(value, np.floating):
        if np.isnan(value):
            return ""
        if float(value).is_integer():
            return str(int(float(value)))
    if isinstance(value, int | np.integer):
        return str(int(value))
    s = str(value).strip()
    if s.lower() in {"nan", "none", "null", "na", "n/a", "-", "--"}:
        return ""
    return s


def _summary_date_key(value) -> str:
    if value is None:
        return ""
    try:
        if isinstance(value, str):
            raw = value.strip()
            if not raw:
                return ""
            ts = pd.to_datetime(raw, errors="coerce", format="%Y-%m-%d")
            if pd.isna(ts):
                ts = pd.to_datetime(raw, errors="coerce", dayfirst=True)
            if pd.isna(ts):
                ts = pd.to_datetime(raw, errors="coerce")
        else:
            ts = pd.to_datetime(value, errors="coerce")
    except Exception:
        return ""
    if pd.isna(ts):
        return ""
    return pd.Timestamp(ts).normalize().strftime("%Y-%m-%d")


def _summary_site_key(value) -> str:
    return "".join(ch for ch in _summary_text(value).upper() if ch.isalnum())


def _summary_candidate_files(bdt_files: list[str]) -> list[str]:
    bdt_paths = {os.path.normcase(os.path.abspath(p)) for p in bdt_files}
    scan_dirs = {os.path.dirname(os.path.abspath(p)) for p in bdt_files}
    candidates: set[str] = set()

    for directory in scan_dirs:
        try:
            names = os.listdir(directory)
        except OSError:
            continue
        for name in names:
            if name.startswith("~$") or name.startswith("._"):
                continue
            ext = os.path.splitext(name)[1].lower()
            if ext not in _SUMMARY_FILE_EXTS:
                continue
            full = os.path.abspath(os.path.join(directory, name))
            if os.path.normcase(full) in bdt_paths:
                continue
            if "bdt" in name.lower():
                continue
            candidates.add(full)

    def _sort_key(path: str):
        name = os.path.basename(path).lower()
        score = 0
        if "summary" in name:
            score += 8
        if "weekly" in name:
            score += 5
        if "battery" in name:
            score += 4
        if "update" in name:
            score += 3
        return (-score, name)

    return sorted(candidates, key=_sort_key)


def _extract_summary_rows(file_path: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    workbook = None

    for engine in ("calamine", "openpyxl"):
        try:
            workbook = pd.ExcelFile(file_path, engine=engine)
            break
        except Exception:
            continue

    if workbook is None:
        return rows

    try:
        for sheet_name in workbook.sheet_names:
            try:
                df = pd.read_excel(workbook, sheet_name=sheet_name, dtype=object)
            except Exception:
                continue
            if df is None or df.empty:
                continue

            mapped_cols: dict[str, str] = {}
            for col in df.columns:
                canonical = _SUMMARY_ALIAS_TO_CANONICAL.get(_normalize_summary_key(col))
                if canonical and canonical not in mapped_cols.values():
                    mapped_cols[col] = canonical

            if "Short Code" not in mapped_cols.values():
                continue
            if "Test Date" not in mapped_cols.values():
                continue

            for _, row in df.iterrows():
                summary_row: dict[str, str] = {}
                for source_col, canonical in mapped_cols.items():
                    raw = row.get(source_col, None)
                    if canonical == "Test Date":
                        summary_row[canonical] = _summary_date_key(raw)
                    elif canonical == "Short Code":
                        summary_row[canonical] = _summary_text(raw).upper()
                    else:
                        summary_row[canonical] = _summary_text(raw)

                if not summary_row.get("Short Code"):
                    continue
                if not any(summary_row.get(k, "") for k in _SUMMARY_CANONICAL_KEYS
                           if k not in {"Short Code", "Test Date"}):
                    continue
                rows.append(summary_row)
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    return rows


def _load_external_summary_lookup(bdt_files: list[str]) -> dict[str, dict]:
    """Build summary lookup maps keyed by (site, test_date) and by site."""
    lookup: dict[str, dict] = {"by_site_date": {}, "by_site": {}}
    for path in _summary_candidate_files(bdt_files):
        for row in _extract_summary_rows(path):
            site_key = _summary_site_key(row.get("Short Code"))
            if not site_key:
                continue
            date_key = _summary_date_key(row.get("Test Date"))
            row_copy = dict(row)
            if date_key:
                row_copy["Test Date"] = date_key
                lookup["by_site_date"].setdefault((site_key, date_key), row_copy)

            by_site = lookup["by_site"].setdefault(site_key, {})
            by_site.setdefault(date_key, row_copy)
    return lookup


def _match_external_summary_row(bdt_data, summary_lookup: dict[str, dict]) -> dict[str, str] | None:
    if not summary_lookup:
        return None

    site_key = _summary_site_key(getattr(bdt_data, "site_code", ""))
    if not site_key:
        return None

    by_site_date = summary_lookup.get("by_site_date", {})
    by_site = summary_lookup.get("by_site", {})
    test_date_key = _summary_date_key(getattr(bdt_data, "test_date", None))

    if test_date_key:
        exact = by_site_date.get((site_key, test_date_key))
        if exact:
            return dict(exact)

    site_rows = by_site.get(site_key, {})
    if not site_rows:
        return None
    if len(site_rows) == 1:
        return dict(next(iter(site_rows.values())))
    if "" in site_rows and len(site_rows) == 2 and test_date_key:
        return dict(site_rows[""])
    return None


def discover_alarm_files(directory: str) -> list[dict]:
    """
    Recursively walk *directory* and return metadata dicts for every
    alarm-format .csv / .xlsx / .xls file found.

    Performs a fast header check on each file so only genuine alarm
    data appears in the list. BDT files, random spreadsheets, etc.
    are silently skipped.
    """
    results: list[dict] = []
    for root, _dirs, files in os.walk(directory):
        for fname in sorted(files):
            if fname.startswith("._") or fname.startswith("~$"):
                continue
            ext = os.path.splitext(fname)[1].lower()
            if ext not in _EXTS:
                continue
            # Fast skip: BDT files by name
            fl = fname.lower()
            if "bdt" in fl:
                continue
            full = os.path.join(root, fname)
            # Trust filenames that clearly indicate alarm data
            name_ok = any(hint in fl for hint in _ALARM_NAME_HINTS)
            if not name_ok:
                # Unknown file — verify header before listing
                if not _quick_header_check(full, ext):
                    continue
            rel  = os.path.relpath(full, directory)
            kb   = os.path.getsize(full) / 1024
            results.append({
                "path": full, "rel_path": rel,
                "filename": fname, "ext": ext,
                "size_kb": kb,
            })
    return results


def parse_alarm_file(info: dict) -> pd.DataFrame | None:
    """Parse one alarm file described by a discovery dict.  Returns None on failure."""
    fp, ext, fname = info["path"], info["ext"], info["filename"]
    df = None

    for enc in _ENCODINGS:
        try:
            if ext == ".csv":
                df = pd.read_csv(fp, encoding=enc,
                                 on_bad_lines="skip", low_memory=False)
            else:
                # calamine (Rust-based) is ~3x faster than openpyxl for read-only
                try:
                    df = pd.read_excel(fp, engine="calamine")
                except Exception:
                    df = pd.read_excel(fp, engine="openpyxl")
            if df is not None and not df.empty:
                break
        except Exception:
            continue

    if df is None or df.empty:
        return None

    # Normalise headers
    df.columns = [str(c).strip() for c in df.columns]

    # Early reject: skip files that don't match any alarm schema
    if not _is_alarm_header(df.columns.tolist()):
        return None

    cols = set(df.columns)

    # Choose schema
    if "FM Office" in cols or ("Site ID" in cols and "Site Name" not in cols):
        rmap = {k: v for k, v in SCHEMA_2_MAP.items() if k in cols}
    else:
        rmap = {k: v for k, v in SCHEMA_1_MAP.items() if k in cols}

    df = df.rename(columns=rmap)
    fname_lower = fname.lower()
    if "temp" in fname_lower:
        df["alarm_category"] = "Temp"
    elif "power" in fname_lower:
        df["alarm_category"] = "Power"
    elif "down" in fname_lower:
        df["alarm_category"] = "Down"
    elif "door" in fname_lower:
        df["alarm_category"] = "Door"
    else:
        df["alarm_category"] = ""
    df["file_source"]    = fname

    # Ensure every expected column exists
    for col in ALL_INTERNAL_COLS:
        if col not in df.columns:
            df[col] = np.nan

    return df[ALL_INTERNAL_COLS]


# ─────────────────────────────────────────────────────────────────
# Classification helpers — re-exported from core.classify
# Duration helpers — re-exported from core.duration
# ─────────────────────────────────────────────────────────────────

_ROW_HASH_COLUMNS = (
    "site_id",
    "alarm_name",
    "alarm_id",
    "network_type",
    "vendor",
    "occurred_on",
    "cleared_on",
    "duration",
    "clearance_status",
    "alarm_source",
    "alarm_category",
)


def deduplicate_alarm_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Drop duplicate canonical rows using vectorized key columns."""
    if df is None or df.empty:
        return df, 0

    key_parts: dict[str, pd.Series] = {}
    for col in _ROW_HASH_COLUMNS:
        if col not in df.columns:
            key_parts[col] = pd.Series("", index=df.index, dtype="object")
            continue
        values = df[col]
        if pd.api.types.is_datetime64_any_dtype(values):
            key_parts[col] = values.dt.strftime("%Y-%m-%d %H:%M:%S").fillna("")
        else:
            key_parts[col] = values.where(values.notna(), "").astype(str).str.strip().str.lower()

    duplicate_mask = pd.DataFrame(key_parts, index=df.index).duplicated(keep="first")
    dropped = int(duplicate_mask.sum())
    return df.loc[~duplicate_mask].copy(), dropped
