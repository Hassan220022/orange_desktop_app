"""
Parsers — file discovery, CSV/XLSX parsing, and threaded background loader.

Optimisations:
 • Encoding detection tries utf-8-sig first (covers BOM and plain UTF-8)
   before falling back to latin-1.  cp1252 is skipped because latin-1 is
   a strict superset for all byte values.
 • Column renaming uses a single dict-comprehension instead of two passes.
 • datetime conversion uses format= hint when possible.
 • Concatenation uses copy=False to avoid redundant memory copies.
"""

import os
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import numpy as np

from PyQt5.QtCore import QThread, pyqtSignal

try:
    from .constants import SCHEMA_1_MAP, SCHEMA_2_MAP, ALL_INTERNAL_COLS
except ImportError:
    from constants import SCHEMA_1_MAP, SCHEMA_2_MAP, ALL_INTERNAL_COLS

_EXTS = frozenset((".csv", ".xlsx", ".xls"))
_ENCODINGS = ("utf-8-sig", "latin-1")          # utf-8-sig covers plain utf-8 too


# ─────────────────────────────────────────────────────────────────
# File discovery
# ─────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────
# Header validation (shared by discovery + parsing)
# ─────────────────────────────────────────────────────────────────
_SCHEMA_KEYS_1 = frozenset(SCHEMA_1_MAP.keys())
_SCHEMA_KEYS_2 = frozenset(SCHEMA_2_MAP.keys())
_MIN_MATCH = 3  # minimum columns that must match to qualify as alarm data


def _is_alarm_header(columns: list[str]) -> bool:
    """Quick check whether column headers look like alarm data."""
    col_set = {str(c).strip() for c in columns}
    return (len(col_set & _SCHEMA_KEYS_1) >= _MIN_MATCH
            or len(col_set & _SCHEMA_KEYS_2) >= _MIN_MATCH)


def _quick_header_check(path: str, ext: str) -> bool:
    """Read only the header row and verify it matches an alarm schema.

    Uses calamine (fast Rust reader) for xlsx/xls, falling back to
    openpyxl if calamine is unavailable or fails.  Returns False for
    files that aren't alarm data (BDT files, random spreadsheets, etc.).
    """
    try:
        if ext == ".csv":
            for enc in _ENCODINGS:
                try:
                    df = pd.read_csv(path, encoding=enc, nrows=0)
                    return _is_alarm_header(df.columns.tolist())
                except Exception:
                    continue
            return False
        # xlsx / xls — try calamine first (Rust-based, ~3x faster)
        try:
            df = pd.read_excel(path, engine="calamine", nrows=0)
            return _is_alarm_header(df.columns.tolist())
        except Exception:
            pass
        # Fallback to openpyxl read_only for header-only access
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb.active
            if ws is None:
                return False
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            return _is_alarm_header(headers)
        finally:
            wb.close()
    except Exception:
        return False


# Filename patterns that strongly indicate alarm data — skip header check
_ALARM_NAME_HINTS = ("alarm", "power_alarm", "down_alarm")


# External weekly-summary workbook support (R11 checklist matching).
_SUMMARY_FILE_EXTS = frozenset((".xlsx", ".xlsm", ".xls", ".xlsb", ".ods"))
_SUMMARY_CANONICAL_KEYS = (
    "Short Code",
    "PLVD Value",
    "Rectifier Brand",
    "# of Modules",
    "Battery Brand",
    "Battery Volt",
    "No of String",
    "No of Batteries",
    "Start Volt",
    "Start Amp",
    "End Volt",
    "End Amp",
    "Discharge time( Mins)",
    "Test Date",
)
_SUMMARY_KEY_ALIASES = {
    "Short Code": ("Short Code", "Site Code"),
    "PLVD Value": ("PLVD Value", "PLD Value"),
    "Rectifier Brand": ("Rectifier Brand",),
    "# of Modules": ("# of Modules", "Number of Modules", "No of Modules"),
    "Battery Brand": ("Battery Brand",),
    "Battery Volt": ("Battery Volt", "Battery Voltage"),
    "No of String": ("No of String", "No of Strings", "Number of Strings"),
    "No of Batteries": ("No of Batteries", "No of Batteries ", "Number of Batteries"),
    "Start Volt": ("Start Volt", "Start Voltage"),
    "Start Amp": ("Start Amp",),
    "End Volt": ("End Volt", "End Voltage"),
    "End Amp": ("End Amp",),
    "Discharge time( Mins)": ("Discharge time( Mins)", "Discharge Time (mins)", "Discharge Time (min)"),
    "Test Date": ("Test Date",),
}


def _normalize_summary_key(key) -> str:
    return "".join(ch for ch in str(key or "").strip().lower() if ch.isalnum())


_SUMMARY_ALIAS_TO_CANONICAL = {
    _normalize_summary_key(alias): canonical
    for canonical, aliases in _SUMMARY_KEY_ALIASES.items()
    for alias in aliases
}


def _summary_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            return str(int(value))
    if isinstance(value, np.floating):
        if np.isnan(value):
            return ""
        if float(value).is_integer():
            return str(int(float(value)))
    if isinstance(value, int | np.integer):
        return str(int(value))
    s = str(value).strip()
    if s.lower() in {"nan", "none", "null", "na", "n/a", "-", "--"}:
        return ""
    return s


def _summary_date_key(value) -> str:
    if value is None:
        return ""
    try:
        if isinstance(value, str):
            raw = value.strip()
            if not raw:
                return ""
            ts = pd.to_datetime(raw, errors="coerce", format="%Y-%m-%d")
            if pd.isna(ts):
                ts = pd.to_datetime(raw, errors="coerce", dayfirst=True)
            if pd.isna(ts):
                ts = pd.to_datetime(raw, errors="coerce")
        else:
            ts = pd.to_datetime(value, errors="coerce")
    except Exception:
        return ""
    if pd.isna(ts):
        return ""
    return pd.Timestamp(ts).normalize().strftime("%Y-%m-%d")


def _summary_site_key(value) -> str:
    return "".join(ch for ch in _summary_text(value).upper() if ch.isalnum())


def _summary_candidate_files(bdt_files: list[str]) -> list[str]:
    bdt_paths = {os.path.normcase(os.path.abspath(p)) for p in bdt_files}
    scan_dirs = {os.path.dirname(os.path.abspath(p)) for p in bdt_files}
    candidates: set[str] = set()

    for directory in scan_dirs:
        try:
            names = os.listdir(directory)
        except OSError:
            continue
        for name in names:
            if name.startswith("~$") or name.startswith("._"):
                continue
            ext = os.path.splitext(name)[1].lower()
            if ext not in _SUMMARY_FILE_EXTS:
                continue
            full = os.path.abspath(os.path.join(directory, name))
            if os.path.normcase(full) in bdt_paths:
                continue
            if "bdt" in name.lower():
                continue
            candidates.add(full)

    def _sort_key(path: str):
        name = os.path.basename(path).lower()
        score = 0
        if "summary" in name:
            score += 8
        if "weekly" in name:
            score += 5
        if "battery" in name:
            score += 4
        if "update" in name:
            score += 3
        return (-score, name)

    return sorted(candidates, key=_sort_key)


def _extract_summary_rows(file_path: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    workbook = None

    for engine in ("calamine", "openpyxl"):
        try:
            workbook = pd.ExcelFile(file_path, engine=engine)
            break
        except Exception:
            continue

    if workbook is None:
        return rows

    try:
        for sheet_name in workbook.sheet_names:
            try:
                df = pd.read_excel(workbook, sheet_name=sheet_name, dtype=object)
            except Exception:
                continue
            if df is None or df.empty:
                continue

            mapped_cols: dict[str, str] = {}
            for col in df.columns:
                canonical = _SUMMARY_ALIAS_TO_CANONICAL.get(_normalize_summary_key(col))
                if canonical and canonical not in mapped_cols.values():
                    mapped_cols[col] = canonical

            if "Short Code" not in mapped_cols.values():
                continue
            if "Test Date" not in mapped_cols.values():
                continue

            for _, row in df.iterrows():
                summary_row: dict[str, str] = {}
                for source_col, canonical in mapped_cols.items():
                    raw = row.get(source_col, None)
                    if canonical == "Test Date":
                        summary_row[canonical] = _summary_date_key(raw)
                    elif canonical == "Short Code":
                        summary_row[canonical] = _summary_text(raw).upper()
                    else:
                        summary_row[canonical] = _summary_text(raw)

                if not summary_row.get("Short Code"):
                    continue
                if not any(summary_row.get(k, "") for k in _SUMMARY_CANONICAL_KEYS
                           if k not in {"Short Code", "Test Date"}):
                    continue
                rows.append(summary_row)
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    return rows


def _load_external_summary_lookup(bdt_files: list[str]) -> dict[str, dict]:
    """Build summary lookup maps keyed by (site, test_date) and by site."""
    lookup = {"by_site_date": {}, "by_site": {}}
    for path in _summary_candidate_files(bdt_files):
        for row in _extract_summary_rows(path):
            site_key = _summary_site_key(row.get("Short Code"))
            if not site_key:
                continue
            date_key = _summary_date_key(row.get("Test Date"))
            row_copy = dict(row)
            if date_key:
                row_copy["Test Date"] = date_key
                lookup["by_site_date"].setdefault((site_key, date_key), row_copy)

            by_site = lookup["by_site"].setdefault(site_key, {})
            by_site.setdefault(date_key, row_copy)
    return lookup


def _match_external_summary_row(bdt_data, summary_lookup: dict[str, dict]) -> dict[str, str] | None:
    if not summary_lookup:
        return None

    site_key = _summary_site_key(getattr(bdt_data, "site_code", ""))
    if not site_key:
        return None

    by_site_date = summary_lookup.get("by_site_date", {})
    by_site = summary_lookup.get("by_site", {})
    test_date_key = _summary_date_key(getattr(bdt_data, "test_date", None))

    if test_date_key:
        exact = by_site_date.get((site_key, test_date_key))
        if exact:
            return dict(exact)

    site_rows = by_site.get(site_key, {})
    if not site_rows:
        return None
    if len(site_rows) == 1:
        return dict(next(iter(site_rows.values())))
    if "" in site_rows and len(site_rows) == 2 and test_date_key:
        return dict(site_rows[""])
    return None


def discover_alarm_files(directory: str) -> list[dict]:
    """
    Recursively walk *directory* and return metadata dicts for every
    alarm-format .csv / .xlsx / .xls file found.

    Performs a fast header check on each file so only genuine alarm
    data appears in the list — BDT files, random spreadsheets, etc.
    are silently skipped.
    """
    results: list[dict] = []
    for root, _dirs, files in os.walk(directory):
        for fname in sorted(files):
            if fname.startswith("._") or fname.startswith("~$"):
                continue
            ext = os.path.splitext(fname)[1].lower()
            if ext not in _EXTS:
                continue
            # Fast skip: BDT files by name
            fl = fname.lower()
            if "bdt" in fl:
                continue
            full = os.path.join(root, fname)
            # Trust filenames that clearly indicate alarm data
            name_ok = any(hint in fl for hint in _ALARM_NAME_HINTS)
            if not name_ok:
                # Unknown file — verify header before listing
                if not _quick_header_check(full, ext):
                    continue
            rel  = os.path.relpath(full, directory)
            kb   = os.path.getsize(full) / 1024
            results.append({
                "path": full, "rel_path": rel,
                "filename": fname, "ext": ext,
                "size_kb": kb,
            })
    return results


def parse_alarm_file(info: dict) -> pd.DataFrame | None:
    """Parse one alarm file described by a discovery dict.  Returns None on failure."""
    fp, ext, fname = info["path"], info["ext"], info["filename"]
    df = None

    for enc in _ENCODINGS:
        try:
            if ext == ".csv":
                df = pd.read_csv(fp, encoding=enc,
                                 on_bad_lines="skip", low_memory=False)
            else:
                # calamine (Rust-based) is ~3x faster than openpyxl for read-only
                try:
                    df = pd.read_excel(fp, engine="calamine")
                except Exception:
                    df = pd.read_excel(fp, engine="openpyxl")
            if df is not None and not df.empty:
                break
        except Exception:
            continue

    if df is None or df.empty:
        return None

    # Normalise headers
    df.columns = [str(c).strip() for c in df.columns]

    # Early reject: skip files that don't match any alarm schema
    if not _is_alarm_header(df.columns.tolist()):
        return None

    cols = set(df.columns)

    # Choose schema
    if "FM Office" in cols or ("Site ID" in cols and "Site Name" not in cols):
        rmap = {k: v for k, v in SCHEMA_2_MAP.items() if k in cols}
    else:
        rmap = {k: v for k, v in SCHEMA_1_MAP.items() if k in cols}

    df = df.rename(columns=rmap)
    fname_lower = fname.lower()
    if "power" in fname_lower:
        df["alarm_category"] = "Power"
    elif "down" in fname_lower:
        df["alarm_category"] = "Down"
    elif "door" in fname_lower:
        df["alarm_category"] = "Door"
    else:
        df["alarm_category"] = ""
    df["file_source"]    = fname

    # Ensure every expected column exists
    for col in ALL_INTERNAL_COLS:
        if col not in df.columns:
            df[col] = np.nan

    return df[ALL_INTERNAL_COLS]


def classify_by_alarm_id(df: pd.DataFrame, alarm_ids: dict) -> pd.DataFrame:
    """Classify alarm_category based on alarm_id matching configured ID lists.

    Args:
        df: DataFrame with 'alarm_id' and 'alarm_category' columns.
        alarm_ids: {"power": [...], "down": [...], "door": [...]}
    Returns:
        DataFrame with updated 'alarm_category' column.
    """
    if df.empty or "alarm_id" not in df.columns:
        return df
    power_set = set(alarm_ids.get("power", []))
    down_set  = set(alarm_ids.get("down", []))
    door_set  = set(alarm_ids.get("door", []))
    # Normalize: floats like 300.0 → "300", strings stay as-is
    aid = (df["alarm_id"].fillna("").astype(str).str.strip()
           .str.replace(r'\.0$', '', regex=True))
    df = df.copy()
    df.loc[aid.isin(power_set), "alarm_category"] = "Power"
    df.loc[aid.isin(down_set),  "alarm_category"] = "Down"
    df.loc[aid.isin(door_set),  "alarm_category"] = "Door"

    # Heuristic fallback so door alarms are visible even without configured IDs.
    door_mask = pd.Series(False, index=df.index)
    door_rx = r"(?:^|[^a-z])door(?:[^a-z]|$)"
    for col in ("alarm_name", "file_source", "alarm_source"):
        if col in df.columns:
            door_mask |= df[col].astype(str).str.contains(
                door_rx, case=False, na=False, regex=True)
    df.loc[door_mask, "alarm_category"] = "Door"
    return df


def compute_site_down_flag(df: pd.DataFrame) -> pd.DataFrame:
    """Compute site_down_flag.

    - Down alarms → always 'Yes' (site went down by definition).
    - Power alarms → 'Yes' only if a Down alarm occurred on the same site
      within the Power alarm's [occurred_on, cleared_on] window
      (meaning the battery didn't hold and the site went down).
    - Everything else → 'No'.
    """
    if df.empty or "alarm_category" not in df.columns:
        return df

    df = df.copy()
    df["site_down_flag"] = "No"

    # All Down alarms = site is down
    df.loc[df["alarm_category"] == "Down", "site_down_flag"] = "Yes"

    need = ["site_id", "occurred_on", "cleared_on"]
    if not all(c in df.columns for c in need):
        return df

    pwr = df[df["alarm_category"] == "Power"].dropna(subset=["site_id", "occurred_on"])
    dwn = df[df["alarm_category"] == "Down"].dropna(subset=["site_id", "occurred_on"])

    if pwr.empty or dwn.empty:
        return df

    # For Power alarms: flag 'Yes' if a Down alarm fell inside its window
    pwr_data = pwr[["site_id", "occurred_on", "cleared_on"]].copy()
    pwr_data["_pwr_idx"] = pwr.index
    dwn_data = dwn[["site_id", "occurred_on"]].rename(
        columns={"occurred_on": "down_time"})

    merged = pwr_data.merge(dwn_data, on="site_id")
    mask = merged["down_time"] >= merged["occurred_on"]
    mask = mask & (merged["down_time"] <= merged["cleared_on"].fillna(pd.Timestamp.max))

    matched_power_idx = merged.loc[mask, "_pwr_idx"].unique()
    df.loc[matched_power_idx, "site_down_flag"] = "Yes"

    return df


# ─────────────────────────────────────────────────────────────────
# Duration helpers
# ─────────────────────────────────────────────────────────────────
def _duration_to_secs(val) -> float:
    """Convert a duration value (str, time, Timestamp) to seconds."""
    if pd.isna(val) or val is None:
        return 0.0
    # datetime.time object
    import datetime as _dt
    if isinstance(val, _dt.time):
        return val.hour * 3600 + val.minute * 60 + val.second
    # Timestamp (Excel serial date like 1900-01-01 00:02:20)
    if isinstance(val, pd.Timestamp):
        return val.hour * 3600 + val.minute * 60 + val.second
    # String "HH:MM:SS"
    s = str(val).strip()
    parts = s.split(":")
    if len(parts) >= 3:
        try:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
        except (ValueError, TypeError):
            pass
    return 0.0


def _secs_to_hhmmss(s) -> str:
    """Convert seconds to HH:MM:SS string."""
    if s <= 0:
        return ""
    h = int(s // 3600)
    m = int((s % 3600) // 60)
    sec = int(s % 60)
    return f"{h:02d}:{m:02d}:{sec:02d}"


# ─────────────────────────────────────────────────────────────────
# Background loader thread
# ─────────────────────────────────────────────────────────────────
class LoaderThread(QThread):
    """Load selected files in a background thread.

    Signals:
        progress(int, str)  — percentage + status message
        finished(DataFrame, str) — merged data + summary message
        error(str) — traceback on failure
    """
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(object, str)
    error    = pyqtSignal(str)

    def __init__(self, file_infos: list[dict]):
        super().__init__()
        self.file_infos = file_infos

    def run(self):
        try:
            dfs: list[pd.DataFrame] = []
            total = len(self.file_infos)

            # Sort largest-first so big files start immediately
            ordered = sorted(
                enumerate(self.file_infos),
                key=lambda t: t[1].get("size_kb", 0),
                reverse=True,
            )

            workers = min(total, os.cpu_count() or 1, 6)
            done_count = 0

            with ThreadPoolExecutor(max_workers=workers) as pool:
                futures = {
                    pool.submit(parse_alarm_file, info): idx
                    for idx, info in ordered
                }
                for future in as_completed(futures):
                    done_count += 1
                    idx = futures[future]
                    info = self.file_infos[idx]
                    self.progress.emit(
                        int(done_count / total * 90),
                        f"[{done_count}/{total}]  {info['filename']}",
                    )
                    try:
                        df = future.result()
                        if df is not None and not df.empty:
                            dfs.append(df)
                    except Exception:
                        pass  # individual file failures silently skipped

            if not dfs:
                self.error.emit(
                    "No readable alarm records found in selected files.")
                return

            self.progress.emit(95, "Merging records …")
            combined = pd.concat(dfs, ignore_index=True)

            # Fast vectorised datetime conversion
            for col in ("occurred_on", "cleared_on"):
                if col in combined.columns:
                    combined[col] = pd.to_datetime(
                        combined[col], errors="coerce", format="mixed")

            if "site_id" in combined.columns:
                combined["site_id"] = (
                    combined["site_id"].astype(str).str.strip())

            # Compute duration for records that don't have it (Nokia)
            if "duration" in combined.columns:
                missing_dur = combined["duration"].fillna("").astype(str).str.strip().eq("")
                if missing_dur.any():
                    has_times = (missing_dur
                                 & combined["occurred_on"].notna()
                                 & combined["cleared_on"].notna())
                    if has_times.any():
                        td = combined.loc[has_times, "cleared_on"] - combined.loc[has_times, "occurred_on"]
                        total_secs = td.dt.total_seconds().fillna(0)
                        h = (total_secs // 3600).astype(int)
                        m = ((total_secs % 3600) // 60).astype(int)
                        s = (total_secs % 60).astype(int)
                        combined.loc[has_times, "duration"] = (
                            h.astype(str).str.zfill(2) + ":"
                            + m.astype(str).str.zfill(2) + ":"
                            + s.astype(str).str.zfill(2))

            # Pre-computed duration seconds for fast filtering
            # Handles str "HH:MM:SS", datetime.time, and Timestamp objects
            if "duration" in combined.columns:
                combined["_duration_secs"] = combined["duration"].apply(_duration_to_secs)
                # Normalize duration display to HH:MM:SS strings
                combined["duration"] = combined["_duration_secs"].apply(_secs_to_hhmmss)

            self.progress.emit(100, "Done!")
            self.finished.emit(
                combined,
                f"Loaded {len(combined):,} records from {len(dfs)} file(s)",
            )
        except Exception:
            self.error.emit(traceback.format_exc())


# ─────────────────────────────────────────────────────────────────
# Background export thread
# ─────────────────────────────────────────────────────────────────
class ExportThread(QThread):
    """Write one or more DataFrames to Excel in a background thread.

    Signals:
        progress(int, str)  — percentage + status message
        finished(str)       — file path on success
        error(str)          — error message on failure
    """
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(str)
    error    = pyqtSignal(str)

    def __init__(self, df: pd.DataFrame | dict[str, pd.DataFrame], path: str):
        super().__init__()
        self._df = df
        self._path = path

    def run(self):
        try:
            self.progress.emit(30, "Writing Excel file …")
            if isinstance(self._df, dict):
                with pd.ExcelWriter(self._path, engine="openpyxl") as writer:
                    total = max(len(self._df), 1)
                    for idx, (sheet_name, df) in enumerate(self._df.items(), start=1):
                        frame = df if isinstance(df, pd.DataFrame) else pd.DataFrame(df)
                        frame.to_excel(writer, sheet_name=sheet_name, index=False)
                        pct = 30 + int(60 * idx / total)
                        self.progress.emit(min(pct, 95), f"Writing sheet: {sheet_name}")
            else:
                self._df.to_excel(self._path, index=False, engine="openpyxl")
            self.progress.emit(100, "Export complete")
            self.finished.emit(self._path)
        except Exception:
            self.error.emit(traceback.format_exc())


class BDTValidationThread(QThread):
    """Parse and validate BDT files in a background thread.

    Signals:
        progress(int, str) — percentage + status message
        finished(list, dict) — (ValidationResult list, site→BDTData dict)
        error(str)         — error message on failure
    """
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(object, object)
    error    = pyqtSignal(str)

    def __init__(self, bdt_files: list[str], alarm_df,
                 tolerance: float, health_pct: float):
        super().__init__()
        self._files = bdt_files
        self._alarm_df = alarm_df
        self._tolerance = tolerance
        self._health_pct = health_pct

    def run(self):
        try:
            from .bdt_parser import parse_bdt_file, load_bdt_photos
            from .bdt_validator import validate_bdt
        except ImportError:
            from bdt_parser import parse_bdt_file, load_bdt_photos
            from bdt_validator import validate_bdt
        from datetime import datetime

        try:
            total = len(self._files)
            results = []
            by_site: dict[str, list] = {}
            done = 0
            summary_lookup = _load_external_summary_lookup(self._files)

            workers = min(total, os.cpu_count() or 1, 8)

            with ThreadPoolExecutor(max_workers=workers) as pool:
                futures = {
                    pool.submit(parse_bdt_file, fp, skip_photos=True): fp
                    for fp in self._files
                }
                for future in as_completed(futures):
                    done += 1
                    fp = futures[future]
                    fname = os.path.basename(fp)
                    pct = int(done / total * 90)
                    self.progress.emit(
                        pct, f"[{done}/{total}]  {fname}")

                    try:
                        bdt_data = future.result()
                    except Exception:
                        continue

                    # Skip non-template files and hard parser failures to avoid
                    # polluting validation table with unusable "Unknown" rows.
                    parse_errors_lc = [str(e).lower() for e in getattr(bdt_data, "errors", [])]
                    hard_file_error = any(
                        ("cannot open file" in err) or ("failed to read bdt sheet" in err)
                        for err in parse_errors_lc
                    )
                    no_extractable_data = (
                        not getattr(bdt_data, "site_code", "")
                        and not getattr(bdt_data, "test_date", None)
                        and not getattr(bdt_data, "discharge_readings", [])
                        and getattr(bdt_data, "start_voltage", None) is None
                        and getattr(bdt_data, "start_ampere", None) is None
                    )
                    if hard_file_error or no_extractable_data:
                        self.progress.emit(
                            pct, f"[{done}/{total}]  skipped invalid BDT: {fname}")
                        continue

                    # R1 photo rule must evaluate actual image availability.
                    # Bulk parsing may defer photos for speed, so load now.
                    if getattr(bdt_data, "photos_deferred", False):
                        load_bdt_photos(bdt_data)

                    if summary_lookup:
                        matched_summary = _match_external_summary_row(
                            bdt_data, summary_lookup)
                        if matched_summary:
                            bdt_data.summary_data = matched_summary

                    result = validate_bdt(
                        bdt_data, self._alarm_df,
                        self._tolerance, self._health_pct)
                    results.append(result)

                    # Persist test record for historical comparison
                    try:
                        try:
                            from .bdt_history import save_test_record
                        except ImportError:
                            from bdt_history import save_test_record
                        save_test_record(bdt_data, result.overall)
                    except Exception:
                        pass  # history saving is best-effort

                    if bdt_data.site_code:
                        key = bdt_data.site_code.strip().upper()
                        by_site.setdefault(key, []).append(bdt_data)

            self.progress.emit(95, "Sorting results…")

            # Sort each site's tests by date (newest first)
            for key in by_site:
                by_site[key].sort(
                    key=lambda b: b.test_date or datetime.min,
                    reverse=True)

            self.progress.emit(100, "Done!")
            self.finished.emit(results, by_site)

        except Exception:
            self.error.emit(traceback.format_exc())
