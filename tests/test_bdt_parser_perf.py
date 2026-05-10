"""Performance test — synthetic 200-sheet BDT workbook parse."""

import time

import pytest

try:
    from alarm_app.bdt._workbook import WorkbookEngine
except ImportError:
    WorkbookEngine = None  # type: ignore[assignment]


@pytest.mark.slow
def test_parse_200_sheet_workbook_under_5_seconds(tmp_path):
    """Synthetic 200-sheet BDT workbook must parse in under 5 seconds."""
    from datetime import datetime

    from openpyxl import Workbook

    path = tmp_path / "big_workbook.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for i in range(200):
        ws = wb.create_sheet(f"Sheet{i}")
        ws["A1"] = f"data_{i}"

    ws = wb.create_sheet("BDT")
    ws["C4"] = "0167DE"
    ws["T3"] = datetime(2026, 1, 15)
    wb.save(str(path))

    engine = WorkbookEngine(str(path))
    start = time.perf_counter()
    names = engine.sheet_names
    for name in names:
        list(engine.sheet_rows(name))
    elapsed = time.perf_counter() - start
    engine.close()

    assert elapsed < 5.0, f"200-sheet parse took {elapsed:.2f}s, expected <5s"
    assert len(names) == 201
