"""Tests for uncovered Temp alarms outside Power coverage windows."""

from zipfile import ZipFile

import pandas as pd
import pytest
from openpyxl import load_workbook

from alarm_app.core.temp_alarm import (
    DEFAULT_HT_CLEARANCE_GAP_X_SECS,
    DEFAULT_HT_HISTORY_START_WEEK,
    DEFAULT_HT_SUMMARY_MIN_DURATION_Y_SECS,
    HtWorkbookFilterSettings,
    HtWorkbookPrecomputed,
    build_temp_alarm_summary,
    compute_ht_consolidated_meet_source,
    compute_ht_meet_rows,
    compute_temp_alarm_matches,
    compute_temp_alarm_matches_for_query,
    describe_meet_preview_empty_state,
    enrich_source_with_site_metadata,
    export_temp_alarm_workbook,
    filter_temp_matches_to_query,
    filter_temp_matches_to_selected_temps,
    filter_temps_by_power_clearance_gap,
    ht_export_filename,
    ht_export_week_from_date,
    ht_export_week_range,
    infer_export_week_label,
)
from alarm_app.data.alarm_store import AlarmQuery
from alarm_app.ui.viewer import AlarmViewer


def _make_df(rows):
    defaults = {
        "site_id": "SITE_A",
        "site_name": "Site Alpha",
        "site_code": "SITE_A",
        "area": "AGLI",
        "contractor": "Orascom",
        "battery_type": "Power safe 155",
        "battery_status": "Good (2 - 3 Hrs)",
        "support": "AUTIN",
        "cleared_by": "oss_user",
        "alarm_reporting_type": "Normal",
        "alarm_source": "U_G_SITE_A_TEST",
        "alarm_name": "MAIN POWER CUT OFF",
        "occurred_on": None,
        "cleared_on": None,
        "duration": "",
        "alarm_category": "Power",
        "network_type": "4G",
        "vendor": "Huawei",
        "clearance_status": "Cleared",
    }
    records = []
    for row in rows:
        rec = {**defaults, **row}
        for col in ("occurred_on", "cleared_on"):
            if isinstance(rec[col], str):
                rec[col] = pd.Timestamp(rec[col])
        records.append(rec)
    return pd.DataFrame(records)


_NO_GAP_FILTER = HtWorkbookFilterSettings(clearance_gap_x_secs=None)


def test_temp_during_active_power_outage_is_covered():
    """Temp alarm during an ongoing (uncleared) power outage must be excluded from uncovered."""
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            # cleared_on intentionally omitted → NaT (ongoing outage)
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 11:00:00",
            "cleared_on": "2026-02-01 11:30:00",
            "duration": "00:30:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=60)

    assert result.empty, f"Expected empty, got {len(result)} rows"
    assert "No uncovered Temp alarms" in err


def test_multiple_temps_during_active_outage_all_covered():
    """Multiple temp alarms during the same ongoing power outage are all excluded from uncovered."""
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            # cleared_on intentionally omitted → NaT (ongoing)
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 11:00:00",
            "cleared_on": "2026-02-01 11:30:00",
            "duration": "00:30:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "External AL 9",
            "occurred_on": "2026-02-01 14:00:00",
            "cleared_on": "2026-02-01 14:15:00",
            "duration": "00:15:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-02 08:00:00",
            "cleared_on": "2026-02-02 08:45:00",
            "duration": "00:45:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=60)

    assert result.empty, f"Expected empty, got {len(result)} rows"
    assert "No uncovered Temp alarms" in err


def test_active_power_outage_has_unbounded_temp_coverage():
    """An uncleared Power alarm covers Temp alarms indefinitely after it starts."""
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2100-01-01 11:00:00",
            "cleared_on": "2100-01-01 11:30:00",
            "duration": "00:30:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=60)

    assert result.empty, f"Expected empty, got {len(result)} rows"
    assert "No uncovered Temp alarms" in err


def test_temp_before_active_power_outage_is_uncovered():
    """Temp alarm occurring BEFORE the power outage started is still uncovered."""
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 12:00:00",
            # cleared_on intentionally omitted → NaT (ongoing)
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 10:30:00",
            "duration": "00:30:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=60)

    assert err == ""
    assert len(result) == 1, f"Expected 1 uncovered, got {len(result)}"
    assert result.iloc[0]["match_window"] == "No same-site Power alarm before Temp"


def test_mix_of_cleared_and_uncleared_power_coverage():
    """Site with both cleared and uncleared power alarms: coverage works for both."""
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": "2026-02-01 10:00:00",  # cleared — covers up to 11:00 with Y=60
        },
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 12:00:00",
            # cleared_on omitted → NaT (ongoing) — covers everything from 12:00 onward
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 09:00:00",  # inside cleared power window → covered
            "cleared_on": "2026-02-01 09:30:00",
            "duration": "00:30:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "External AL 9",
            "occurred_on": "2026-02-01 11:30:00",  # after cleared window, before ongoing → UNCOVERED
            "cleared_on": "2026-02-01 11:45:00",
            "duration": "00:15:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 13:00:00",  # inside ongoing power window → covered
            "cleared_on": "2026-02-01 13:30:00",
            "duration": "00:30:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=60)

    assert err == ""
    assert len(result) == 1, f"Expected 1 uncovered, got {len(result)}"
    assert result.iloc[0]["temp_alarm_name"] == "External AL 9"


def test_excludes_temp_inside_power_window():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 11:00:00",
            "cleared_on": "2026-02-01 11:30:00",
            "duration": "00:30:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=60)

    assert result.empty
    assert "No uncovered Temp alarms" in err


def test_excludes_temp_after_power_clearance_within_y_margin():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 12:45:00",
            "cleared_on": "2026-02-01 13:00:00",
            "duration": "00:15:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=60)

    assert result.empty
    assert "No uncovered Temp alarms" in err


def test_excludes_multiple_temp_rows_inside_same_power_window():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 11:30:00",
            "cleared_on": "2026-02-01 11:40:00",
            "duration": "00:10:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 10:30:00",
            "cleared_on": "2026-02-01 10:40:00",
            "duration": "00:10:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=60)

    assert result.empty
    assert "No uncovered Temp alarms" in err


def test_includes_temp_after_y_margin():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 13:01:00",
            "cleared_on": "2026-02-01 13:30:00",
            "duration": "00:29:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=60)

    assert err == ""
    assert len(result) == 1
    row = result.iloc[0]
    assert row["match_window"] == "Outside Power coverage"
    assert row["power_time"] == "2026-02-01 10:00:00"
    assert row["power_cleared"] == "2026-02-01 12:00:00"
    assert row["temp_delay_after_power"] == "03:01:00"
    assert row["temp_delay_after_power_clearance"] == "01:01:00"


def test_uncovered_temp_after_prior_power_includes_power_context():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 13:01:00",
            "cleared_on": "2026-02-01 13:30:00",
            "duration": "00:29:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=60)

    assert err == ""
    assert len(result) == 1
    row = result.iloc[0]
    assert row["power_time"] == "2026-02-01 10:00:00"
    assert row["power_cleared"] == "2026-02-01 12:00:00"
    assert row["x_duration"] == "02:00:00"
    assert row["temp_delay_after_power"] == "03:01:00"
    assert row["temp_delay_after_power_clearance"] == "01:01:00"
    assert row["match_window"] == "Outside Power coverage"


def test_margin_can_exceed_sixty_minutes():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 13:30:00",
            "cleared_on": "2026-02-01 14:00:00",
            "duration": "00:30:00",
        },
    ])

    result, err = compute_temp_alarm_matches(df, margin_minutes=120)

    assert result.empty
    assert "No uncovered Temp alarms" in err


def test_summary_filters_to_requested_week_and_keeps_site_metadata():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SRAN_LWG_TEST_SITE_A",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 13:15:00",
            "cleared_on": "2026-02-01 13:45:00",
            "duration": "00:30:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SRAN_LWG_TEST_SITE_A_2",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 14:15:00",
            "cleared_on": "2026-02-01 15:00:00",
            "duration": "00:45:00",
        },
        {
            "site_id": "SITE_C",
            "site_name": "Site Charlie",
            "site_code": "SITE_C",
            "area": "CAIR",
            "contractor": "Huawei",
            "battery_type": "Lithium 100Ah",
            "battery_status": "Good (2 - 3 Hrs)",
            "alarm_category": "Temp",
            "alarm_source": "SRAN_LWG_TEST_SITE_C",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-01-21 14:15:00",
            "cleared_on": "2026-01-21 15:15:00",
            "duration": "01:00:00",
        },
    ])
    matches, err = compute_temp_alarm_matches(df, margin_minutes=60)
    assert err == ""

    summary = build_temp_alarm_summary(matches, week_label="W05-26")

    assert list(summary.columns) == [
        "##", "Site Name", "Site Code", "Area", "Contractor",
        "No. Of HT Alarms", "HT Duration", "Batteries Types",
        "Batteries Status", "Week No.", "W05-26", "W04-26", "W03-26",
        "W02-26", "W01-26", "W52-25", "W51-25", "W50-25",
    ]
    assert len(summary) == 1
    site_a = summary[summary["Site Code"] == "SITE_A"].iloc[0]
    assert site_a["Site Name"] == "Site Alpha"
    assert site_a["Area"] == "AGLI"
    assert site_a["Contractor"] == "Orascom"
    assert site_a["Batteries Types"] == "Power safe 155"
    assert site_a["Batteries Status"] == "Good (2 - 3 Hrs)"
    assert site_a["No. Of HT Alarms"] == 2
    assert site_a["HT Duration"] == "01:15"
    assert site_a["Week No."] == "W05-26"
    assert site_a["W05-26"] == "W05-26"
    assert site_a["W03-26"] == ""


def test_summary_separates_weeks_when_no_week_filter_is_requested():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SRAN_LWG_TEST_SITE_A",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 13:15:00",
            "cleared_on": "2026-02-01 13:45:00",
            "duration": "00:30:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SRAN_LWG_TEST_SITE_A_2",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-01-21 14:15:00",
            "cleared_on": "2026-01-21 15:15:00",
            "duration": "01:00:00",
        },
    ])
    matches, err = compute_temp_alarm_matches(df, margin_minutes=60)
    assert err == ""

    summary = build_temp_alarm_summary(matches, week_label=None)

    assert list(summary["Week No."]) == ["W03-26", "W05-26"]
    week_03 = summary[summary["Week No."] == "W03-26"].iloc[0]
    week_05 = summary[summary["Week No."] == "W05-26"].iloc[0]
    assert week_03["No. Of HT Alarms"] == 1
    assert week_03["HT Duration"] == "01:00"
    assert week_03["W03-26"] == "W03-26"
    assert week_05["No. Of HT Alarms"] == 1
    assert week_05["HT Duration"] == "00:30"
    assert week_05["W05-26"] == "W05-26"


def test_ht_export_week_uses_jan1_fixed_seven_day_blocks():
    week = ht_export_week_from_date("2024-06-30")

    assert week["week_label"] == "W26-24"
    assert week["short_week_label"] == "W26"
    assert str(week["start"].date()) == "2024-06-24"
    assert str(week["end"].date()) == "2024-07-01"
    assert week["filename"] == "2024-HT-Alarms-W26.xlsx"
    assert ht_export_filename("W26-24") == "2024-HT-Alarms-W26.xlsx"


def test_ht_export_week_jan1_block_boundaries():
    assert ht_export_week_from_date("2026-01-01")["week_label"] == "W01-26"
    assert ht_export_week_from_date("2026-01-07")["week_label"] == "W01-26"
    assert ht_export_week_from_date("2026-01-08")["week_label"] == "W02-26"


def test_ht_export_week_may_2026_search_example():
    assert ht_export_week_from_date("2026-05-24")["week_label"] == "W21-26"
    assert ht_export_week_from_date("2026-05-31")["week_label"] == "W22-26"
    start, end = ht_export_week_range("W19-26")
    assert str(start.date()) == "2026-05-07"
    assert str(end.date()) == "2026-05-14"
    start, end = ht_export_week_range("W21-26")
    assert str(start.date()) == "2026-05-21"
    assert str(end.date()) == "2026-05-28"
    start, end = ht_export_week_range("W22-26")
    assert str(start.date()) == "2026-05-28"
    assert str(end.date()) == "2026-06-04"


def test_ht_export_week_rejects_invalid_week_for_year():
    with pytest.raises(ValueError):
        ht_export_week_range("W54-28")


def test_infer_export_week_label_prefers_date_from_and_earliest_manual_day():
    assert infer_export_week_label(
        date_from=pd.Timestamp("2026-05-24"),
        date_to=pd.Timestamp("2026-05-31"),
    ) == "W21-26"
    assert infer_export_week_label(
        manual_days=[pd.Timestamp("2026-05-31"), pd.Timestamp("2026-05-24")],
        date_to=pd.Timestamp("2026-05-31"),
    ) == "W21-26"
    assert infer_export_week_label(date_to=pd.Timestamp("2026-05-31")) == "W22-26"
    assert infer_export_week_label(
        fallback_times=pd.to_datetime(["2026-05-31 10:00:00", "2026-05-24 08:00:00"])
    ) == "W22-26"


def test_meet_engine_uses_daily_power_unavailable_or_diff_greater_than_seven_hours():
    unavailable = _make_df([
        {
            "alarm_category": "Temp",
            "alarm_source": "NO_POWER",
            "occurred_on": "2024-06-30 08:00:00",
            "cleared_on": "2024-06-30 16:00:00",
            "duration": "08:00:00",
        }
    ])
    _, meet, _ = compute_ht_meet_rows(
        unavailable, week_label="W26-24", filter_settings=_NO_GAP_FILTER
    )
    assert meet["Alarm Source"].tolist() == ["NO_POWER"]

    greater_than = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2024-06-30 06:00:00",
            "cleared_on": "2024-06-30 06:59:00",
            "duration": "00:59:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "GT_7H",
            "occurred_on": "2024-06-30 08:00:00",
            "cleared_on": "2024-06-30 16:00:00",
            "duration": "08:00:00",
        },
    ])
    study, meet, _ = compute_ht_meet_rows(
        greater_than, week_label="W26-24", filter_settings=_NO_GAP_FILTER
    )
    assert study["Meet"].tolist() == ["Yes"]
    assert meet["Alarm Source"].tolist() == ["GT_7H"]

    exactly = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2024-06-30 06:00:00",
            "cleared_on": "2024-06-30 07:00:00",
            "duration": "01:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "EQ_7H",
            "occurred_on": "2024-06-30 08:00:00",
            "cleared_on": "2024-06-30 16:00:00",
            "duration": "08:00:00",
        },
    ])
    study, meet, _ = compute_ht_meet_rows(
        exactly, week_label="W26-24", filter_settings=_NO_GAP_FILTER
    )
    assert study["Meet"].tolist() == [""]
    assert meet.empty


def test_export_workbook_contains_reference_style_temp_alarm_sheets(tmp_path):
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 10:30:00",
            "duration": "00:30:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SRAN_LWG_TEST_SITE_A_MEET",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 11:15:00",
            "cleared_on": "2026-02-01 15:15:00",
            "duration": "04:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SRAN_LWG_TEST_SITE_A",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 16:00:00",
            "cleared_on": "2026-02-01 20:00:00",
            "duration": "04:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SRAN_LWG_TEST_SITE_A_OLD_WEEK",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-01-21 13:15:00",
            "cleared_on": "2026-01-21 14:15:00",
            "duration": "01:00:00",
        },
    ])
    matches, err = compute_temp_alarm_matches(df, margin_minutes=60)
    assert err == ""
    out = tmp_path / "temp_alarm_export.xlsx"

    export_temp_alarm_workbook(
        matches, out, week_label="W05-26", source_df=df, filter_settings=_NO_GAP_FILTER
    )

    wb = load_workbook(out, data_only=False)
    with ZipFile(out) as archive:
        for sheet_index in range(1, 7):
            xml = archive.read(f"xl/worksheets/sheet{sheet_index}.xml")
            assert b"<dimension " in xml
            assert xml.find(b"<sheetPr") < xml.find(b"<dimension ")
    assert wb.sheetnames == [
        "W05 AUTIN HT",
        "W05 AUTIN Power",
        "W05 AUTIN HT Study",
        "Meet",
        "W05",
        "Consolidated",
    ]
    assert wb["W05 AUTIN HT"].max_column == 12
    assert wb["W05 AUTIN Power"].max_column == 11
    assert wb["W05 AUTIN HT Study"].max_column == 16
    assert wb["Meet"].max_column == 10
    assert wb["W05 AUTIN HT"].max_row == 3
    assert wb["W05 AUTIN HT Study"].max_row == 3

    ht_raw = wb["W05 AUTIN HT"]
    assert [ht_raw.cell(row=1, column=col).value for col in range(1, 13)] == [
        "Alarm Source",
        "Site Name",
        "Site ID",
        "Last Occurred On",
        "Cleared On",
        "Duration\n(hh:mm:ss)",
        "Alarm Name",
        "Clearance Status",
        "Cleared By",
        "Alarm Reporting Type",
        "Week",
        "Area",
    ]
    assert ht_raw["A2"].value == "SRAN_LWG_TEST_SITE_A_MEET"
    assert ht_raw["D2"].value == "2/1/26 11:15"
    assert ht_raw["F2"].value == "04:00:00"
    assert ht_raw["I2"].value == "oss_user"
    assert ht_raw["J2"].value == "Normal"
    assert ht_raw["K2"].value == 5
    assert ht_raw["C2"].value == "SITE_A"

    power_raw = wb["W05 AUTIN Power"]
    assert [power_raw.cell(row=1, column=col).value for col in range(1, 12)] == [
        "Alarm Source",
        "Site Name",
        "Site ID",
        "Last Occurred On",
        "Cleared On",
        "Duration(hh:mm:ss)",
        "Alarm Name",
        "Clearance Status",
        "Cleared By",
        "Alarm Reporting Type",
        "SUM",
    ]
    assert power_raw["A2"].value == "U_G_SITE_A_TEST"
    assert power_raw["C2"].value == "SITE_A"
    assert power_raw["D2"].value == "2/1/26 10:00"
    assert power_raw["F2"].number_format == "[hh]:mm"
    assert power_raw["K2"].value == (
        '=SUMIFS($F:$F,$B:$B,B2,$D:$D,">="&INT($D2),$D:$D,"<"&INT($D2)+1)'
    )

    study = wb["W05 AUTIN HT Study"]
    assert [study.cell(row=1, column=col).value for col in range(1, 17)] == [
        "Alarm Source",
        "Site Name",
        "Site ID",
        "Support",
        "Day",
        "Last Occurred On",
        "Cleared On",
        "Duration\n(hh:mm:ss)",
        "Alarm Name",
        "Clearance Status",
        "Cleared By",
        "Alarm Reporting Type",
        "HT SUM IFS",
        "Powr SUM IFS",
        "Diff",
        "Meet",
    ]
    assert study.max_row == 3
    assert study["A2"].value == "SRAN_LWG_TEST_SITE_A_MEET"
    assert study["F2"].value == "2/1/26 11:15"
    assert study["D2"].value == '=B2&" "&E2'
    assert study["E2"].value == "=DAY(F2)"
    assert study["M2"].value == "=SUMIFS(H:H,B:B,B2,E:E,E2)"
    assert study["N2"].value == (
        "=SUMIFS('W05 AUTIN Power'!$F:$F,'W05 AUTIN Power'!$B:$B,$B2,"
        "'W05 AUTIN Power'!$D:$D,\">=\"&INT($F2),'W05 AUTIN Power'!$D:$D,\"<\"&INT($F2)+1)"
    )
    assert study["O2"].value == "=M2-N2"
    assert study["P2"].value == "Yes"

    meet = wb["Meet"]
    assert [meet.cell(row=1, column=col).value for col in range(1, 11)] == [
        "Site Name",
        "Site ID",
        "Alarm Source",
        "Last Occurred On",
        "Cleared On",
        "Duration(hh:mm:ss)",
        "Alarm Name",
        "Clearance Status",
        "Cleared By",
        "Alarm Reporting Type",
    ]
    assert meet["A2"].value == "Site Alpha"
    assert meet["B2"].value == "SITE_A"
    assert meet["C2"].value == "SRAN_LWG_TEST_SITE_A_MEET"
    assert meet["D2"].value == "2/1/26 11:15"

    summary = wb["W05"]
    consolidated = wb["Consolidated"]
    assert summary.max_column == 18
    assert consolidated.max_column == 18
    assert [summary.cell(row=1, column=col).value for col in range(1, 19)] == [
        "##", "Site Name", "Site Code", "Area", "Contractor",
        "No. Of HT Alarms", "HT Duration", "Batteries Types",
        "Batteries Status", "Week No.", "W05-26", "W04-26", "W03-26",
        "W02-26", "W01-26", "W52-25", "W51-25", "W50-25",
    ]
    assert summary["B2"].value == "Site Alpha"
    assert summary["C2"].value == "SITE_A"
    assert summary["G2"].value == "08:00"
    assert summary["A1"].fill.fgColor.rgb == "004F81BD"
    assert summary["A1"].border.left.color.rgb == "FF000000"
    assert wb["W05 AUTIN HT"]["A1"].fill.fgColor.rgb == "00FFC000"
    assert wb["W05 AUTIN HT Study"]["M1"].fill.fgColor.rgb == "0092D050"
    assert wb["W05 AUTIN Power"]["D1"].fill.fgColor.rgb == "00FFC000"
    assert wb["Meet"].max_row == 3
    assert consolidated.max_row >= summary.max_row


def test_export_enriches_site_metadata_and_keeps_six_sheets_when_complete(tmp_path):
    df = _make_df([
        {
            "site_id": "site-a",
            "site_name": "Alarm Site Name",
            "site_code": "site-a",
            "area": "OLD",
            "contractor": "OLD",
            "battery_type": "OLD",
            "battery_status": "OLD",
            "alarm_category": "Temp",
            "alarm_source": "SRAN_SITE_A_HT",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": "2026-02-01 16:00:00",
            "duration": "08:00:00",
        }
    ])
    metadata = pd.DataFrame([
        {
            "site_id": "SITEA",
            "site_name": "Network Site A",
            "orange_area": "AGLI",
            "subcontractor": "Huawei",
            "battery_type": "Lithium",
            "backup_status": "Good",
        }
    ])
    out = tmp_path / "metadata_complete.xlsx"

    warnings = export_temp_alarm_workbook(
        pd.DataFrame(), out, week_label="W05-26", source_df=df, site_metadata_df=metadata,
        return_warnings=True, filter_settings=_NO_GAP_FILTER,
    )

    wb = load_workbook(out, data_only=False)
    assert wb.sheetnames == ["W05 AUTIN HT", "W05 AUTIN Power", "W05 AUTIN HT Study", "Meet", "W05", "Consolidated"]
    assert warnings["missing_metadata_site_ids"] == []
    assert warnings["missing_metadata_count"] == 0
    assert "stage_timings" in warnings
    summary = wb["W05"]
    assert summary["B2"].value == "Network Site A"
    assert summary["C2"].value == "SITEA"
    assert summary["D2"].value == "AGLI"
    assert summary["E2"].value == "Huawei"
    assert summary["H2"].value == "Lithium"
    assert summary["I2"].value == "Good"


def test_export_metadata_fallback_parses_alarm_source_and_missing_adds_sheet(tmp_path):
    df = _make_df([
        {
            "site_id": "",
            "site_name": "",
            "site_code": "",
            "alarm_category": "Temp",
            "alarm_source": "SRAN_SITEB_HT",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": "2026-02-01 16:00:00",
            "duration": "08:00:00",
        },
        {
            "site_id": "SITE_C",
            "alarm_category": "Temp",
            "alarm_source": "SRAN_SITE_C_HT",
            "occurred_on": "2026-02-01 09:00:00",
            "cleared_on": "2026-02-01 17:00:00",
            "duration": "08:00:00",
        },
    ])
    metadata = pd.DataFrame([
        {"site_id": "SITEB", "site_name": "Fallback Site B", "area": "CAIR"}
    ])
    out = tmp_path / "metadata_missing.xlsx"

    warnings = export_temp_alarm_workbook(
        pd.DataFrame(), out, week_label="W05-26", source_df=df, site_metadata_df=metadata,
        return_warnings=True, filter_settings=_NO_GAP_FILTER,
    )

    wb = load_workbook(out, data_only=False)
    assert "Missing Metadata" in wb.sheetnames
    assert warnings["missing_metadata_site_ids"] == ["SITEC"]
    assert wb["W05"]["B2"].value == "Fallback Site B"
    assert wb["W05"]["C2"].value == "SITEB"
    missing = wb["Missing Metadata"]
    assert missing["A2"].value == "SITEC"


def test_filter_matches_to_original_date_scope_uses_temp_time():
    matches = pd.DataFrame([
        {"site_id": "A", "temp_time": "2026-02-01 23:30:00"},
        {"site_id": "B", "temp_time": "2026-02-02 00:30:00"},
    ])
    query = AlarmQuery(date_from=pd.Timestamp("2026-02-02"), date_to=pd.Timestamp("2026-02-02"))

    filtered = filter_temp_matches_to_query(matches, query)

    assert filtered["site_id"].tolist() == ["B"]


def test_filter_matches_to_selected_temp_rows_uses_site_and_temp_time():
    matches = pd.DataFrame([
        {"site_id": "A", "temp_time": "2026-02-01 11:00:00", "temp_cleared": "2026-02-01 11:10:00", "temp_alarm_name": "Shelter High Temperature", "temp_alarm_source": "SRC-1"},
        {"site_id": "A", "temp_time": "2026-02-01 11:00:00", "temp_cleared": "2026-02-01 11:20:00", "temp_alarm_name": "Shelter High Temperature", "temp_alarm_source": "SRC-2"},
    ])
    selected_temp = pd.DataFrame([
        {"site_id": "A", "occurred_on": pd.Timestamp("2026-02-01 11:00:00"), "cleared_on": pd.Timestamp("2026-02-01 11:20:00"), "alarm_name": "Shelter High Temperature", "alarm_source": "SRC-2"},
    ])

    filtered = filter_temp_matches_to_selected_temps(matches, selected_temp)

    assert filtered["temp_alarm_source"].tolist() == ["SRC-2"]


def test_filter_matches_to_selected_temp_rows_handles_blank_cleared_time():
    matches = pd.DataFrame([
        {
            "site_id": "A",
            "temp_time": "2026-02-01 11:00:00",
            "temp_cleared": "",
            "temp_alarm_name": "Shelter High Temperature",
            "temp_alarm_source": "SRC-1",
        }
    ])
    selected_temp = pd.DataFrame([
        {
            "site_id": "A",
            "occurred_on": pd.Timestamp("2026-02-01 11:00:00"),
            "cleared_on": pd.NaT,
            "alarm_name": "Shelter High Temperature",
            "alarm_source": "SRC-1",
        }
    ])

    filtered = filter_temp_matches_to_selected_temps(matches, selected_temp)

    assert len(filtered) == 1


def test_query_path_without_result_filter_query_still_loads_query(monkeypatch):
    source_rows = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 13:15:00",
            "cleared_on": "2026-02-01 13:45:00",
            "duration": "00:30:00",
        },
    ])
    monkeypatch.setattr("alarm_app.core.temp_alarm.query_alarms", lambda query: source_rows)

    result, err, source_df = compute_temp_alarm_matches_for_query(AlarmQuery(), margin_minutes=60)

    assert err == ""
    assert len(result) == 1
    assert len(source_df) == 2


def test_date_scoped_temp_excludes_power_more_than_one_day_earlier():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-04 12:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-04 10:00:00",
            "cleared_on": "2026-02-04 11:00:00",
            "duration": "01:00:00",
        },
    ])
    query = AlarmQuery(date_from=pd.Timestamp("2026-02-04"), date_to=pd.Timestamp("2026-02-04"))

    matches, err = compute_temp_alarm_matches(df, margin_minutes=60)
    filtered = filter_temp_matches_to_query(matches, query)

    assert filtered.empty
    assert "No uncovered Temp alarms" in err


def test_temp_source_query_preserves_viewer_scope_for_ht_preview():
    query = AlarmQuery(date_from=pd.Timestamp("2026-02-04"), date_to=pd.Timestamp("2026-02-04"))

    source_query = AlarmViewer._build_temp_alarm_source_query(query)

    assert source_query.date_from == query.date_from
    assert source_query.date_to == query.date_to
    assert source_query.manual_days == query.manual_days


def test_query_path_preserves_viewer_scope_for_ht_power_source(monkeypatch):
    selected_temp = _make_df([
        {
            "site_id": "SITE_A",
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-04 13:30:00",
            "cleared_on": "2026-02-04 14:30:00",
            "duration": "01:00:00",
        }
    ])
    power_rows = _make_df([
        {
            "site_id": "SITE_A",
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-04 12:00:00",
        },
    ])
    calls = []

    def fake_query_alarms(query):
        calls.append(query)
        if len(calls) == 1:
            return selected_temp
        return power_rows

    monkeypatch.setattr("alarm_app.core.temp_alarm.query_alarms", fake_query_alarms)
    selected_query = AlarmQuery(
        date_from=pd.Timestamp("2026-02-04"),
        date_to=pd.Timestamp("2026-02-04"),
        manual_days=[pd.Timestamp("2026-02-04")],
        min_duration_secs=900,
        vendor="HUAWEI",
        network_type="4G",
        allowed_values={"office": ["Cairo"]},
        column_filters={"area": ["North"]},
        col_filters={"region": ["East"]},
    )

    result, err, source_df = compute_temp_alarm_matches_for_query(
        selected_query,
        margin_minutes=60,
        result_filter_query=selected_query,
    )

    assert err == ""
    assert len(result) == 1
    assert len(source_df) == 2
    assert calls[0].category == "Temp"
    assert calls[0].date_from == pd.Timestamp("2026-02-04")
    assert calls[1].category == "Power"
    assert calls[1].site_text == ""
    assert calls[1].vendor == "HUAWEI"
    assert calls[1].network_type == "4G"
    assert calls[1].min_duration_secs == 900
    assert calls[1].allowed_values == {"office": ["Cairo"]}
    assert calls[1].column_filters == {"area": ["North"]}
    assert calls[1].col_filters == {"region": ["East"]}
    assert calls[1].date_from == pd.Timestamp("2026-02-04")
    assert calls[1].date_to == pd.Timestamp("2026-02-04")
    assert list(calls[1].manual_days or []) == [pd.Timestamp("2026-02-04")]
    assert set(calls[1].site_scope_keys) == {"SITE_A"}


def test_query_path_can_return_full_temp_source_for_ht_consolidated(monkeypatch):
    selected_temp = _make_df([
        {
            "site_id": "SITE_A",
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-04 13:30:00",
            "cleared_on": "2026-02-04 22:30:00",
            "duration": "09:00:00",
        }
    ])
    power_rows = _make_df([
        {
            "site_id": "SITE_A",
            "alarm_category": "Power",
            "occurred_on": "2026-02-04 10:00:00",
            "cleared_on": "2026-02-04 11:00:00",
        },
    ])
    historical_temp = _make_df([
        {
            "site_id": "SITE_A",
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2025-12-20 13:30:00",
            "cleared_on": "2025-12-20 22:30:00",
            "duration": "09:00:00",
        },
        {
            "site_id": "SITE_B",
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2025-12-20 13:30:00",
            "cleared_on": "2025-12-20 22:30:00",
            "duration": "09:00:00",
        },
        selected_temp.iloc[0].to_dict(),
    ])
    calls = []

    def fake_query_alarms(query):
        calls.append(query)
        if len(calls) == 1:
            return selected_temp
        if len(calls) == 2:
            return power_rows
        return historical_temp

    monkeypatch.setattr("alarm_app.core.temp_alarm.query_alarms", fake_query_alarms)
    selected_query = AlarmQuery(date_from=pd.Timestamp("2026-02-04"), date_to=pd.Timestamp("2026-02-04"))

    _result, _err, source_df = compute_temp_alarm_matches_for_query(
        selected_query,
        margin_minutes=60,
        result_filter_query=selected_query,
        include_full_temp_source=True,
    )

    assert len(calls) == 3
    assert calls[2].category == "Temp"
    assert set(calls[2].site_scope_keys) == {"SITE_A"}
    assert source_df["occurred_on"].astype(str).str.contains("2025-12-20").any()
    assert "SITE_B" not in set(source_df["site_id"].astype(str))


def test_ht_export_week_allows_year_end_w53_label():
    metadata = ht_export_week_from_date(pd.Timestamp("2028-12-31"))

    assert metadata["week_label"] == "W53-28"
    assert metadata["filename"] == "2028-HT-Alarms-W53.xlsx"


def test_query_path_full_temp_source_filters_selected_sites_when_scope_exceeds_query_cutoff(monkeypatch):
    selected_temp = _make_df([
        {
            "site_id": f"SITE_{idx}",
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-04 13:30:00",
            "cleared_on": "2026-02-04 22:30:00",
            "duration": "09:00:00",
        }
        for idx in range(501)
    ])
    historical_temp = _make_df([
        selected_temp.iloc[0].to_dict(),
        {
            "site_id": "UNSELECTED_SITE",
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2025-12-20 13:30:00",
            "cleared_on": "2025-12-20 22:30:00",
            "duration": "09:00:00",
        },
    ])
    calls = []

    def fake_query_alarms(query):
        calls.append(query)
        if len(calls) == 1:
            return selected_temp
        if len(calls) == 2:
            return _make_df([])
        return historical_temp

    monkeypatch.setattr("alarm_app.core.temp_alarm.query_alarms", fake_query_alarms)

    _result, _err, source_df = compute_temp_alarm_matches_for_query(
        AlarmQuery(date_from=pd.Timestamp("2026-02-04"), date_to=pd.Timestamp("2026-02-04")),
        result_filter_query=AlarmQuery(date_from=pd.Timestamp("2026-02-04"), date_to=pd.Timestamp("2026-02-04")),
        include_full_temp_source=True,
    )

    assert calls[2].site_scope_keys is None
    assert set(source_df["site_id"].astype(str)) == {"SITE_0"}


def test_query_path_full_temp_source_relaxes_duration_filter_for_historical_rows(monkeypatch):
    selected_temp = _make_df([
        {
            "site_id": "SITE_A",
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-04 13:30:00",
            "cleared_on": "2026-02-04 22:30:00",
            "duration": "09:00:00",
        }
    ])
    power_rows = _make_df([
        {
            "site_id": "SITE_A",
            "alarm_category": "Power",
            "occurred_on": "2026-02-04 10:00:00",
            "cleared_on": "2026-02-04 11:00:00",
        },
    ])
    short_historical_temp = _make_df([
        {
            "site_id": "SITE_A",
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2025-12-20 13:30:00",
            "cleared_on": "2025-12-20 14:00:00",
            "duration": "00:30:00",
        },
        selected_temp.iloc[0].to_dict(),
    ])
    calls = []

    def fake_query_alarms(query):
        calls.append(query)
        if len(calls) == 1:
            return selected_temp
        if len(calls) == 2:
            return power_rows
        return short_historical_temp

    monkeypatch.setattr("alarm_app.core.temp_alarm.query_alarms", fake_query_alarms)
    selected_query = AlarmQuery(
        date_from=pd.Timestamp("2026-02-04"),
        date_to=pd.Timestamp("2026-02-04"),
        min_duration_secs=900,
        vendor="HUAWEI",
        network_type="4G",
    )

    _result, _err, source_df = compute_temp_alarm_matches_for_query(
        selected_query,
        margin_minutes=60,
        result_filter_query=selected_query,
        include_full_temp_source=True,
    )

    assert len(calls) == 3
    assert calls[2].category == "Temp"
    assert calls[2].min_duration_secs is None
    assert calls[2].vendor == "All"
    assert calls[2].network_type == "All"
    history_start = ht_export_week_range(DEFAULT_HT_HISTORY_START_WEEK)[0].date()
    assert calls[1].date_from == history_start
    assert calls[2].date_from == history_start
    assert calls[2].date_to == pd.Timestamp("2026-02-04")
    assert source_df["occurred_on"].astype(str).str.contains("2025-12-20").any()
    short_rows = source_df[
        source_df["occurred_on"].astype(str).str.contains("2025-12-20")
    ]
    assert short_rows.iloc[0]["duration"] == "00:30:00"


def test_filter_temps_by_power_clearance_gap_drops_when_gap_within_x():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
            "duration": "02:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "DROP",
            "occurred_on": "2026-02-01 13:01:00",
            "cleared_on": "2026-02-01 14:00:00",
            "duration": "00:59:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "KEEP",
            "occurred_on": "2026-02-01 15:01:00",
            "cleared_on": "2026-02-01 16:00:00",
            "duration": "00:59:00",
        },
    ])
    filtered = filter_temps_by_power_clearance_gap(df, x_secs=7200)
    assert filtered[filtered["alarm_category"] == "Temp"]["alarm_source"].tolist() == ["KEEP"]


def test_filter_temps_by_power_clearance_gap_exact_boundary_drops():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 12:00:00",
            "duration": "02:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "BOUNDARY",
            "occurred_on": "2026-02-01 14:00:00",
            "cleared_on": "2026-02-01 15:00:00",
            "duration": "01:00:00",
        },
    ])
    filtered = filter_temps_by_power_clearance_gap(df, x_secs=7200)
    assert filtered[filtered["alarm_category"] == "Temp"].empty


def test_filter_temps_by_power_clearance_gap_x_off_keeps_all_temps():
    df = _make_df([
        {
            "alarm_category": "Temp",
            "alarm_source": "ANY",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": "2026-02-01 08:50:00",
            "duration": "00:50:00",
        },
    ])
    filtered = filter_temps_by_power_clearance_gap(df, x_secs=None)
    assert filtered["alarm_source"].tolist() == ["ANY"]


def test_filter_temps_by_power_clearance_gap_keeps_power_rows():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": "2026-02-01 08:05:00",
            "duration": "00:05:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SHORT",
            "occurred_on": "2026-02-01 09:00:00",
            "cleared_on": "2026-02-01 09:10:00",
            "duration": "00:10:00",
        },
    ])
    filtered = filter_temps_by_power_clearance_gap(df, x_secs=7200)
    assert filtered["alarm_category"].tolist() == ["Power"]


def test_filter_temps_by_power_clearance_gap_drops_without_prior_power():
    df = _make_df([
        {
            "alarm_category": "Temp",
            "alarm_source": "NO_POWER",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": "2026-02-01 09:00:00",
            "duration": "01:00:00",
        },
    ])
    filtered = filter_temps_by_power_clearance_gap(df, x_secs=7200)
    assert filtered.empty


def test_filter_temps_by_power_clearance_gap_drops_with_uncleared_power():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": None,
            "duration": "",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "UNCLEARED",
            "occurred_on": "2026-02-01 09:00:00",
            "cleared_on": "2026-02-01 10:00:00",
            "duration": "01:00:00",
        },
    ])
    filtered = filter_temps_by_power_clearance_gap(df, x_secs=7200)
    assert filtered[filtered["alarm_category"] == "Temp"].empty


def test_filter_temps_by_power_clearance_gap_uses_latest_cleared_before_temp_with_active_power():
    """When an uncleared Power started after a cleared one, pair on latest cleared_on before Temp."""
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": "2026-02-01 10:00:00",
            "duration": "02:00:00",
        },
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 11:00:00",
            "cleared_on": None,
            "duration": "",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "KEEP",
            "occurred_on": "2026-02-01 13:01:00",
            "cleared_on": "2026-02-01 14:00:00",
            "duration": "00:59:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "DROP",
            "occurred_on": "2026-02-01 10:30:00",
            "cleared_on": "2026-02-01 11:00:00",
            "duration": "00:30:00",
        },
    ])
    filtered = filter_temps_by_power_clearance_gap(df, x_secs=7200)
    assert filtered[filtered["alarm_category"] == "Temp"]["alarm_source"].tolist() == ["KEEP"]


def test_filter_temps_by_power_clearance_gap_keeps_cross_week_power():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-01-21 10:00:00",
            "cleared_on": "2026-01-21 12:00:00",
            "duration": "02:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "CROSS_WEEK",
            "occurred_on": "2026-02-01 15:01:00",
            "cleared_on": "2026-02-01 16:00:00",
            "duration": "00:59:00",
        },
    ])
    filtered = filter_temps_by_power_clearance_gap(df, x_secs=7200, week_label="W05-26")
    assert filtered[filtered["alarm_category"] == "Temp"]["alarm_source"].tolist() == ["CROSS_WEEK"]


def test_compute_ht_meet_rows_applies_clearance_gap_before_meet_rule():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2024-06-30 06:00:00",
            "cleared_on": "2024-06-30 07:00:00",
            "duration": "01:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SHORT_GAP",
            "occurred_on": "2024-06-30 08:00:00",
            "cleared_on": "2024-06-30 17:00:00",
            "duration": "09:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "LONG_GAP",
            "occurred_on": "2024-06-30 10:00:00",
            "cleared_on": "2024-06-30 17:00:00",
            "duration": "07:00:00",
        },
    ])
    _, meet, _ = compute_ht_meet_rows(
        df,
        week_label="W26-24",
        filter_settings=HtWorkbookFilterSettings(clearance_gap_x_secs=7200, apply_meet_threshold=False),
    )
    assert meet["Alarm Source"].tolist() == ["LONG_GAP"]


def test_export_autin_ht_scopes_export_week_after_duration_filter(tmp_path):
    df = _make_df([
        {
            "alarm_category": "Temp",
            "alarm_source": "W06_ROW",
            "occurred_on": "2026-02-01 11:15:00",
            "cleared_on": "2026-02-01 15:15:00",
            "duration": "04:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "OLD_WEEK",
            "occurred_on": "2026-01-21 13:15:00",
            "cleared_on": "2026-01-21 17:15:00",
            "duration": "04:00:00",
        },
    ])
    out = tmp_path / "scoped.xlsx"
    export_temp_alarm_workbook(
        pd.DataFrame(),
        out,
        week_label="W05-26",
        source_df=df,
        filter_settings=_NO_GAP_FILTER,
    )
    wb = load_workbook(out, data_only=True)
    ht_sources = [row[0] for row in wb["W05 AUTIN HT"].iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
    assert ht_sources == ["W06_ROW"]


def test_export_consolidated_uses_historical_filtered_temps(tmp_path):
    df = _make_df([
        {
            "alarm_category": "Temp",
            "alarm_source": "W06_ROW",
            "occurred_on": "2026-02-01 11:15:00",
            "cleared_on": "2026-02-01 19:15:00",
            "duration": "08:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "HIST_SHORT",
            "occurred_on": "2025-10-05 11:15:00",
            "cleared_on": "2025-10-05 11:45:00",
            "duration": "00:30:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "HIST_LONG",
            "occurred_on": "2025-10-12 11:15:00",
            "cleared_on": "2025-10-12 19:15:00",
            "duration": "08:00:00",
        },
    ])
    out = tmp_path / "consolidated.xlsx"
    export_temp_alarm_workbook(
        pd.DataFrame(),
        out,
        week_label="W05-26",
        source_df=df,
        filter_settings=_NO_GAP_FILTER,
    )
    wb = load_workbook(out, data_only=True)
    consolidated = wb["Consolidated"]
    rows = [row for row in consolidated.iter_rows(min_row=2, values_only=True) if row and row[0]]
    week_nos = {row[9] for row in rows}
    assert "W05-26" in week_nos
    assert "W41-25" in week_nos
    assert sum(int(row[5] or 0) for row in rows) == 2


def test_export_consolidated_x_filter_uses_power_before_history_start(tmp_path):
    """Power before W40-22 must remain visible for X-gap pairing on early history Temps."""
    history_start = ht_export_week_range(DEFAULT_HT_HISTORY_START_WEEK)[0]
    power_cleared = history_start - pd.Timedelta(hours=1)
    power_occurred = power_cleared - pd.Timedelta(hours=1)
    temp_occurred = history_start + pd.Timedelta(hours=3)
    temp_cleared = temp_occurred + pd.Timedelta(hours=8)
    df = _make_df([
        {
            "site_id": "S1",
            "alarm_category": "Power",
            "occurred_on": power_occurred,
            "cleared_on": power_cleared,
            "duration": "01:00:00",
        },
        {
            "site_id": "S1",
            "alarm_category": "Temp",
            "alarm_source": "EARLY_HIST_TEMP",
            "occurred_on": temp_occurred,
            "cleared_on": temp_cleared,
            "duration": "08:00:00",
        },
    ])
    export_week = ht_export_week_from_date(temp_occurred)["week_label"]
    out = tmp_path / "consolidated_x.xlsx"
    export_temp_alarm_workbook(
        pd.DataFrame(),
        out,
        week_label=export_week,
        source_df=df,
    )
    wb = load_workbook(out, data_only=True)
    consolidated = wb["Consolidated"]
    target_week = ht_export_week_from_date(temp_occurred)["week_label"]
    week_rows = [
        row
        for row in consolidated.iter_rows(min_row=2, values_only=True)
        if row and row[9] == target_week
    ]
    assert week_rows, f"expected Consolidated row for {target_week}"
    assert any(int(row[5] or 0) >= 1 for row in week_rows)


def test_enrich_source_with_site_metadata_vectorized_matches_row_logic():
    source = _make_df([
        {
            "site_id": "AAA111",
            "site_code": "AAA111",
            "alarm_source": "AAA-111 temp",
            "site_name": "Alarm Name",
        },
        {
            "site_id": "",
            "site_code": "",
            "alarm_source": "BBB222 shelter",
            "site_name": "Fallback",
        },
        {
            "site_id": "MISSING1",
            "alarm_source": "MISSING1 temp",
            "site_name": "No Catalog",
        },
    ])
    metadata = _make_df([
        {"site_id": "AAA111", "site_name": "Catalog Alpha", "area": "North", "contractor": "One"},
        {"site_id": "BBB222", "site_name": "Catalog Beta", "area": "South", "contractor": "Two"},
    ])
    enriched, missing = enrich_source_with_site_metadata(source, metadata)
    assert enriched.loc[0, "site_name"] == "Catalog Alpha"
    assert enriched.loc[1, "site_name"] == "Catalog Beta"
    assert enriched.loc[1, "site_id"] == "BBB222"
    assert set(missing["Site ID"].astype(str)) == {"MISSING1"}


def test_describe_meet_preview_empty_state_distinguishes_gap_filter_from_threshold():
    short_gap = _make_df([
        {
            "site_id": "SITE_A",
            "alarm_category": "Power",
            "occurred_on": "2024-06-30 08:00:00",
            "cleared_on": "2024-06-30 09:30:00",
            "duration": "01:30:00",
        },
        {
            "site_id": "SITE_A",
            "alarm_category": "Temp",
            "alarm_source": "SHORT_GAP",
            "occurred_on": "2024-06-30 10:00:00",
            "cleared_on": "2024-06-30 18:00:00",
            "duration": "08:00:00",
        },
    ])
    gap_message = describe_meet_preview_empty_state(
        short_gap,
        "W26-24",
        filter_settings=HtWorkbookFilterSettings(clearance_gap_x_secs=7200),
    )
    assert "Power-cleared" in gap_message

    below_threshold = _make_df([
        {
            "site_id": "SITE_A",
            "alarm_category": "Temp",
            "alarm_source": "LONG",
            "occurred_on": "2024-06-30 08:00:00",
            "cleared_on": "2024-06-30 17:00:00",
            "duration": "09:00:00",
        },
        {
            "site_id": "SITE_A",
            "alarm_category": "Power",
            "occurred_on": "2024-06-30 08:00:00",
            "cleared_on": "2024-06-30 16:00:00",
            "duration": "08:00:00",
        },
    ])
    threshold_message = describe_meet_preview_empty_state(
        below_threshold,
        "W26-24",
        filter_settings=_NO_GAP_FILTER,
    )
    assert ">7 hour" in threshold_message


def test_build_temp_alarm_preview_enriches_full_source_before_meet_compute(monkeypatch):
    from alarm_app.ui.dialogs import _build_temp_alarm_preview

    source = _make_df([
        {
            "site_id": "AAA111",
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2024-06-30 08:00:00",
            "cleared_on": "2024-06-30 18:00:00",
            "duration": "10:00:00",
        },
        {
            "site_id": "BBB222",
            "alarm_category": "Temp",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2024-07-10 08:00:00",
            "cleared_on": "2024-07-10 18:00:00",
            "duration": "10:00:00",
        },
    ])
    metadata = _make_df([
        {"site_id": "AAA111", "site_name": "Catalog Alpha"},
        {"site_id": "BBB222", "site_name": "Catalog Beta"},
    ])
    seen_sizes = []

    def _spy_enrich(frame, catalog):
        seen_sizes.append(len(frame))
        return enrich_source_with_site_metadata(frame, catalog)

    monkeypatch.setattr("alarm_app.ui.dialogs.enrich_source_with_site_metadata", _spy_enrich)
    precomputed, _reason = _build_temp_alarm_preview(
        source,
        metadata,
        "",
        "W26-24",
        filter_settings=_NO_GAP_FILTER,
    )
    meet = precomputed.meet
    assert seen_sizes == [2]
    assert set(meet["Site Name"]) == {"Catalog Alpha"}


def test_build_temp_alarm_summary_y_filter_keeps_and_drops_site_weeks():
    meet_source = _make_df([
        {
            "alarm_category": "Temp",
            "site_id": "SITE_A",
            "site_code": "SITE_A",
            "site_name": "Alpha",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": "2026-02-01 09:10:00",
            "duration": "01:10:00",
        },
        {
            "alarm_category": "Temp",
            "site_id": "SITE_B",
            "site_code": "SITE_B",
            "site_name": "Beta",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": "2026-02-01 08:45:00",
            "duration": "00:45:00",
        },
        {
            "alarm_category": "Temp",
            "site_id": "SITE_C",
            "site_code": "SITE_C",
            "site_name": "Charlie",
            "occurred_on": "2026-02-01 08:00:00",
            "cleared_on": "2026-02-01 09:01:00",
            "duration": "01:01:00",
        },
    ])
    summary = build_temp_alarm_summary(
        meet_source,
        week_label="W05-26",
        min_ht_duration_secs=3600,
    )
    site_codes = set(summary["Site Code"].tolist())
    assert site_codes == {"SITE_A", "SITE_C"}
    charlie = summary[summary["Site Code"] == "SITE_C"].iloc[0]
    assert charlie["HT Duration"] == "01:01"


def test_compute_ht_meet_rows_7h_toggle_off_includes_gap_passing_temps():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2024-06-30 06:00:00",
            "cleared_on": "2024-06-30 07:00:00",
            "duration": "01:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "FAILS_7H",
            "occurred_on": "2024-06-30 10:00:00",
            "cleared_on": "2024-06-30 16:00:00",
            "duration": "06:00:00",
        },
    ])
    _, meet_on, _ = compute_ht_meet_rows(
        df,
        week_label="W26-24",
        filter_settings=HtWorkbookFilterSettings(clearance_gap_x_secs=None, apply_meet_threshold=True),
    )
    _, meet_off, _ = compute_ht_meet_rows(
        df,
        week_label="W26-24",
        filter_settings=HtWorkbookFilterSettings(clearance_gap_x_secs=None, apply_meet_threshold=False),
    )
    assert meet_on.empty
    assert meet_off["Alarm Source"].tolist() == ["FAILS_7H"]


def test_compute_ht_consolidated_meet_source_matches_full_meet_source():
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2024-06-30 06:00:00",
            "cleared_on": "2024-06-30 07:00:00",
            "duration": "01:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "MEET_ROW",
            "occurred_on": "2024-06-30 08:00:00",
            "cleared_on": "2024-06-30 16:00:00",
            "duration": "08:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "NON_MEET",
            "occurred_on": "2024-06-30 09:00:00",
            "cleared_on": "2024-06-30 12:00:00",
            "duration": "03:00:00",
        },
    ])
    _, _, full_meet_source = compute_ht_meet_rows(
        df,
        week_label=None,
        filter_settings=_NO_GAP_FILTER,
    )
    consolidated = compute_ht_consolidated_meet_source(df, filter_settings=_NO_GAP_FILTER)
    pd.testing.assert_frame_equal(
        consolidated.reset_index(drop=True),
        full_meet_source.reset_index(drop=True),
        check_dtype=False,
    )


def test_export_with_precomputed_matches_full_export(tmp_path):
    df = _make_df([
        {
            "alarm_category": "Power",
            "occurred_on": "2026-02-01 10:00:00",
            "cleared_on": "2026-02-01 10:30:00",
            "duration": "00:30:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SRAN_LWG_TEST_SITE_A_MEET",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 11:15:00",
            "cleared_on": "2026-02-01 15:15:00",
            "duration": "04:00:00",
        },
        {
            "alarm_category": "Temp",
            "alarm_source": "SRAN_LWG_TEST_SITE_A",
            "alarm_name": "Shelter High Temperature",
            "occurred_on": "2026-02-01 16:00:00",
            "cleared_on": "2026-02-01 20:00:00",
            "duration": "04:00:00",
        },
    ])
    settings = _NO_GAP_FILTER
    study, meet, meet_source = compute_ht_meet_rows(
        df,
        week_label="W05-26",
        filter_settings=settings,
        power_sheet="W05 AUTIN Power",
    )
    precomputed = HtWorkbookPrecomputed(
        enriched_source=df,
        filter_text="",
        week_label="W05-26",
        filter_settings=settings,
        study=study,
        meet=meet,
        meet_source=meet_source,
        missing_metadata=pd.DataFrame(columns=["Site ID", "Alarm Source", "Reason"]),
    )
    full_path = tmp_path / "full.xlsx"
    cached_path = tmp_path / "cached.xlsx"
    export_temp_alarm_workbook(pd.DataFrame(), full_path, week_label="W05-26", source_df=df, filter_settings=settings)
    export_temp_alarm_workbook(
        pd.DataFrame(),
        cached_path,
        week_label="W05-26",
        source_df=df,
        filter_settings=settings,
        precomputed=precomputed,
    )
    full_wb = load_workbook(full_path, data_only=False)
    cached_wb = load_workbook(cached_path, data_only=False)
    assert full_wb.sheetnames == cached_wb.sheetnames
    for sheet_name in full_wb.sheetnames:
        full_ws = full_wb[sheet_name]
        cached_ws = cached_wb[sheet_name]
        assert full_ws.max_row == cached_ws.max_row
        assert full_ws.max_column == cached_ws.max_column
        if sheet_name == "W05 AUTIN HT Study" and full_ws.max_row >= 2:
            assert full_ws["D2"].value == cached_ws["D2"].value
            assert full_ws["M2"].value == cached_ws["M2"].value


def test_preview_export_meet_row_parity_with_metadata_filter(tmp_path):
    from alarm_app.ui.dialogs import _build_temp_alarm_preview

    source = _make_df([
        {
            "site_id": "AAA111",
            "site_name": "Alpha Site",
            "alarm_category": "Temp",
            "alarm_source": "ALPHA_TEMP",
            "occurred_on": "2026-02-01 11:15:00",
            "cleared_on": "2026-02-01 19:15:00",
            "duration": "08:00:00",
        },
        {
            "site_id": "BBB222",
            "site_name": "Beta Site",
            "alarm_category": "Temp",
            "alarm_source": "BETA_TEMP",
            "occurred_on": "2026-02-01 11:15:00",
            "cleared_on": "2026-02-01 19:15:00",
            "duration": "08:00:00",
        },
    ])
    filter_text = "Alpha"
    precomputed, _reason = _build_temp_alarm_preview(
        source,
        None,
        filter_text,
        "W05-26",
        filter_settings=_NO_GAP_FILTER,
    )
    out = tmp_path / "filtered_export.xlsx"
    export_temp_alarm_workbook(
        pd.DataFrame(),
        out,
        week_label="W05-26",
        source_df=source,
        filter_settings=_NO_GAP_FILTER,
        precomputed=precomputed,
        metadata_filter_text=filter_text,
    )
    wb = load_workbook(out, data_only=True)
    meet_sources = [
        row[2]
        for row in wb["Meet"].iter_rows(min_row=2, max_col=3, values_only=True)
        if row and row[2]
    ]
    assert len(meet_sources) == len(precomputed.meet)
    assert set(meet_sources) == set(precomputed.meet["Alarm Source"])


@pytest.mark.slow
def test_export_large_workbook_under_baseline(tmp_path):
    import time

    rows = []
    for site_idx in range(200):
        site_id = f"SITE_{site_idx:03d}"
        for day in range(7):
            occurred = pd.Timestamp("2026-05-24") + pd.Timedelta(days=day, hours=8)
            cleared = occurred + pd.Timedelta(hours=9)
            rows.append({
                "site_id": site_id,
                "alarm_category": "Power",
                "occurred_on": occurred - pd.Timedelta(hours=2),
                "cleared_on": occurred - pd.Timedelta(hours=1),
                "duration": "01:00:00",
            })
            rows.append({
                "site_id": site_id,
                "alarm_category": "Temp",
                "alarm_source": f"TEMP_{site_id}_{day}",
                "occurred_on": occurred,
                "cleared_on": cleared,
                "duration": "09:00:00",
            })
    df = _make_df(rows)
    out = tmp_path / "large_export.xlsx"
    started = time.perf_counter()
    result = export_temp_alarm_workbook(
        pd.DataFrame(),
        out,
        week_label="W21-26",
        source_df=df,
        filter_settings=_NO_GAP_FILTER,
        return_warnings=True,
    )
    elapsed = time.perf_counter() - started
    assert out.exists()
    assert result is not None
    assert "stage_timings" in result
    assert elapsed <= 120
