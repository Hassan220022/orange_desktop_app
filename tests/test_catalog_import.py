"""Integration tests for data/catalog_import.py using tiny xlsx fixtures."""

import json
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest

from alarm_app.data.catalog_import import (
    import_bdt_summary_workbook,
    import_network_summary_db_sheet,
)

# ---------------------------------------------------------------------------
# helper: create tiny xlsx files in tmp_path
# ---------------------------------------------------------------------------


def _write_network_summary_xlsx(
    path: Path,
    *,
    include_code: bool = True,
    rows: list[list] | None = None,
) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DB"

    headers = ["Code", "Site  Name", "Area", "Subcontractor", "Battery Type"]
    ws.append(headers)

    data_rows = rows or [
        ["ABC-123", "Alpha Site", "North", "SubCo A", "VRLA"],
        ["XYZ-456", "Beta Site", "South", "SubCo B", "Li-Ion"],
        ["DEF-789", "Gamma Site", "East", "SubCo A", "VRLA"],
    ]
    for r in data_rows:
        ws.append(r)

    if not include_code:
        # replace Code header with something else
        ws["A1"] = "NOT_CODE"
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
            row[0].value = ""

    wb.save(path)


def _write_bdt_summary_xlsx(path: Path, sheets: list[tuple[str, list[list]]] | None = None) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    # remove default sheet
    wb.remove(wb.active)

    if sheets is None:
        sheets = [
            (
                "W27-24",
                [
                    ["Site ID", "Week", "Test Date", "Battery", "Result"],
                    ["ABC", "W27", "2024-07-01", "VRLA", "Pass"],
                    ["XYZ", "W27", "2024-07-02", "Li-Ion", "Pass"],
                ],
            ),
            (
                "W28-24",
                [
                    ["Site ID", "Week", "Test Date", "Battery", "Result"],
                    ["DEF", "W28", "2024-07-08", "VRLA", "Fail"],
                ],
            ),
        ]

    for sheet_name, rows in sheets:
        ws = wb.create_sheet(title=sheet_name)
        for r in rows:
            ws.append(r)

    wb.save(path)


# ---------------------------------------------------------------------------
# fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _temp_db(tmp_path, monkeypatch):
    """Route SQLite and DuckDB to temp paths for all import tests."""
    monkeypatch.setattr("alarm_app.db.engine.STATE_DIR", tmp_path)
    monkeypatch.setattr("alarm_app.db.engine.DB_PATH", tmp_path / "test.db")
    monkeypatch.setattr(
        "alarm_app.data.catalog_store.CATALOG_DB_FILE",
        tmp_path / "catalog_test.duckdb",
    )
    # Also patch get_app_engine so get_shared_session uses the temp db
    monkeypatch.setattr("alarm_app.db.engine._app_engine", None)
    monkeypatch.setattr("alarm_app.db.engine._app_session_factory", None)
    # Ensure tables are created in the fresh temp db
    from alarm_app.db.engine import init_app_db

    init_app_db()


# ---------------------------------------------------------------------------
# Network Summary import tests
# ---------------------------------------------------------------------------


class TestNetworkSummaryImport:
    def test_successful_import(self, tmp_path):
        xlsx_path = tmp_path / "network_summary.xlsx"
        _write_network_summary_xlsx(xlsx_path, include_code=True)

        count = import_network_summary_db_sheet(xlsx_path)
        assert count == 3

        # verify SQLite
        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.repos.catalog_repo import query_site_metadata

        session = get_shared_session()
        try:
            row = query_site_metadata(session, "ABC123")
            assert row is not None
            raw = json.loads(row.raw_data_json)
            assert raw["site_name"] == "Alpha Site"

            # verify original headers preserved
            hdr = json.loads(row.original_headers_json)
            assert "Code" in hdr
            assert hdr["Code"] == "code"

            row2 = query_site_metadata(session, "XYZ456")
            assert row2 is not None
            raw2 = json.loads(row2.raw_data_json)
            assert raw2["area"] == "South"
        finally:
            session.close()

        # verify DuckDB
        from alarm_app.data.catalog_store import query_site_metadata as duckdb_query

        result = duckdb_query("DEF789")
        assert len(result) == 1
        raw3 = json.loads(result.iloc[0]["raw_data_json"])
        assert raw3["area"] == "East"

    def test_import_merges_with_existing_site_metadata_by_site_id(self, tmp_path):
        initial_path = tmp_path / "network_summary_initial.xlsx"
        _write_network_summary_xlsx(initial_path, include_code=True)
        import_network_summary_db_sheet(initial_path)

        update_path = tmp_path / "network_summary_update.xlsx"
        _write_network_summary_xlsx(
            update_path,
            include_code=True,
            rows=[
                ["ABC-123", "Alpha Site Updated", "Central", "SubCo C", "Li-Ion"],
                ["NEW-001", "New Site", "West", "SubCo D", "VRLA"],
            ],
        )

        count = import_network_summary_db_sheet(update_path)

        assert count == 2

        from alarm_app.data.catalog_store import query_site_metadata as duckdb_query
        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.repos.catalog_repo import query_site_metadata

        session = get_shared_session()
        try:
            updated = query_site_metadata(session, "ABC123")
            preserved = query_site_metadata(session, "DEF789")
            inserted = query_site_metadata(session, "NEW001")
            assert updated is not None
            assert preserved is not None
            assert inserted is not None
            assert json.loads(updated.raw_data_json)["site_name"] == "Alpha Site Updated"
            assert json.loads(preserved.raw_data_json)["site_name"] == "Gamma Site"
            assert json.loads(inserted.raw_data_json)["area"] == "West"
        finally:
            session.close()

        assert json.loads(duckdb_query("ABC123").iloc[0]["raw_data_json"])["site_name"] == "Alpha Site Updated"
        assert not duckdb_query("DEF789").empty
        assert not duckdb_query("NEW001").empty

    def test_import_accepts_mixed_date_and_placeholder_columns(self, tmp_path):
        import openpyxl

        xlsx_path = tmp_path / "network_summary_mixed_dates.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DB"
        ws.append(["Code", "Site Name", "On Air Date"])
        ws.append(["ABC-123", "Alpha", datetime(2024, 1, 1)])
        ws.append(["XYZ-456", "Beta", "_"])
        wb.save(xlsx_path)

        count = import_network_summary_db_sheet(xlsx_path)

        from alarm_app.data.catalog_store import query_site_metadata as duckdb_query

        assert count == 2
        assert duckdb_query("ABC123").iloc[0]["on_air_date"] == "2024-01-01 00:00:00"
        assert duckdb_query("XYZ456").iloc[0]["on_air_date"] == "_"

    def test_sqlite_commit_failure_restores_duckdb_snapshot(self, tmp_path, monkeypatch):
        xlsx_path = tmp_path / "network_summary.xlsx"
        _write_network_summary_xlsx(xlsx_path, include_code=True)

        from alarm_app.data.catalog_store import query_site_metadata as duckdb_query
        from alarm_app.data.catalog_store import replace_site_metadata
        from alarm_app.db.engine import get_shared_session as real_get_shared_session

        replace_site_metadata(
            pd.DataFrame([
                {"site_id": "OLD1", "site_name": "Old Site", "original_headers_json": "{}", "raw_data_json": '{"site_id":"OLD1"}'}
            ])
        )
        session = real_get_shared_session()

        def fail_commit():
            raise RuntimeError("commit failed")

        monkeypatch.setattr("alarm_app.db.engine.get_shared_session", lambda: session)
        monkeypatch.setattr(session, "commit", fail_commit)

        with pytest.raises(RuntimeError, match="commit failed"):
            import_network_summary_db_sheet(xlsx_path)

        assert len(duckdb_query("OLD1")) == 1
        assert duckdb_query("ABC123").empty

    def test_missing_code_column_raises(self, tmp_path):
        xlsx_path = tmp_path / "no_code.xlsx"
        _write_network_summary_xlsx(xlsx_path, include_code=False)

        with pytest.raises(ValueError, match="missing 'Code' column"):
            import_network_summary_db_sheet(xlsx_path)

    def test_no_valid_rows_raises(self, tmp_path):
        import openpyxl

        xlsx_path = tmp_path / "empty_db.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DB"
        ws.append(["Code", "Area"])
        # only blank code rows
        ws.append(["", "North"])
        ws.append([None, "South"])
        wb.save(xlsx_path)

        with pytest.raises(ValueError, match="No rows with non-empty Code"):
            import_network_summary_db_sheet(xlsx_path)

    def test_no_db_sheet_raises(self, tmp_path):
        import openpyxl

        xlsx_path = tmp_path / "no_db.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "OtherSheet"
        ws = wb.active
        ws.append(["Col1", "Col2"])
        ws.append(["a", "b"])
        wb.save(xlsx_path)

        with pytest.raises(ValueError, match="No 'DB' sheet found"):
            import_network_summary_db_sheet(xlsx_path)

    def test_read_only_workbook_closed_when_network_import_validation_fails(self, tmp_path, monkeypatch):
        import openpyxl

        xlsx_path = tmp_path / "no_db.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "OtherSheet"
        wb.save(xlsx_path)

        real_load = openpyxl.load_workbook
        opened = {}

        def load_and_track(*args, **kwargs):
            workbook = real_load(*args, **kwargs)
            opened["workbook"] = workbook
            original_close = workbook.close

            def close_and_track():
                opened["closed"] = True
                return original_close()

            workbook.close = close_and_track
            return workbook

        monkeypatch.setattr(openpyxl, "load_workbook", load_and_track)

        with pytest.raises(ValueError, match="No 'DB' sheet found"):
            import_network_summary_db_sheet(xlsx_path)
        assert opened.get("closed") is True

    def test_original_header_mapping_preserved(self, tmp_path):
        xlsx_path = tmp_path / "network_summary.xlsx"
        _write_network_summary_xlsx(xlsx_path, include_code=True)
        import_network_summary_db_sheet(xlsx_path)

        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.repos.catalog_repo import query_site_metadata

        session = get_shared_session()
        try:
            row = query_site_metadata(session, "ABC123")
            hdr = json.loads(row.original_headers_json)
            # all original headers mapped
            assert "Code" in hdr
            assert "Site  Name" in hdr
            assert "Subcontractor" in hdr
            assert "Battery Type" in hdr
            # normalized values (multiple spaces collapse to single underscore)
            assert hdr["Code"] == "code"
            assert hdr["Site  Name"] == "site_name"
            assert hdr["Battery Type"] == "battery_type"
        finally:
            session.close()


# ---------------------------------------------------------------------------
# BDT Summary import tests
# ---------------------------------------------------------------------------


class TestBDTSummaryImport:
    def test_multi_sheet_import(self, tmp_path):
        xlsx_path = tmp_path / "bdt_summary.xlsx"
        _write_bdt_summary_xlsx(xlsx_path)

        periods = import_bdt_summary_workbook(xlsx_path)
        assert set(periods.keys()) == {"W27-24", "W28-24"}
        assert periods["W27-24"] == 2
        assert periods["W28-24"] == 1

        # verify SQLite
        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.models import BDTSummaryCatalog

        session = get_shared_session()
        try:
            all_rows = session.query(BDTSummaryCatalog).all()
            site_ids = {r.site_id for r in all_rows}
            assert site_ids == {"ABC", "XYZ", "DEF"}

            periods_in_db = {r.reporting_period for r in all_rows}
            assert periods_in_db == {"W27-24", "W28-24"}
        finally:
            session.close()

        # verify DuckDB
        from alarm_app.data.catalog_store import query_bdt_summary

        result = query_bdt_summary(reporting_period="W27-24")
        assert len(result) == 2

        result = query_bdt_summary(reporting_period="W28-24")
        assert len(result) == 1

    def test_merge_by_period(self, tmp_path):
        xlsx_path = tmp_path / "bdt_summary.xlsx"
        _write_bdt_summary_xlsx(xlsx_path)

        # first import
        import_bdt_summary_workbook(xlsx_path)

        # create a second workbook that only has W27-24 with different rows
        xlsx2_path = tmp_path / "bdt_summary_v2.xlsx"
        import openpyxl

        wb2 = openpyxl.Workbook()
        wb2.remove(wb2.active)
        ws = wb2.create_sheet("W27-24")
        ws.append(["Site ID", "Week", "Test Date", "Battery", "Result"])
        ws.append(["NEW1", "W27", "2024-07-03", "VRLA", "Pass"])
        ws.append(["NEW2", "W27", "2024-07-04", "Li-Ion", "Fail"])
        wb2.save(xlsx2_path)

        # second import should replace W27-24 but leave W28-24
        periods = import_bdt_summary_workbook(xlsx2_path)
        assert set(periods.keys()) == {"W27-24"}
        assert periods["W27-24"] == 2

        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.models import BDTSummaryCatalog

        session = get_shared_session()
        try:
            all_rows = session.query(BDTSummaryCatalog).all()
            site_ids = {r.site_id for r in all_rows}
            # W27-24 rows replaced (ABC/XYZ gone, NEW1/NEW2 in)
            # W28-24 still has DEF
            assert site_ids == {"NEW1", "NEW2", "DEF"}

            periods_in_db = {r.reporting_period for r in all_rows}
            assert periods_in_db == {"W27-24", "W28-24"}
        finally:
            session.close()

        # DuckDB should match
        from alarm_app.data.catalog_store import query_bdt_summary

        result_w27 = query_bdt_summary(reporting_period="W27-24")
        assert len(result_w27) == 2
        w27_sites = set(result_w27["site_id"])
        assert w27_sites == {"NEW1", "NEW2"}

        result_w28 = query_bdt_summary(reporting_period="W28-24")
        assert len(result_w28) == 1

    def test_week_test_date_test_year_extracted(self, tmp_path):
        xlsx_path = tmp_path / "bdt_date_test.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("P1")
        ws.append(["Site ID", "Week", "Test Date", "Test Year", "Result"])
        ws.append(["S1", "W27", "2024-07-01", 2024, "Pass"])
        wb.save(xlsx_path)

        import_bdt_summary_workbook(xlsx_path)

        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.models import BDTSummaryCatalog

        session = get_shared_session()
        try:
            row = session.query(BDTSummaryCatalog).first()
            assert row is not None
            assert row.week == "W27"
            assert row.test_date is not None
            assert row.test_date.year == 2024
            assert row.test_year == 2024
        finally:
            session.close()

    def test_invalid_test_year_values_do_not_abort_import(self, tmp_path):
        xlsx_path = tmp_path / "bdt_bad_years.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("P1")
        ws.append(["Site ID", "Week", "Test Date", "Test Year", "Result"])
        ws.append(["S1", "W27", "2024-07-01", "202", "Pass"])
        ws.append(["S2", "W27", "2024-07-02", "1e400", "Pass"])
        wb.save(xlsx_path)

        periods = import_bdt_summary_workbook(xlsx_path)

        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.models import BDTSummaryCatalog

        assert periods == {"P1": 2}
        session = get_shared_session()
        try:
            rows = session.query(BDTSummaryCatalog).order_by(BDTSummaryCatalog.site_id).all()
            assert [row.site_id for row in rows] == ["S1", "S2"]
            assert [row.test_year for row in rows] == [None, None]
        finally:
            session.close()

    def test_zero_row_bdt_period_clears_existing_rows(self, tmp_path):
        xlsx_path = tmp_path / "bdt_summary.xlsx"
        _write_bdt_summary_xlsx(
            xlsx_path,
            sheets=[
                (
                    "W27-24",
                    [
                        ["Site ID", "Week", "Test Date", "Result"],
                        ["ABC", "W27", "2024-07-01", "Pass"],
                    ],
                )
            ],
        )
        import_bdt_summary_workbook(xlsx_path)

        clear_path = tmp_path / "bdt_summary_clear.xlsx"
        _write_bdt_summary_xlsx(
            clear_path,
            sheets=[
                (
                    "W27-24",
                    [
                        ["Site ID", "Week", "Test Date", "Result"],
                        [None, "W27", "2024-07-01", "Pass"],
                    ],
                )
            ],
        )

        periods = import_bdt_summary_workbook(clear_path)

        from alarm_app.data.catalog_store import query_bdt_summary
        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.models import BDTSummaryCatalog

        assert periods == {"W27-24": 0}
        assert query_bdt_summary(reporting_period="W27-24").empty
        session = get_shared_session()
        try:
            assert session.query(BDTSummaryCatalog).count() == 0
        finally:
            session.close()

    def test_restore_failure_preserves_original_import_exception(self, tmp_path, monkeypatch):
        xlsx_path = tmp_path / "network_summary.xlsx"
        _write_network_summary_xlsx(xlsx_path, include_code=True)

        from alarm_app.db.engine import get_shared_session as real_get_shared_session

        session = real_get_shared_session()
        calls = {"replace": 0}

        def replace_then_fail_on_restore(_df):
            calls["replace"] += 1
            if calls["replace"] > 1:
                raise RuntimeError("restore failed")
            return len(_df)

        def fail_commit():
            raise RuntimeError("commit failed")

        monkeypatch.setattr("alarm_app.db.engine.get_shared_session", lambda: session)
        monkeypatch.setattr("alarm_app.data.catalog_store.replace_site_metadata", replace_then_fail_on_restore)
        monkeypatch.setattr(session, "commit", fail_commit)

        with pytest.raises(RuntimeError, match="commit failed"):
            import_network_summary_db_sheet(xlsx_path)

    def test_empty_workbook_returns_empty(self, tmp_path):
        xlsx_path = tmp_path / "empty_bdt.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Header1"])  # header only, no data rows
        wb.save(xlsx_path)

        periods = import_bdt_summary_workbook(xlsx_path)
        assert periods == {"Sheet1": 0}

    def test_completely_empty_sheet_records_zero_row_period(self, tmp_path):
        """A sheet with no rows at all (not even headers) is recorded as 0 rows."""
        xlsx_path = tmp_path / "empty_sheet_bdt.xlsx"
        _write_bdt_summary_xlsx(
            xlsx_path,
            sheets=[
                ("W30-24", [["Site ID", "Week", "Test Date", "Result"], ["S1", "W30", "2024-07-15", "Pass"]]),
                ("W31-24", []),  # completely empty — no rows at all
            ],
        )

        periods = import_bdt_summary_workbook(xlsx_path)
        assert set(periods.keys()) == {"W30-24", "W31-24"}
        assert periods["W30-24"] == 1
        assert periods["W31-24"] == 0

        # verify only W30-24 row exists
        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.models import BDTSummaryCatalog

        session = get_shared_session()
        try:
            all_rows = session.query(BDTSummaryCatalog).all()
            assert len(all_rows) == 1
            assert all_rows[0].site_id == "S1"
        finally:
            session.close()

    def test_completely_empty_sheet_clears_previous_period(self, tmp_path):
        """Importing a completely empty sheet for a previously populated period clears stale rows."""
        # populate a period with rows
        xlsx_path = tmp_path / "bdt_populated.xlsx"
        _write_bdt_summary_xlsx(
            xlsx_path,
            sheets=[
                ("W32-24", [["Site ID", "Week", "Test Date", "Result"], ["S1", "W32", "2024-08-01", "Pass"]]),
            ],
        )
        import_bdt_summary_workbook(xlsx_path)

        # now import a completely empty sheet for the same period
        clear_path = tmp_path / "bdt_clear.xlsx"
        _write_bdt_summary_xlsx(
            clear_path,
            sheets=[("W32-24", [])],  # completely empty sheet, no rows
        )

        periods = import_bdt_summary_workbook(clear_path)
        assert periods == {"W32-24": 0}

        from alarm_app.data.catalog_store import query_bdt_summary
        from alarm_app.db.engine import get_shared_session
        from alarm_app.db.models import BDTSummaryCatalog

        assert query_bdt_summary(reporting_period="W32-24").empty
        session = get_shared_session()
        try:
            assert session.query(BDTSummaryCatalog).count() == 0
        finally:
            session.close()

    def test_read_only_workbook_closed_when_bdt_import_empty(self, tmp_path, monkeypatch):
        xlsx_path = tmp_path / "empty_bdt.xlsx"
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Header1"])
        wb.save(xlsx_path)

        real_load = openpyxl.load_workbook
        opened = {}

        def load_and_track(*args, **kwargs):
            workbook = real_load(*args, **kwargs)
            opened["workbook"] = workbook
            original_close = workbook.close

            def close_and_track():
                opened["closed"] = True
                return original_close()

            workbook.close = close_and_track
            return workbook

        monkeypatch.setattr(openpyxl, "load_workbook", load_and_track)

        periods = import_bdt_summary_workbook(xlsx_path)
        assert periods == {"Sheet1": 0}
        assert opened.get("closed") is True
