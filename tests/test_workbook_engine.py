"""Tests for alarm_app.bdt._workbook.WorkbookEngine."""

import pytest

try:
    from alarm_app.bdt._workbook import WorkbookEngine
except ImportError:
    WorkbookEngine = None  # type: ignore[assignment]


@pytest.fixture
def sample_xlsx(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws["A1"] = "Header"
    ws["B1"] = "Value"
    ws["A2"] = "data1"
    ws["B2"] = "data2"
    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return path


def test_repr_shows_filepath(sample_xlsx):
    engine = WorkbookEngine(str(sample_xlsx))
    rep = repr(engine)
    assert "test.xlsx" in rep
    engine.close()


def test_sheet_names_returns_list(sample_xlsx):
    engine = WorkbookEngine(str(sample_xlsx))
    names = engine.sheet_names
    assert isinstance(names, list)
    assert "TestSheet" in names
    engine.close()


def test_sheet_rows_returns_data(sample_xlsx):
    engine = WorkbookEngine(str(sample_xlsx))
    rows = engine.sheet_rows("TestSheet")
    assert rows[0] == ["Header", "Value"]
    assert rows[1] == ["data1", "data2"]
    engine.close()


def test_close_clears_cache(sample_xlsx):
    engine = WorkbookEngine(str(sample_xlsx))
    engine.sheet_rows("TestSheet")
    engine.close()
    with pytest.raises(RuntimeError, match="closed"):
        _ = engine.sheet_names


def test_context_manager_closes(sample_xlsx):
    with WorkbookEngine(str(sample_xlsx)) as engine:
        _ = engine.sheet_rows("TestSheet")
    # after exit the engine is closed
    with pytest.raises(RuntimeError, match="closed"):
        _ = engine.sheet_names


def test_calamine_fallback_promotes_to_openpyxl(monkeypatch, sample_xlsx):
    import python_calamine

    def mock_from_path(_path):
        raise RuntimeError("calamine unavailable")

    monkeypatch.setattr(
        python_calamine.CalamineWorkbook, "from_path", staticmethod(mock_from_path)
    )

    engine = WorkbookEngine(str(sample_xlsx))
    rows = engine.sheet_rows("TestSheet")
    assert engine.engine_used == "openpyxl"
    assert rows[0] == ["Header", "Value"]
    assert rows[1] == ["data1", "data2"]
    engine.close()


def test_load_workbook_fn_injection(monkeypatch, sample_xlsx):
    import python_calamine

    def mock_from_path(_path):
        raise RuntimeError("calamine unavailable")

    monkeypatch.setattr(
        python_calamine.CalamineWorkbook, "from_path", staticmethod(mock_from_path)
    )

    called = False

    def custom_loader(path, **kwargs):
        nonlocal called
        called = True
        from openpyxl import load_workbook

        return load_workbook(path, **kwargs)

    engine = WorkbookEngine(str(sample_xlsx), load_workbook_fn=custom_loader)
    _ = engine.sheet_rows("TestSheet")
    assert called
    assert engine.engine_used == "openpyxl"
    engine.close()


def test_repeated_sheet_rows_uses_cache(monkeypatch, sample_xlsx):
    import python_calamine

    call_count = 0
    original_from_path = python_calamine.CalamineWorkbook.from_path

    def counting_from_path(path):
        nonlocal call_count
        call_count += 1
        return original_from_path(path)

    monkeypatch.setattr(
        python_calamine.CalamineWorkbook,
        "from_path",
        staticmethod(counting_from_path),
    )

    engine = WorkbookEngine(str(sample_xlsx))
    engine.sheet_rows("TestSheet")
    engine.sheet_rows("TestSheet")
    assert call_count == 1
    engine.close()
