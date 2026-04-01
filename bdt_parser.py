"""
BDT Parser — extract structured data from Battery Discharge Test Excel files.

BDT files have a non-tabular layout with multiple sections in a single sheet.
Data is extracted by known cell positions (row, col) based on the standard
BDT template.
"""

from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


@dataclass
class PhotoSlot:
    """One labelled photo placeholder in the BDT template."""
    label: str                       # e.g. "Battery current", "PLVD set point"
    image_data: bytes | None = None  # raw JPEG/PNG bytes, None if empty
    image_ext: str = ""              # "jpeg" or "png"


@dataclass
class BDTData:
    """Structured data extracted from one BDT file."""
    file_path: str = ""
    filename: str = ""
    site_code: str = ""
    site_name: str = ""
    test_date: datetime | None = None
    time_in: str = ""
    time_out: str = ""

    # Discharge readings: list of (time_label, voltage, ampere)
    discharge_readings: list[tuple[str, float | None, float | None]] = field(
        default_factory=list)
    start_voltage: float | None = None
    start_ampere: float | None = None
    end_voltage: float | None = None
    end_ampere: float | None = None
    after_reconnect_voltage: float | None = None
    after_reconnect_ampere: float | None = None
    discharge_minutes: float = 0.0

    # Battery info
    ibat_before_test: float | None = None
    battery_brand: str = ""
    battery_ah: float | None = None
    battery_voltage: float | None = None
    num_strings: int | None = None

    # Photos
    photo_count: int = 0
    photo_slots: list[PhotoSlot] = field(default_factory=list)
    photos_deferred: bool = False

    # Parse errors
    errors: list[str] = field(default_factory=list)


def _safe_float(val) -> float | None:
    """Try to convert a cell value to float, return None on failure."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        s = str(val).strip().replace(",", ".")
        # Strip units like "AH", "AM", "V"
        for suffix in ("ah", "am", "a", "v", "vdc"):
            if s.lower().endswith(suffix):
                s = s[:-len(suffix)].strip()
        return float(s)
    except (ValueError, TypeError):
        return None


def _safe_str(val) -> str:
    """Convert cell value to stripped string, empty on None/NaN."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


_BRAND_KEYWORDS = (
    "lithium", "lith", "huawei", "zte", "narada", "shoto",
    "sacred sun", "ritar", "vision", "coslight", "byd",
    "pylontech", "gel", "agm", "vrla", "opzv", "opzs",
)


def _parse_battery_info(max_column, cell_fn, data: BDTData):
    """Extract battery specs from the BDT sheet.

    Standard BDT template layout (col 1 = label, col 9 = value):
        row 40: Battery brand
        row 42: Number of batteries connected to the rectifier
        row 44: Battery nominal voltage
        row 46: Battery ampere hour
        row 48: Number of strings

    Falls back to keyword scanning rows 35-65 if fixed positions don't match.
    """
    brand_raw = ""
    ah_raw = None
    voltage_raw = None
    strings_raw = None

    # ── Fixed-position extraction — standard BDT template ──
    candidate = _safe_str(cell_fn(40, 9))
    if candidate:
        brand_raw = candidate

    parsed = _safe_float(cell_fn(44, 9))
    if parsed is not None and parsed > 0:
        voltage_raw = parsed

    parsed = _safe_float(cell_fn(46, 9))
    if parsed is not None and parsed > 0:
        ah_raw = parsed

    parsed = _safe_float(cell_fn(48, 9))
    if parsed is not None and parsed > 0:
        strings_raw = int(parsed)

    # ── Keyword-based scan (rows 35-65) as fallback ──
    for r in range(35, 66):
        val = _safe_str(cell_fn(r, 1)).lower()
        if not val:
            val = _safe_str(cell_fn(r, 2)).lower()
        if not val:
            continue

        # Battery brand
        if not brand_raw and "battery brand" in val:
            candidate = _safe_str(cell_fn(r, 9))
            if candidate:
                brand_raw = candidate

        # Nominal voltage
        if voltage_raw is None and "nominal voltage" in val:
            parsed = _safe_float(cell_fn(r, 9))
            if parsed is not None and parsed > 0:
                voltage_raw = parsed

        # AH
        if ah_raw is None and "ampere hour" in val:
            parsed = _safe_float(cell_fn(r, 9))
            if parsed is not None and parsed > 0:
                ah_raw = parsed

        # Number of strings
        if strings_raw is None and "number of string" in val:
            parsed = _safe_float(cell_fn(r, 9))
            if parsed is not None and parsed > 0:
                strings_raw = int(parsed)

    # ── Brand detection from known manufacturer keywords ──
    if brand_raw:
        data.battery_brand = brand_raw
    else:
        # Try to find brand by scanning battery section for known names
        for r in range(35, 66):
            for c in (1, 2, 9):
                val = _safe_str(cell_fn(r, c)).lower()
                for kw in _BRAND_KEYWORDS:
                    if kw in val:
                        data.battery_brand = _safe_str(cell_fn(r, c))
                        break
                if data.battery_brand:
                    break
            if data.battery_brand:
                break

    data.battery_ah = ah_raw
    data.battery_voltage = voltage_raw
    data.num_strings = strings_raw


import re as _re

_DATE_IN_FILENAME = _re.compile(r"(\d{1,2})-(\d{1,2})-(\d{4})")


def _parse_test_date(cell_val, filename: str) -> datetime | None:
    """Coerce a cell value to a date, falling back to filename extraction.

    Calamine may return datetime.date, datetime.datetime, datetime.time,
    or garbage strings.
    """
    import datetime as _dt

    # datetime.date (calamine returns this for date-only cells)
    if isinstance(cell_val, _dt.date) and not isinstance(cell_val, datetime):
        if cell_val.year > 1900:
            return datetime(cell_val.year, cell_val.month, cell_val.day)

    # Already a proper datetime with a real year
    if isinstance(cell_val, datetime) and cell_val.year > 1900:
        return cell_val

    # String — try common formats
    if isinstance(cell_val, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(cell_val.strip(), fmt)
            except ValueError:
                continue

    # Fallback: extract date from filename (e.g. "14-1-2026")
    m = _DATE_IN_FILENAME.search(filename)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    return None


def parse_bdt_file(file_path: str, *, skip_photos: bool = False) -> BDTData:
    """Parse a single BDT Excel file and return structured data.

    Uses python-calamine (Rust-based reader, ~100x faster than openpyxl)
    for cell data extraction.  Falls back to openpyxl if calamine fails.

    Args:
        skip_photos: If True, skip photo extraction for faster batch
            processing.  Photos can be loaded later via
            ``load_bdt_photos()``.
    """
    import os
    data = BDTData(file_path=file_path, filename=os.path.basename(file_path))

    # ── Read sheet data with calamine (fast path) ────────
    rows = None
    try:
        import python_calamine
        wb = python_calamine.CalamineWorkbook.from_path(file_path)
        if "BDT sheet" not in wb.sheet_names:
            data.errors.append("Missing 'BDT sheet'")
            return data
        rows = wb.get_sheet_by_name("BDT sheet").to_python()
    except Exception:
        pass

    # ── Fallback to openpyxl whenever calamine didn't yield rows ──
    if rows is None:
        try:
            owb = load_workbook(file_path, data_only=True)
        except Exception as e:
            data.errors.append(f"Cannot open file: {e}")
            return data
        if "BDT sheet" not in owb.sheetnames:
            data.errors.append("Missing 'BDT sheet'")
            owb.close()
            return data
        ows = owb["BDT sheet"]
        rows = []
        for row_cells in ows.iter_rows(min_row=1, max_row=ows.max_row,
                                        max_col=ows.max_column):
            rows.append([c.value for c in row_cells])
        owb.close()

    if rows is None:
        if not data.errors:
            data.errors.append("Failed to read BDT sheet")
        return data

    max_row = len(rows)
    max_col = max((len(r) for r in rows), default=0)

    def cell(row, col):
        """1-indexed cell access (matches Excel row/col numbers)."""
        r, c = row - 1, col - 1
        if r < 0 or r >= max_row:
            return None
        row_data = rows[r]
        if c < 0 or c >= len(row_data):
            return None
        return row_data[c]

    # Site info
    data.site_name = _safe_str(cell(4, 3))
    data.site_code = _safe_str(cell(4, 9))
    data.test_date = _parse_test_date(cell(3, 15), data.filename)
    data.time_in  = _safe_str(cell(4, 15))
    data.time_out = _safe_str(cell(5, 15))

    # Discharge test table — find by scanning for "Batteries discharge test"
    discharge_start_row = None
    for r in range(1, max_row + 1):
        v = _safe_str(cell(r, 2))
        if "batteries discharge test" in v.lower():
            discharge_start_row = r
            break

    if discharge_start_row:
        # Scan for "Before disconnecting Rectifier" row
        data_row = None
        for r in range(discharge_start_row + 1, discharge_start_row + 10):
            lbl = _safe_str(cell(r, 1)).lower()
            if "before disconnecting" in lbl:
                data_row = r
                break
        if data_row is None:
            data_row = discharge_start_row + 3

        data.start_voltage = _safe_float(cell(data_row, 4))
        data.start_ampere  = _safe_float(cell(data_row, 5))

        string_amps = []
        for sc in range(7, 22, 2):
            sa = _safe_float(cell(data_row, sc))
            if sa is not None:
                string_amps.append(sa)
        if string_amps:
            data.ibat_before_test = max(string_amps)

        # Discharge time-series
        last_filled_mins = 0.0
        last_filled_voltage = None
        last_filled_ampere = None
        r = data_row + 1
        while r <= data_row + 30:
            lbl = _safe_str(cell(r, 1))
            if "after connecting" in lbl.lower():
                data.after_reconnect_voltage = _safe_float(cell(r, 4))
                data.after_reconnect_ampere  = _safe_float(cell(r, 5))
                break
            if not lbl:
                r += 1
                continue
            v = _safe_float(cell(r, 4))
            a = _safe_float(cell(r, 5))
            data.discharge_readings.append((lbl, v, a))
            if v is not None:
                last_filled_voltage = v
            if a is not None:
                last_filled_ampere = a
            if v is not None or a is not None:
                try:
                    last_filled_mins = float(lbl.split()[0])
                except ValueError:
                    pass
            r += 1

        data.discharge_minutes = last_filled_mins
        data.end_voltage = last_filled_voltage
        data.end_ampere  = last_filled_ampere

    _parse_battery_info(max_col, cell, data)

    # Photo slots
    if not skip_photos:
        data.photo_slots, data.photo_count = _extract_photo_slots(file_path)
        data.photos_deferred = False
    else:
        data.photos_deferred = True

    return data


def load_bdt_photos(bdt: BDTData) -> None:
    """Lazy-load photo slots for a BDTData that was parsed with skip_photos."""
    if bdt.photo_slots:
        bdt.photos_deferred = False
        return  # already loaded
    bdt.photo_slots, bdt.photo_count = _extract_photo_slots(bdt.file_path)
    bdt.photos_deferred = False


# ── Photo slot definitions ────────────────────────────────────
# Each slot: (label_row_1indexed, label_col_1indexed)
# 5 row bands × 3 column groups = 15 photo placeholders.
# Row bands (0-indexed anchor rows): 9-19, 21-32, 34-44, 46-56, 58-68
# Column groups (0-indexed): 12-16, 17-21, 22-26
_SLOT_DEFS: list[tuple[int, int]] = [
    # Band 0 — Rectifier
    (9, 13), (9, 18), (9, 23),
    # Band 1 — Batteries
    (21, 13), (21, 18), (21, 23),
    # Band 2 — CBs / Rack / LVD
    (34, 13), (34, 18), (34, 23),
    # Band 3 — Current / Load / PLVD
    (46, 13), (46, 18), (46, 23),
    # Band 4 — Charging / Disconnect / Reconnect
    (58, 13), (58, 18), (58, 23),
]

# Map (band_index, col_group_index) → slot index in _SLOT_DEFS
_BAND_RANGES = [(9, 19), (21, 32), (34, 44), (46, 56), (58, 68)]
_COL_GROUPS = [(12, 16), (17, 21), (22, 26)]


def _anchor_to_slot(from_row: int, from_col: int) -> int | None:
    """Map a 0-indexed anchor position to a slot index (0-14), or None."""
    band = None
    for bi, (lo, hi) in enumerate(_BAND_RANGES):
        if lo <= from_row <= hi:
            band = bi
            break
    if band is None:
        return None

    col_grp = None
    for ci, (lo, hi) in enumerate(_COL_GROUPS):
        if lo <= from_col <= hi:
            col_grp = ci
            break
    if col_grp is None:
        return None

    return band * 3 + col_grp


def _extract_photo_slots(file_path: str) -> tuple[list[PhotoSlot], int]:
    """Extract labelled photo slots from a BDT xlsx file.

    Returns (list_of_PhotoSlot, total_media_count).
    """
    import zipfile
    import xml.etree.ElementTree as ET

    ns_xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"
    ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_rel = "http://schemas.openxmlformats.org/package/2006/relationships"

    try:
        zf = zipfile.ZipFile(file_path)
    except Exception:
        return [], 0

    try:
        # Total media count (for backwards compat)
        media_files = [n for n in zf.namelist() if n.startswith("xl/media/")]
        total_media = len(media_files)

        # Find all drawing XML files in the zip (not just drawing1.xml)
        all_names = zf.namelist()
        drawing_paths = sorted(
            n for n in all_names
            if n.startswith("xl/drawings/") and n.endswith(".xml")
            and "/_rels/" not in n
        )
        if not drawing_paths:
            return [], total_media

        # Build rId → media zip path map from all drawing rels
        rid_to_path: dict[str, str] = {}
        for dp in drawing_paths:
            dp_basename = dp.rsplit("/", 1)[-1]
            rels_path = f"xl/drawings/_rels/{dp_basename}.rels"
            if rels_path in all_names:
                rels_xml = ET.fromstring(zf.read(rels_path))
                for rel in rels_xml:
                    rid = rel.get("Id", "")
                    target = rel.get("Target", "")
                    if target.startswith("../media/"):
                        rid_to_path[rid] = "xl/media/" + target.split("/")[-1]

        # Parse all drawing XMLs — extract twoCellAnchor positions + rIds
        # slot_index → (rId, media_path)
        slot_images: dict[int, tuple[str, str]] = {}

        for dp in drawing_paths:
            drawing_xml = ET.fromstring(zf.read(dp))
            for anchor in drawing_xml.findall(f"{{{ns_xdr}}}twoCellAnchor"):
                frm = anchor.find(f"{{{ns_xdr}}}from")
                if frm is None:
                    continue
                from_col = int(frm.find(f"{{{ns_xdr}}}col").text)
                from_row = int(frm.find(f"{{{ns_xdr}}}row").text)

                # Skip non-photo anchors (logo at col 0, etc.)
                if from_col < 12:
                    continue

                slot_idx = _anchor_to_slot(from_row, from_col)
                if slot_idx is None or slot_idx in slot_images:
                    continue

                # Find embedded image rId
                blip = anchor.find(f".//{{{ns_a}}}blip")
                if blip is None:
                    continue
                rid = blip.get(f"{{{ns_r}}}embed", "")
                if not rid or rid not in rid_to_path:
                    continue
                slot_images[slot_idx] = (rid, rid_to_path[rid])

        # Read labels from the worksheet and build PhotoSlot list
        # Re-open with openpyxl just for label reading
        wb = load_workbook(file_path, data_only=True)
        ws = wb["BDT sheet"] if "BDT sheet" in wb.sheetnames else None

        slots: list[PhotoSlot] = []
        for idx, (label_row, label_col) in enumerate(_SLOT_DEFS):
            # Read label from worksheet
            label = ""
            if ws is not None:
                raw = ws.cell(row=label_row, column=label_col).value
                label = _safe_str(raw) if raw else ""
            if not label:
                label = f"Slot {idx + 1}"

            img_data = None
            img_ext = ""
            if idx in slot_images:
                _, media_path = slot_images[idx]
                try:
                    img_data = zf.read(media_path)
                    img_ext = media_path.rsplit(".", 1)[-1].lower()
                except Exception:
                    pass

            slots.append(PhotoSlot(
                label=label, image_data=img_data, image_ext=img_ext))

        wb.close()
        return slots, total_media

    finally:
        zf.close()
