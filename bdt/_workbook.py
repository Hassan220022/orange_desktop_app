"""WorkbookEngine — single-open, calamine-first Excel reader for BDT parsing.

Replaces the duplicated calamine → openpyxl fallback chains scattered across
bdt/parser.py and friends. Opens the file once and reuses the loaded workbook
for sheet-name queries and row data extraction.

Failures of the calamine fast path are logged at WARNING level so operators
can see why a file is slow, then the engine transparently promotes itself
to openpyxl. The first calamine sheet-read failure promotes the *whole*
workbook so subsequent reads use openpyxl too.
"""

from __future__ import annotations

import logging
from typing import Any, Callable

logger = logging.getLogger(__name__)


def _default_load_workbook(path: str, **kwargs: Any) -> Any:
    """Default openpyxl loader.

    Imported lazily so the engine module doesn't pull openpyxl in until
    the fallback is actually needed.
    """
    from openpyxl import load_workbook
    return load_workbook(path, **kwargs)


class WorkbookEngine:
    """Read Excel workbooks via calamine (Rust, fast) with openpyxl fallback.

    Opens lazily on first access.  Use as a context manager to guarantee
    cleanup::

        with WorkbookEngine(path) as wb:
            for name in wb.sheet_names:
                rows = wb.sheet_rows(name)

    Sheet rows are cached after the first read so callers can request the
    same sheet multiple times without paying the parse cost twice.
    """

    __slots__ = (
        "_path",
        "_calamine_wb",
        "_openpyxl_wb",
        "_engine_used",
        "_closed",
        "_rows_cache",
        "_load_workbook_fn",
    )

    def __init__(
        self,
        file_path: str,
        *,
        load_workbook_fn: Callable[..., Any] | None = None,
    ) -> None:
        """Create an engine for *file_path*.

        Args:
            file_path: Absolute path to the .xlsx/.xlsm/.xls file.
            load_workbook_fn: Optional override for the openpyxl loader.
                Useful for tests and for callers that want their own
                ``load_workbook`` reference patched.  Defaults to
                ``openpyxl.load_workbook``.
        """
        self._path: str = file_path
        self._calamine_wb: Any = None
        self._openpyxl_wb: Any = None
        self._engine_used: str | None = None  # "calamine" or "openpyxl"
        self._closed: bool = False
        self._rows_cache: dict[str, list[list[Any]]] = {}
        self._load_workbook_fn: Callable[..., Any] = (
            load_workbook_fn or _default_load_workbook
        )

    # ── context manager ──────────────────────────────────────────

    def __enter__(self) -> WorkbookEngine:
        return self

    def __exit__(self, *args: object) -> None:
        self.close()

    def __repr__(self) -> str:
        return (
            f"WorkbookEngine({self._path!r}, engine={self._engine_used or 'unopened'})"
        )

    # ── properties ───────────────────────────────────────────────

    @property
    def engine_used(self) -> str | None:
        return self._engine_used

    @property
    def sheet_names(self) -> list[str]:
        self._ensure_open()
        if self._engine_used == "calamine":
            return list(self._calamine_wb.sheet_names)  # type: ignore[union-attr]
        return list(self._openpyxl_wb.sheetnames)  # type: ignore[union-attr]

    # ── data access ──────────────────────────────────────────────

    def sheet_rows(self, sheet_name: str) -> list[list[Any]]:
        """Return all rows of *sheet_name* as a list of row-lists.

        Each row is a list of cell values (None for empty cells).
        Row 0 corresponds to Excel row 1.

        Cached after the first call for the same sheet.  If calamine
        succeeds at opening the workbook but fails to read this sheet,
        the engine promotes itself to openpyxl mid-stream and retries.
        """
        self._ensure_open()
        if sheet_name in self._rows_cache:
            return self._rows_cache[sheet_name]

        rows: list[list[Any]]
        if self._engine_used == "calamine":
            try:
                rows = self._calamine_wb.get_sheet_by_name(  # type: ignore[union-attr]
                    sheet_name
                ).to_python()
            except Exception:
                logger.warning(
                    "calamine sheet_rows failed for %s/%s; promoting to openpyxl",
                    self._path,
                    sheet_name,
                    exc_info=True,
                )
                self._promote_to_openpyxl()
                rows = self._openpyxl_sheet_rows(sheet_name)
        else:
            rows = self._openpyxl_sheet_rows(sheet_name)

        self._rows_cache[sheet_name] = rows
        return rows

    # ── internal ─────────────────────────────────────────────────

    def _ensure_open(self) -> None:
        if self._engine_used is not None:
            return
        if self._closed:
            raise RuntimeError("WorkbookEngine is closed")

        # 1st attempt: calamine (Rust, ~100× faster)
        try:
            import python_calamine  # noqa: F811

            self._calamine_wb = python_calamine.CalamineWorkbook.from_path(self._path)
            self._engine_used = "calamine"
            return
        except Exception:
            logger.warning(
                "calamine failed for %s — falling back to openpyxl",
                self._path,
                exc_info=True,
            )

        # 2nd attempt: openpyxl (pure Python, full parse)
        self._openpyxl_wb = self._load_workbook_fn(self._path, data_only=True)
        self._engine_used = "openpyxl"

    def _promote_to_openpyxl(self) -> None:
        """Switch from calamine to openpyxl mid-stream and drop cached rows.

        Cached rows came from calamine; we drop them so subsequent reads
        come from openpyxl consistently.  Sheet names cached on calamine
        are also forgotten — the next ``sheet_names`` call rereads them
        from the openpyxl workbook.
        """
        self._openpyxl_wb = self._load_workbook_fn(self._path, data_only=True)
        self._engine_used = "openpyxl"
        self._calamine_wb = None
        self._rows_cache.clear()

    def _openpyxl_sheet_rows(self, sheet_name: str) -> list[list[Any]]:
        ows = self._openpyxl_wb[sheet_name]  # type: ignore[index]
        if ows.max_row is None or ows.max_column is None:
            return []
        rows: list[list[Any]] = []
        for row_cells in ows.iter_rows(
            min_row=1, max_row=ows.max_row, max_col=ows.max_column
        ):
            rows.append([c.value for c in row_cells])
        return rows

    def close(self) -> None:
        if self._closed:
            return
        self._calamine_wb = None
        if self._openpyxl_wb is not None:
            try:
                self._openpyxl_wb.close()
            except Exception:
                pass
            self._openpyxl_wb = None
        self._engine_used = None
        self._closed = True
        self._rows_cache.clear()
