"""
BDT Parser — extract structured data from Battery Discharge Test Excel files.

BDT files have a non-tabular layout with multiple sections in a single sheet.
Data is extracted by known cell positions (row, col) based on the standard
BDT template.
"""

import logging
import hashlib
from dataclasses import dataclass, field
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class PhotoSlot:
    """One labelled photo placeholder in the BDT template."""
    label: str                       # e.g. "Battery current", "PLVD set point"
    image_data: bytes | None = None  # raw JPEG/PNG bytes, None if empty
    image_ext: str = ""              # "jpeg" or "png"
    category: str = "other"


@dataclass
class BDTData:
    """Structured data extracted from one BDT file."""
    file_path: str = ""
    filename: str = ""
    site_code: str = ""
    site_name: str = ""
    test_date: datetime | None = None
    time_in: str = ""
    time_out: str = ""

    # Discharge readings: list of (time_label, voltage, ampere)
    discharge_readings: list[tuple[str, float | None, float | None]] = field(
        default_factory=list)
    start_voltage: float | None = None
    start_ampere: float | None = None
    end_voltage: float | None = None
    end_ampere: float | None = None
    after_reconnect_voltage: float | None = None
    after_reconnect_ampere: float | None = None
    discharge_minutes: float = 0.0

    # Battery info
    ibat_before_test: float | None = None
    starting_ibattery_ampere: float | None = None
    battery_brand: str = ""
    battery_model: str = ""
    battery_ah: float | None = None
    battery_voltage: float | None = None
    num_strings: int | None = None
    num_batteries: int | None = None

    # Additional site and rectifier info
    power_source: str = ""
    site_category: str = ""
    site_type: str = ""
    rectifier_capacity: float | None = None

    # Core layout detection result (A or B)
    core_layout: str = ""  # "Layout A" or "Layout B"
    num_modules: int | None = None
    rectifier_brand: str = ""
    pld_value: str = ""
    string_discharge_readings: list[list[tuple[float | None, float | None]]] = field(default_factory=list)
    summary_data: dict[str, str] = field(default_factory=dict)
    door_alarm_condition: bool | None = None

    # Photos
    photo_count: int = 0
    photo_slots: list[PhotoSlot] = field(default_factory=list)
    photo_layout_id: str = "LAYOUT_PHOTO_16"
    required_photo_count: int = 16
    photos_deferred: bool = False

    # Layout family detection metadata (FR-001, FR-002)
    core_layout_family: str = ""  # "A", "B1", "B2", "C", "UNKNOWN", "SUMMARY_EXCLUDED"
    detection_confidence: str = ""  # "high", "medium", "low"
    detection_reasons: list[str] = field(default_factory=list)

    # Photo category mapping metadata (FR-003, FR-004)
    photo_categories_found: list[str] = field(default_factory=list)
    photo_mapping_confidence: str = ""  # "high", "medium", "low"
    photo_detection_mode: str = ""  # "structural", "deferred"
    required_photo_categories: list[str] = field(default_factory=list)
    parser_mode: str = ""  # "fast_family" or "structural"
    structural_signature: str = ""

    # Parse errors
    errors: list[str] = field(default_factory=list)


def _safe_float(val) -> float | None:
    """Try to convert a cell value to float, return None on failure."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        s = str(val).strip().replace(",", ".")
        # Strip units like "AH", "AM", "V"
        for suffix in ("ah", "am", "a", "v", "vdc"):
            if s.lower().endswith(suffix):
                s = s[:-len(suffix)].strip()
        return float(s)
    except (ValueError, TypeError):
        return None


def _safe_str(val) -> str:
    """Convert cell value to stripped string, empty on None/NaN."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in {"nan", "none", "null", "unknown", "n/a", "na", "-", "--"}:
        return ""
    return s


def _has_meaningful_numeric(*vals: float | None, zero_eps: float = 0.01) -> bool:
    """True when at least one numeric value is meaningfully non-zero."""
    for v in vals:
        if v is None:
            continue
        try:
            if abs(float(v)) > zero_eps:
                return True
        except (TypeError, ValueError):
            continue
    return False


def _find_text_in_row_window(cell_fn, max_col: int, row_start: int, row_end: int,
                             needles: tuple[str, ...], search_cols: int | None = None):
    """Return (row, col) of the first cell containing all needle tokens."""
    end_col = min(max_col, search_cols) if search_cols else max_col
    for r in range(max(1, row_start), max(1, row_end) + 1):
        for c in range(1, end_col + 1):
            txt = _safe_str(cell_fn(r, c)).lower()
            if txt and all(n in txt for n in needles):
                return r, c
    return None, None


def _extract_row_label(cell_fn, row: int, preferred_col: int | None,
                       max_scan_col: int = 6) -> str:
    """Read a row label from preferred column, then fallback scan of first columns."""
    if preferred_col is not None:
        label = _safe_str(cell_fn(row, preferred_col))
        if label:
            return label
    for c in range(1, max_scan_col + 1):
        label = _safe_str(cell_fn(row, c))
        if label:
            return label
    return ""


_BRAND_KEYWORDS = (
    "lithium", "lith", "huawei", "zte", "narada", "shoto",
    "sacred sun", "ritar", "vision", "coslight", "byd",
    "pylontech", "gel", "agm", "vrla", "opzv", "opzs",
)


_BRAND_LABEL_KEYWORDS = ("battery brand", "battery type", "brand", "make",
                         "manufacturer")
_VOLTAGE_LABEL_KEYWORDS = ("nominal voltage", "nominal volt", "voltage",
                           "volt")
_AH_LABEL_KEYWORDS = ("ampere hour", "ampere-hour", "capacity", "ah",
                       "battery capacity")
_STRINGS_LABEL_KEYWORDS = ("number of string", "number of strings",
                           "no. of string", "strings", "string count")

# ── Layout cell-position maps ─────────────────────────────────
# All positions are 1-indexed (row, col) matching Excel coordinates.
# Layout A = real production template confirmed on 11 files.
# Layout B = older template variant (fallback).

_LAYOUT_A = {
    "site_name":      (4,  3),   # C4
    "site_code":      (4, 12),   # L4
    "test_date":      (3, 20),   # T3
    "time_in":        (5, 21),   # U5
    "time_out":       (6, 21),   # U6
    "battery_brand":  (28, 12),  # L28
    "num_batteries":  (30, 12),  # L30
    "battery_voltage":(32, 12),  # L32
    "battery_ah":     (34, 12),  # L34
    "num_strings":    (36, 12),  # L36
    "rectifier_brand":(13, 12),  # L13
    "num_modules":    (15, 12),  # L15
    "pld_value":      (36, 26),  # Z36  (PLVD set point location)
    "power_source":   (11, 12),  # L11
}

_LAYOUT_B = {
    "site_name":      (4,  3),   # C4  (same)
    "site_code":      (4,  9),   # I4
    "test_date":      (3, 15),   # O3
    "time_in":        (4, 15),   # O4
    "time_out":       (5, 15),   # O5
    "battery_brand":  (40,  9),  # I40
    "num_batteries":  (43,  9),  # I43
    "battery_voltage":(44,  9),  # I44
    "battery_ah":     (46,  9),  # I46
    "num_strings":    (48,  9),  # I48
    "rectifier_brand":(13,  9),  # I13
    "num_modules":    (17,  9),  # I17
    "pld_value":      (29,  9),  # I29
    "power_source":   (11,  9),  # I11
}

# Layout C: "BDT sheet" production files (e.g. 0167DE) have an empty Excel row 1.
# Calamine skips it, so calamine row N = Excel row N+1.
# All row-sensitive positions are shifted by -1 relative to _LAYOUT_B.
_LAYOUT_C = {
    **_LAYOUT_B,
    "rectifier_brand":(12,  9),  # Excel I13 → calamine row 12
    "num_modules":    (16,  9),  # Excel I17 → calamine row 16
    "pld_value":      (28,  9),  # Excel I29 → calamine row 28
}


def _read_value_near_label(cell_fn, r: int, c: int, max_col: int):
    """Return the first non-empty value to the right of (r, c)."""
    for offset in (1, 2):
        if c + offset <= max_col:
            val = _safe_str(cell_fn(r, c + offset))
            if val:
                return val
    # Try up to 3 more columns to the right as last resort
    for offset in (3, 4, 5):
        if c + offset <= max_col:
            val = _safe_str(cell_fn(r, c + offset))
            if val:
                return val
    return ""


def _parse_battery_info(max_column, cell_fn, data: BDTData, layout: dict | None = None):
    """Extract battery specs from the BDT sheet.

    Uses layout-specific cell positions for fixed extraction, then falls
    back to keyword scanning rows 20-100 if fixed positions miss.
    A second-pass broad scan (rows 1-150) catches templates with unusual
    layouts.
    """
    if layout is None:
        layout = _LAYOUT_A
    brand_raw = ""
    ah_raw = None
    voltage_raw = None
    strings_raw = None

    # ── Layout-based fixed-position extraction ──
    r_brand, c_brand = layout["battery_brand"]
    r_batt, c_batt = layout["num_batteries"]
    r_volt, c_volt = layout["battery_voltage"]
    r_ah, c_ah = layout["battery_ah"]
    r_str, c_str = layout["num_strings"]

    candidate = _safe_str(cell_fn(r_brand, c_brand))
    if candidate:
        brand_raw = candidate

    parsed = _safe_float(cell_fn(r_volt, c_volt))
    if parsed is not None and parsed > 0:
        voltage_raw = parsed

    parsed = _safe_float(cell_fn(r_ah, c_ah))
    if parsed is not None and parsed > 0:
        ah_raw = parsed

    parsed = _safe_float(cell_fn(r_str, c_str))
    if parsed is not None and parsed > 0:
        strings_raw = int(parsed)

    # ── Combined keyword-based scan (rows 1-150) - single pass with early exit ──
    # Use restricted keywords to avoid matching discharge table headers.
    _VOLTAGE_BROAD_KEYWORDS = ("nominal voltage", "battery voltage")
    _AH_BROAD_KEYWORDS = ("ampere hour", "battery capacity", "ampere-hour")
    scan_col_end = min(max(max_column, 9), 16)
    
    if not brand_raw or voltage_raw is None or ah_raw is None or strings_raw is None:
        for r in range(1, 151):
            for c in range(1, scan_col_end + 1):
                val = _safe_str(cell_fn(r, c)).lower()
                if not val:
                    continue

                # Battery brand
                if not brand_raw and any(kw in val for kw in _BRAND_LABEL_KEYWORDS):
                    candidate = _read_value_near_label(cell_fn, r, c, max_column)
                    if not candidate:
                        candidate = _safe_str(cell_fn(r + 1, c))
                    if candidate:
                        brand_raw = candidate

                # Nominal voltage
                if voltage_raw is None and any(kw in val for kw in _VOLTAGE_LABEL_KEYWORDS + _VOLTAGE_BROAD_KEYWORDS):
                    raw = _read_value_near_label(cell_fn, r, c, max_column)
                    if not raw:
                        raw = _safe_str(cell_fn(r + 1, c))
                    parsed = _safe_float(raw) if raw else None
                    if parsed is not None and parsed > 0:
                        voltage_raw = parsed

                # AH
                if ah_raw is None and any(kw in val for kw in _AH_LABEL_KEYWORDS + _AH_BROAD_KEYWORDS):
                    raw = _read_value_near_label(cell_fn, r, c, max_column)
                    if not raw:
                        raw = _safe_str(cell_fn(r + 1, c))
                    parsed = _safe_float(raw) if raw else None
                    if parsed is not None and parsed > 0:
                        ah_raw = parsed

                # Number of strings
                if strings_raw is None and any(kw in val for kw in _STRINGS_LABEL_KEYWORDS):
                    raw = _read_value_near_label(cell_fn, r, c, max_column)
                    if not raw:
                        raw = _safe_str(cell_fn(r + 1, c))
                    parsed = _safe_float(raw) if raw else None
                    if parsed is not None and parsed > 0:
                        strings_raw = int(parsed)

            # Stop early if everything is populated
            if brand_raw and voltage_raw is not None and ah_raw is not None and strings_raw is not None:
                break

    # ── Brand detection from known manufacturer keywords ──
    if brand_raw:
        data.battery_brand = brand_raw
    else:
        # Try to find brand by scanning wider range for known names
        for r in range(20, 101):
            for c in range(0, scan_col_end + 1):
                val = _safe_str(cell_fn(r, c)).lower()
                for kw in _BRAND_KEYWORDS:
                    if kw in val:
                        data.battery_brand = _safe_str(cell_fn(r, c))
                        break
                if data.battery_brand:
                    break
            if data.battery_brand:
                break

    data.battery_ah = ah_raw
    data.battery_voltage = voltage_raw
    data.num_strings = strings_raw


import re as _re

_DATE_IN_FILENAME = _re.compile(r"(\d{1,2})-(\d{1,2})-(\d{4})")
_SITE_CODE_TOKEN = _re.compile(r"(?<![A-Z0-9])(\d{4,5}[A-Z]{2})(?![A-Z0-9])")


def _extract_site_code_token(text: str) -> str:
    """Extract probable site code token (e.g. 0482SI, 4415DE) from text."""
    raw = _safe_str(text).upper()
    if not raw:
        return ""

    matches = [m.group(1) for m in _SITE_CODE_TOKEN.finditer(raw)]
    if not matches:
        return ""

    # Prefer the most frequent token; tie-break by first appearance.
    counts: dict[str, int] = {}
    first_idx: dict[str, int] = {}
    for i, token in enumerate(matches):
        counts[token] = counts.get(token, 0) + 1
        if token not in first_idx:
            first_idx[token] = i
    best = sorted(counts.keys(), key=lambda k: (-counts[k], first_idx[k]))[0]
    return best


def _detect_layout(cell_fn, max_row: int, max_col: int, sheet_name: str | None = None) -> dict:
    """Detect which BDT template layout this file uses.

    Strategy: check Layout A site_code position (L4 = row 4, col 12).
    If that cell contains a plausible site code token, use Layout A.
    Otherwise fall back to Layout B.

    Layout B2 (Rec1/Rec2 family) uses Layout A coordinates - these are detected
    as Layout A by the same signals since they share the coordinate system.
    Layout B1 (Rectifier 1) uses Layout B coordinates - detected as Layout B.
    Layout C (test_pms) has unreliable coordinates - falls back to Layout B + scanning.
    """
    # Layout B2 check: if sheet name is Rec1/Rec2 variant, force Layout A coordinates
    # These sheets use Layout A coordinate system despite different names
    if sheet_name:
        sheet_norm = str(sheet_name).strip()
        layout_b2_variants = ["Rec1", "Rec2", "Rec 1", "Rec 2", "Rect.1", "Rect.2"]
        if sheet_norm in layout_b2_variants:
            logger.debug("Layout A selected: Layout B2 sheet name '%s' uses Layout A coordinates", sheet_name)
            return _LAYOUT_A

    # Layout B1 check: Rectifier 1 uses Layout B coordinates (A1:AC132 range)
    if sheet_name and str(sheet_name).strip() == "Rectifier 1":
        logger.debug("Layout B selected: Layout B1 sheet name 'Rectifier 1' uses Layout B coordinates")
        return _LAYOUT_B

    is_bdt_sheet = sheet_name and str(sheet_name).strip() == "BDT sheet"

    if max_col < 12:
        logger.debug("Layout B selected: max_col < 12")
        return _LAYOUT_B

    # Layout A: site code at L4 (row=4, col=12)
    candidate_a = _safe_str(cell_fn(4, 12))
    if candidate_a and _extract_site_code_token(candidate_a):
        logger.debug("Layout A selected: site code token found at L4")
        return _LAYOUT_A

    # Layout A: also check if test_date cell T3 (row=3, col=20) holds a valid date
    if max_col >= 20:
        date_val = cell_fn(3, 20)
        if _parse_test_date(date_val, "") is not None:
            logger.debug("Layout A selected: valid date at T3")
            return _LAYOUT_A

    # Third signal: check L13 (rectifier brand at Layout A row 13, col 12)
    # A non-empty string there is strong evidence for Layout A even if L4 and T3 are blank
    if max_col >= 12:
        rectifier_a = _safe_str(cell_fn(13, 12))
        if rectifier_a:
            logger.debug("Layout A selected: rectifier brand at L13")
            return _LAYOUT_A

    # Layout C: "BDT sheet" production files where calamine skips an empty Excel row 1,
    # shifting rectifier/modules/pld positions up by one row compared to _LAYOUT_B.
    # Probe each of the three offset fields: if the C-position has content while the
    # B-position is empty, the empty-row-1 offset is active.
    if is_bdt_sheet:
        offset_pairs = [
            (12, 13),  # rectifier_brand: C=(12,9) vs B=(13,9)
            (16, 17),  # num_modules:     C=(16,9) vs B=(17,9)
            (28, 29),  # pld_value:       C=(28,9) vs B=(29,9)
        ]
        for c_row, b_row in offset_pairs:
            at_c = _safe_str(cell_fn(c_row, 9))
            at_b = _safe_str(cell_fn(b_row, 9))
            if at_c and not at_b:
                logger.debug(
                    "Layout C detected: value at (%d,9) with empty (%d,9) confirms empty-row-1 offset",
                    c_row, b_row,
                )
                return _LAYOUT_C
        logger.debug("Layout B selected: 'BDT sheet' with no empty-row-1 offset detected")
        return _LAYOUT_B

    logger.debug("Layout B selected: no Layout A signals detected")
    return _LAYOUT_B


def _detect_layout_family(cell_fn, max_row: int, max_col: int, sheet_name: str | None = None, all_sheet_names: list[str] | None = None) -> tuple[str, str, list[str]]:  # noqa: ARG001 (max_row reserved for future row-level probing)
    """Detect layout family with metadata for FR-001.

    Returns (family, confidence, reasons).
    Family values: "A", "B1", "B2", "C", "UNKNOWN", "SUMMARY_EXCLUDED".
    Confidence: "high", "medium", "low".
    Reasons: list of machine-readable detection signals.
    """
    reasons: list[str] = []
    all_sheet_names = all_sheet_names or []
    
    # Summary/aggregate exclusion check (FR-001)
    normalized_sheets = [str(s).strip().lower() for s in all_sheet_names]
    has_bdt_sheet = any("bdt" in s for s in normalized_sheets)
    has_summary = any("summary" in s for s in normalized_sheets)
    
    if all_sheet_names and not has_bdt_sheet and has_summary:
        reasons.append("summary_only_workbook")
        return "SUMMARY_EXCLUDED", "high", reasons
    
    # Layout C detection
    if sheet_name and str(sheet_name).strip() == "BDT sheet":
        has_power_alarm = any("power alarm" in s.lower() for s in all_sheet_names)
        has_config = any("config" in s.lower() for s in all_sheet_names)
        if has_power_alarm or has_config:
            reasons.append("multi_sheet_test_pms")
            reasons.append("sheet_name_bdt_sheet")
            return "C", "high", reasons
        reasons.append("sheet_name_bdt_sheet_fallback")
        return "C", "medium", reasons
    
    # Layout B2 detection
    if sheet_name:
        sheet_norm = str(sheet_name).strip()
        layout_b2_variants = ["Rec1", "Rec2", "Rec 1", "Rec 2", "Rect.1", "Rect.2"]
        if sheet_norm in layout_b2_variants:
            reasons.append(f"sheet_name_{sheet_norm}")
            reasons.append("layout_b2_family")
            return "B2", "high", reasons
    
    # Layout B1 detection
    if sheet_name and str(sheet_name).strip() == "Rectifier 1":
        reasons.append("sheet_name_rectifier_1")
        return "B1", "high", reasons
    
    # Layout A detection
    if max_col < 12:
        reasons.append("max_col_lt_12")
        return "UNKNOWN", "low", reasons
    
    candidate_a = _safe_str(cell_fn(4, 12))
    if candidate_a and _extract_site_code_token(candidate_a):
        reasons.append("site_code_at_l4")
        return "A", "high", reasons
    
    if max_col >= 20:
        date_val = cell_fn(3, 20)
        if _parse_test_date(date_val, "") is not None:
            reasons.append("valid_date_at_t3")
            return "A", "high", reasons
    
    rectifier_a = _safe_str(cell_fn(13, 12))
    if rectifier_a:
        reasons.append("rectifier_brand_at_l13")
        return "A", "medium", reasons
    
    reasons.append("no_layout_a_signals")
    return "UNKNOWN", "low", reasons


def _resolve_bdt_sheet_name(sheet_names: list[str],
                            filename: str | None = None) -> str | None:
    """Return the best matching BDT sheet name, with pragmatic fallbacks.

    Handles all production sheet name variants:
    - Layout A: "BDT" (97.5% of files)
    - Layout C: "BDT sheet" (test_pms multi-sheet format)
    - Layout B2: "Rec1"/"Rec2", "Rec 1"/"Rec 2", "Rect.1"/"Rect.2" (1.4% - use Layout A coordinates)
    - Layout B1: "Rectifier 1" (0.06% - uses Layout B coordinates)

    Summary/aggregate exclusion (FR-001):
    - Returns None for summary-only workbooks (no BDT sheet present)
    - This allows early exclusion before parsing
    """
    if not sheet_names:
        return None

    # Summary/aggregate exclusion check (FR-001)
    # If only Summary sheet exists (or Summary + Config/Power Alarm without BDT), exclude early
    normalized_sheets = [str(s).strip().lower() for s in sheet_names]
    has_bdt_sheet = any("bdt" in s for s in normalized_sheets)
    has_summary = any("summary" in s for s in normalized_sheets)
    
    if not has_bdt_sheet and has_summary:
        logger.debug("Summary-only workbook detected (no BDT sheet), excluding from parsing")
        return None

    # Exact canonical names first.
    for name in sheet_names:
        if str(name).strip() == "BDT sheet":
            return name
    for name in sheet_names:
        if str(name).strip() == "BDT":
            return name

    # Layout B1: Rectifier 1 singleton (uses Layout B coordinates)
    for name in sheet_names:
        if str(name).strip() == "Rectifier 1":
            return name

    # Layout B2: Rec1/Rec2 family (uses Layout A coordinates)
    # These sheets use Layout A coordinate system but have different names
    layout_b2_variants = ["Rec1", "Rec2", "Rec 1", "Rec 2", "Rect.1", "Rect.2"]
    for name in sheet_names:
        if str(name).strip() in layout_b2_variants:
            return name

    # Case-insensitive match for canonical names.
    for name in sheet_names:
        normalized = str(name).strip().lower()
        if normalized == "bdt sheet":
            return name
        if normalized == "bdt":
            return name
        if normalized == "rectifier 1":
            return name
        if normalized in [v.lower() for v in layout_b2_variants]:
            return name

    # Flexible normalization: allow variants like "BDT_Sheet", "Bdt-Sheet", "BDT SHEET(1)".
    def _norm(text: str) -> str:
        return "".join(ch for ch in str(text).lower() if ch.isalnum())

    for name in sheet_names:
        nm = _norm(name)
        if "bdt" in nm and "sheet" in nm:
            return name

    for name in sheet_names:
        nm = _norm(name)
        if nm.startswith("bdt"):
            return name

    # Last resort for BDT files: many real exports use a generic single/first sheet name.
    # But exclude summary-only workbooks (FR-001)
    if filename and "bdt" in str(filename).lower():
        # Check if the first sheet is a summary sheet
        first_sheet = str(sheet_names[0]).strip().lower() if sheet_names else ""
        if "summary" not in first_sheet:
            return sheet_names[0]
        # If first sheet is summary, don't use fallback
        logger.debug("Fallback resolution skipped: first sheet appears to be summary sheet")

    return None


def _parse_test_date(cell_val, filename: str) -> datetime | None:
    """Coerce a cell value to a date, falling back to filename extraction.

    Calamine may return datetime.date, datetime.datetime, datetime.time,
    or garbage strings.
    """
    import datetime as _dt

    # datetime.date (calamine returns this for date-only cells)
    if isinstance(cell_val, _dt.date) and not isinstance(cell_val, datetime):
        if cell_val.year > 1900:
            return datetime(cell_val.year, cell_val.month, cell_val.day)

    # Already a proper datetime with a real year
    if isinstance(cell_val, datetime) and cell_val.year > 1900:
        return cell_val

    # String — try common formats
    if isinstance(cell_val, str):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(cell_val.strip(), fmt)
            except ValueError:
                continue

    # Fallback: extract date from filename (e.g. "14-1-2026")
    m = _DATE_IN_FILENAME.search(filename)
    if m:
        try:
            return datetime(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    return None


def _parse_summary_sheet(file_path: str, sheet_names: list[str]) -> dict[str, str]:
    """Read row 2 of the Summary sheet into a header->value dict."""
    target = None
    for name in sheet_names:
        if str(name).strip().lower() == "summary":
            target = name
            break
    if target is None:
        return {}

    rows = None
    try:
        import python_calamine
        wb = python_calamine.CalamineWorkbook.from_path(file_path)
        rows = wb.get_sheet_by_name(target).to_python()
    except Exception:
        pass

    if rows is None:
        try:
            owb = load_workbook(file_path, data_only=True)
            ws = owb[target]
            rows = []
            for row_cells in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 2),
                                          max_col=ws.max_column):
                rows.append([c.value for c in row_cells])
            owb.close()
        except Exception:
            return {}

    if not rows or len(rows) < 2:
        return {}

    headers = rows[0]
    values = rows[1]
    result: dict[str, str] = {}
    for i, hdr in enumerate(headers):
        hdr_str = _safe_str(hdr)
        if not hdr_str:
            continue
        val = values[i] if i < len(values) else None
        result[hdr_str] = _safe_str(val) if val is not None else ""
    return result


def parse_bdt_file(file_path: str, *, skip_photos: bool = False) -> BDTData:
    """Parse a single BDT Excel file and return structured data.

    Uses python-calamine (Rust-based reader, ~100x faster than openpyxl)
    for cell data extraction.  Falls back to openpyxl if calamine fails.

    Args:
        skip_photos: If True, skip photo extraction for faster batch
            processing.  Photos can be loaded later via
            ``load_bdt_photos()``.
    """
    import os
    data = BDTData(file_path=file_path, filename=os.path.basename(file_path))

    # ── Read sheet data with calamine (fast path) ────────
    rows = None
    all_sheet_names: list[str] = []
    try:
        import python_calamine
        wb = python_calamine.CalamineWorkbook.from_path(file_path)
        all_sheet_names = list(wb.sheet_names)
        bdt_sheet_name = _resolve_bdt_sheet_name(all_sheet_names, data.filename)
        if bdt_sheet_name is None:
            data.errors.append("Missing 'BDT sheet'")
            return data
        rows = wb.get_sheet_by_name(bdt_sheet_name).to_python()
    except Exception:
        pass

    # ── Fallback to openpyxl whenever calamine didn't yield rows ──
    if rows is None:
        try:
            owb = load_workbook(file_path, data_only=True)
        except Exception as e:
            data.errors.append(f"Cannot open file: {e}")
            return data
        all_sheet_names = list(owb.sheetnames)
        bdt_sheet_name = _resolve_bdt_sheet_name(all_sheet_names, data.filename)
        if bdt_sheet_name is None:
            data.errors.append("Missing 'BDT sheet'")
            owb.close()
            return data
        ows = owb[bdt_sheet_name]
        rows = []
        for row_cells in ows.iter_rows(min_row=1, max_row=ows.max_row,
                                        max_col=ows.max_column):
            rows.append([c.value for c in row_cells])
        owb.close()

    if rows is None:
        if not data.errors:
            data.errors.append("Failed to read BDT sheet")
        return data

    max_row = len(rows)
    max_col = max((len(r) for r in rows), default=0)

    def cell(row, col):
        """1-indexed cell access (matches Excel row/col numbers)."""
        r, c = row - 1, col - 1
        if r < 0 or r >= max_row:
            return None
        row_data = rows[r]
        if c < 0 or c >= len(row_data):
            return None
        return row_data[c]

    # Detect layout for coordinate selection
    layout = _detect_layout(cell, max_row, max_col, sheet_name=bdt_sheet_name)
    data.core_layout = "Layout A" if layout is _LAYOUT_A else "Layout B"
    
    # Detect layout family for metadata (FR-001)
    family, confidence, reasons = _detect_layout_family(
        cell, max_row, max_col, sheet_name=bdt_sheet_name, all_sheet_names=all_sheet_names
    )
    data.core_layout_family = family
    data.detection_confidence = confidence
    data.detection_reasons = reasons
    data.parser_mode = "fast_family"
    
    # Skip parsing for summary-only workbooks (FR-001)
    if family == "SUMMARY_EXCLUDED":
        data.errors.append("Summary/aggregate workbook excluded from BDT validation")
        return data
    data.site_name = _safe_str(cell(*layout["site_name"]))
    data.site_code = _safe_str(cell(*layout["site_code"]))
    data.test_date = _parse_test_date(cell(*layout["test_date"]), data.filename)
    data.time_in = _safe_str(cell(*layout["time_in"]))
    data.time_out = _safe_str(cell(*layout["time_out"]))

    # Try alternate fixed positions based on detected layout
    if data.test_date is None:
        if layout is _LAYOUT_A:
            alt_date_cells = ((2, 20), (3, 19), (3, 21), (4, 20))
        else:
            alt_date_cells = ((2, 15), (3, 14), (3, 16), (1, 15))
        for rr, cc in alt_date_cells:
            parsed = _parse_test_date(cell(rr, cc), data.filename)
            if parsed is not None:
                data.test_date = parsed
                break

    # Fallback for time_in/time_out if not found at layout position
    if not data.time_in or data.time_in in ("Time In", "Time in:", ""):
        if layout is _LAYOUT_A:
            alt_time_in_cells = ((5, 20), (5, 21), (5, 19))
        else:
            alt_time_in_cells = ((4, 15), (4, 16), (4, 14))
        for rr, cc in alt_time_in_cells:
            val = _safe_str(cell(rr, cc))
            if val and val not in ("Time In", "Time in:", ""):
                data.time_in = val
                break

    if not data.time_out or data.time_out in ("Time Out", "Time Out:", ""):
        if layout is _LAYOUT_A:
            alt_time_out_cells = ((6, 20), (6, 21), (6, 19))
        else:
            alt_time_out_cells = ((5, 15), (5, 16), (5, 14))
        for rr, cc in alt_time_out_cells:
            val = _safe_str(cell(rr, cc))
            if val and val not in ("Time Out", "Time Out:", ""):
                data.time_out = val
                break

    # Keyword-based fallback scan for test_date
    if data.test_date is None:
        for r in range(1, min(max_row, 25) + 1):
            found = False
            for c in range(0, min(max_col, 16) + 1):
                text = _safe_str(cell(r, c)).lower()
                if not text:
                    continue
                if (any(kw in text for kw in ("test date", "date of test",
                                               "التاريخ"))
                        or (text == "date" and c <= 2)):
                    for nc in range(c + 1, min(max_col, c + 8) + 1):
                        parsed = _parse_test_date(cell(r, nc), data.filename)
                        if parsed is not None:
                            data.test_date = parsed
                            found = True
                            break
                    if found:
                        break
            if data.test_date is not None:
                break

    # Final filename-only fallback (in case no cell yielded a date)
    if data.test_date is None:
        data.test_date = _parse_test_date(None, data.filename)

    for r in range(1, min(max_row, 20) + 1):
        label = _safe_str(cell(r, 1)).lower()
        if not label:
            label = _safe_str(cell(r, 2)).lower()
        if not label:
            continue

        if not data.site_code and ("site code" in label or label == "site id"):
            for c in range(2, min(max_col, 12) + 1):
                candidate = _safe_str(cell(r, c))
                token = _extract_site_code_token(candidate)
                if token:
                    data.site_code = token
                    break
                if candidate and "site" not in candidate.lower():
                    data.site_code = candidate
                    break

        if not data.site_name and "site name" in label:
            for c in range(2, min(max_col, 12) + 1):
                candidate = _safe_str(cell(r, c))
                if candidate and "site" not in candidate.lower():
                    data.site_name = candidate
                    break

        if not data.time_in and "time in" in label:
            for c in range(2, min(max_col, 20) + 1):
                candidate = _safe_str(cell(r, c))
                if candidate:
                    data.time_in = candidate
                    break

        if not data.time_out and "time out" in label:
            for c in range(2, min(max_col, 20) + 1):
                candidate = _safe_str(cell(r, c))
                if candidate:
                    data.time_out = candidate
                    break

    # Normalize + fallback site-code extraction:
    # Prefer explicit token sources over any free-text value.
    filename_token = _extract_site_code_token(data.filename)
    sheet_site_token = _extract_site_code_token(data.site_code)
    site_name_token = _extract_site_code_token(data.site_name)

    if sheet_site_token:
        data.site_code = sheet_site_token
    elif filename_token:
        data.site_code = filename_token
    elif site_name_token:
        data.site_code = site_name_token
    elif data.site_code:
        # Keep legacy free-text only as last resort.
        data.site_code = data.site_code.strip().upper()

    # Discharge test table — find by scanning for "Batteries discharge test"
    discharge_start_row = None
    for r in range(1, max_row + 1):
        for c in range(1, min(max_col, 8) + 1):
            v = _safe_str(cell(r, c))
            if "batteries discharge test" in v.lower():
                discharge_start_row = r
                break
        if discharge_start_row is not None:
            break

    if discharge_start_row:
        # Detect table columns by header row ("Rec Bus Bar", "String #")
        rec_v_col, rec_a_col = 4, 5
        string_a_cols: list[int] = []
        hdr_row, hdr_col = _find_text_in_row_window(
            cell, max_col, discharge_start_row, discharge_start_row + 8,
            needles=("rec", "bus", "bar"),
        )
        if hdr_row and hdr_col:
            rec_v_col = hdr_col
            rec_a_col = min(max_col, hdr_col + 1)
            for c in range(hdr_col + 2, max_col + 1):
                txt = _safe_str(cell(hdr_row, c)).lower()
                if "string" in txt:
                    a_col = c + 1
                    if a_col <= max_col:
                        string_a_cols.append(a_col)
        else:
            # Legacy fixed template
            for sc in range(7, 22, 2):
                if sc <= max_col:
                    string_a_cols.append(sc)

        # Scan for "Before disconnecting Rectifier" row
        data_row, label_col = _find_text_in_row_window(
            cell, max_col, discharge_start_row + 1, discharge_start_row + 12,
            needles=("before", "disconnect"),
            search_cols=6,
        )
        if data_row is None:
            data_row = discharge_start_row + 3
        if label_col is None:
            label_col = 1

        data.start_voltage = _safe_float(cell(data_row, rec_v_col))
        data.start_ampere = _safe_float(cell(data_row, rec_a_col))

        string_amps = []
        for sc in string_a_cols:
            sa = _safe_float(cell(data_row, sc))
            if sa is not None:
                string_amps.append(sa)
        if string_amps:
            data.ibat_before_test = max(string_amps)
            data.starting_ibattery_ampere = data.ibat_before_test

        # Capture per-string (V, A) pairs for the "Before disconnecting" row
        all_string_readings: list[list[tuple[float | None, float | None]]] = []
        if string_a_cols:
            before_row_pairs: list[tuple[float | None, float | None]] = []
            for sc in string_a_cols:
                sv = _safe_float(cell(data_row, sc - 1))
                sa = _safe_float(cell(data_row, sc))
                before_row_pairs.append((sv, sa))
            all_string_readings.append(before_row_pairs)

        # Discharge time-series
        last_filled_mins = 0.0
        last_filled_voltage = None
        last_filled_ampere = None
        r = data_row + 1
        while r <= data_row + 30:
            lbl = _extract_row_label(cell, r, label_col)
            if "after connecting" in lbl.lower():
                data.after_reconnect_voltage = _safe_float(cell(r, rec_v_col))
                data.after_reconnect_ampere = _safe_float(cell(r, rec_a_col))
                break
            if not lbl:
                r += 1
                continue
            v = _safe_float(cell(r, rec_v_col))
            a = _safe_float(cell(r, rec_a_col))
            row_string_vals = [_safe_float(cell(r, sc)) for sc in string_a_cols]
            if string_a_cols:
                row_string_pairs: list[tuple[float | None, float | None]] = []
                for sc in string_a_cols:
                    sv = _safe_float(cell(r, sc - 1))
                    sa = _safe_float(cell(r, sc))
                    row_string_pairs.append((sv, sa))
                all_string_readings.append(row_string_pairs)
            has_data = _has_meaningful_numeric(v, a, *row_string_vals)

            if not has_data:
                # Hide formula-driven zeros/placeholder rows from parsed readings.
                v = None
                a = None

            data.discharge_readings.append((lbl, v, a))
            if v is not None:
                last_filled_voltage = v
            if a is not None:
                last_filled_ampere = a
            if has_data:
                try:
                    last_filled_mins = float(str(lbl).split()[0])
                except (ValueError, TypeError, IndexError):
                    pass
            r += 1

        data.discharge_minutes = last_filled_mins
        data.end_voltage = last_filled_voltage
        data.end_ampere  = last_filled_ampere
        data.string_discharge_readings = all_string_readings

    _parse_battery_info(max_col, cell, data, layout)

    # ── Rectifier / module / battery-count / PLVD / power_source fields ──
    r_rect, c_rect = layout["rectifier_brand"]
    r_mod, c_mod = layout["num_modules"]
    r_batt2, c_batt2 = layout["num_batteries"]
    r_pld, c_pld = layout["pld_value"]
    r_psrc, c_psrc = layout["power_source"]  # Now in both Layout A and Layout B maps

    data.rectifier_brand = _safe_str(cell(r_rect, c_rect))
    _mod_raw = _safe_float(cell(r_mod, c_mod))
    if _mod_raw is not None and _mod_raw > 0:
        data.num_modules = int(_mod_raw)
    _batt_raw = _safe_float(cell(r_batt2, c_batt2))
    if _batt_raw is not None and _batt_raw > 0:
        data.num_batteries = int(_batt_raw)
    data.pld_value = _safe_str(cell(r_pld, c_pld))
    data.power_source = _safe_str(cell(r_psrc, c_psrc))

    # Keyword-fallback scanning
    if not data.rectifier_brand:
        _r, _c = _find_text_in_row_window(cell, max_col, 10, 16,
                                           needles=("rectifier", "type"))
        if _r is not None:
            val = (_safe_str(cell(_r, _c + 1))
                   or _safe_str(cell(_r, _c))
                   or _safe_str(cell(_r, c_rect))
                   or _safe_str(cell(_r, c_rect - 1))
                   or _safe_str(cell(_r, c_rect + 1)))
            data.rectifier_brand = val

    if data.num_modules is None:
        _r, _c = _find_text_in_row_window(cell, max_col, 13, 18,
                                           needles=("number", "modules"))
        if _r is not None:
            _v = (_safe_float(cell(_r, c_mod))
                  or _safe_float(cell(_r, _c + 1))
                  or _safe_float(cell(_r, _c)))
            if _v is not None and _v > 0:
                data.num_modules = int(_v)

    if data.num_batteries is None:
        _r, _c = _find_text_in_row_window(cell, max_col, 28, 46,
                                           needles=("number", "batteries", "connected"))
        if _r is not None:
            _v = (_safe_float(cell(_r, c_batt2))
                  or _safe_float(cell(_r, _c + 1))
                  or _safe_float(cell(_r, _c)))
            if _v is not None and _v > 0:
                data.num_batteries = int(_v)

    if not data.pld_value:
        _r, _c = _find_text_in_row_window(cell, max_col, 20, 35,
                                           needles=("plvd", "set"))
        if _r is None:
            _r, _c = _find_text_in_row_window(cell, max_col, 20, 35,
                                               needles=("lvd", "disconnect"))
        if _r is not None:
            data.pld_value = (_safe_str(cell(_r, c_pld))
                              or _safe_str(cell(_r, _c + 1))
                              or _safe_str(cell(_r, _c)))

    # Summary sheet
    data.summary_data = _parse_summary_sheet(file_path, all_sheet_names)

    # Photo slots
    if not skip_photos:
        (
            data.photo_slots,
            data.photo_count,
            data.photo_layout_id,
            data.required_photo_count,
            data.photo_mapping_confidence,
            data.photo_detection_mode,
            data.photo_categories_found,
        ) = _extract_photo_slots(
            file_path,
            family_guess=family,
            family_confidence=confidence,
            bdt_sheet_name=bdt_sheet_name,
        )
        if data.photo_detection_mode == "structural":
            data.parser_mode = "structural"
        data.photos_deferred = False
        # Set required photo categories from constants (FR-003)
        from alarm_app.constants import BDT_REQUIRED_PHOTO_CATEGORIES
        data.required_photo_categories = list(BDT_REQUIRED_PHOTO_CATEGORIES)
    else:
        data.photos_deferred = True
        data.photo_detection_mode = "deferred"

    return data


def load_bdt_photos(bdt: BDTData) -> None:
    """Lazy-load photo slots for a BDTData that was parsed with skip_photos."""
    if bdt.photo_slots:
        bdt.photos_deferred = False
        return  # already loaded
    (
        bdt.photo_slots,
        bdt.photo_count,
        bdt.photo_layout_id,
        bdt.required_photo_count,
        bdt.photo_mapping_confidence,
        bdt.photo_detection_mode,
        bdt.photo_categories_found,
    ) = _extract_photo_slots(
        bdt.file_path,
        family_guess=bdt.core_layout_family,
        family_confidence=bdt.detection_confidence,
    )
    if bdt.photo_detection_mode == "structural":
        bdt.parser_mode = "structural"
    bdt.photos_deferred = False
    # Set required photo categories from constants (FR-003)
    from alarm_app.constants import BDT_REQUIRED_PHOTO_CATEGORIES
    bdt.required_photo_categories = list(BDT_REQUIRED_PHOTO_CATEGORIES)


# ── Photo slot definitions by layout ───────────────────────────────
_PHOTO_LAYOUTS = {
    # 6-photo template (rectifier + batteries only, 3 cols per band)
    "LAYOUT_PHOTO_6": {
        "slot_defs": [
            (9, 13), (9, 18), (9, 23),
            (21, 13), (21, 18), (21, 23),
        ],
        "band_ranges": [(8, 21), (21, 34)],
        "col_groups": [(11, 16), (17, 21), (22, 26)],
        "band_categories": {0: "rectifier", 1: "batteries"},
        "required_count": 6,
    },
    # 15-photo template (5 bands x 3 cols)
    "LAYOUT_PHOTO_15": {
        "slot_defs": [
            (9, 13), (9, 18), (9, 23),
            (21, 13), (21, 18), (21, 23),
            (34, 13), (34, 18), (34, 23),
            (46, 13), (46, 18), (46, 23),
            (58, 13), (58, 18), (58, 23),
        ],
        "band_ranges": [(8, 21), (21, 34), (34, 46), (46, 58), (58, 70)],
        "col_groups": [(11, 16), (17, 21), (22, 26)],
        "band_categories": {
            0: "rectifier",
            1: "batteries",
            2: "modules",
            3: "load",
            4: "charging",
        },
        "required_count": 15,
    },
    # 16-photo+ template using 4th extra column group; require 16 minimum
    "LAYOUT_PHOTO_16": {
        "slot_defs": [
            (9, 13), (9, 18), (9, 23), (9, 28),
            (21, 13), (21, 18), (21, 23), (21, 28),
            (34, 13), (34, 18), (34, 23), (34, 28),
            (46, 13), (46, 18), (46, 23), (46, 28),
            (58, 13), (58, 18), (58, 23), (58, 28),
        ],
        "band_ranges": [(8, 21), (21, 34), (34, 46), (46, 58), (58, 70)],
        "col_groups": [(11, 16), (17, 21), (22, 26), (27, 31)],
        "band_categories": {
            0: "rectifier",
            1: "batteries",
            2: "modules",
            3: "load",
            4: "charging",
        },
        "required_count": 16,
    },
}


def _anchor_to_slot(from_row: int, from_col: int,
                    band_ranges: list[tuple[int, int]],
                    col_groups: list[tuple[int, int]]) -> int | None:
    """Map a 0-indexed anchor position to a slot index, or None."""
    band = None
    for bi, (lo, hi) in enumerate(band_ranges):
        if lo <= from_row < hi:  # exclusive upper bound prevents overlap
            band = bi
            break
    if band is None:
        return None

    col_grp = None
    for ci, (lo, hi) in enumerate(col_groups):
        if lo <= from_col <= hi:
            col_grp = ci
            break
    if col_grp is None:
        return None

    return band * len(col_groups) + col_grp


def _select_photo_layout(anchor_count: int, max_anchor_col: int) -> tuple[str, int]:
    """Select photo layout ID and required photo count from anchor count."""
    if anchor_count <= 7:
        layout_id = "LAYOUT_PHOTO_6"
    elif 13 <= anchor_count <= 15:
        layout_id = "LAYOUT_PHOTO_15"
    elif anchor_count >= 16:
        layout_id = "LAYOUT_PHOTO_16"
    else:
        # Dead zone: 8-12 anchors (corrupt 15-photo file with missing photos)
        # Explicit rule: 8-12 anchors with max_anchor_col >= 22 maps to LAYOUT_PHOTO_15
        # else LAYOUT_PHOTO_6. This handles corrupt files without changing behavior
        # on observed files.
        if 8 <= anchor_count <= 12:
            if max_anchor_col >= 22:
                layout_id = "LAYOUT_PHOTO_15"
            else:
                layout_id = "LAYOUT_PHOTO_6"
        else:
            # Deterministic fallback for sparse/malformed files.
            if max_anchor_col >= 27:
                layout_id = "LAYOUT_PHOTO_16"
            elif max_anchor_col >= 22:
                layout_id = "LAYOUT_PHOTO_15"
            else:
                layout_id = "LAYOUT_PHOTO_6"

    required = int(_PHOTO_LAYOUTS[layout_id]["required_count"])
    return layout_id, required


def _extract_photo_slots_structural(
    file_path: str,
    *,
    family_guess: str,
    family_confidence: str,
    bdt_sheet_name: str | None,
) -> tuple[list[PhotoSlot], int, str, int, str, str, list[str]]:
    """Section-first photo extraction from OOXML primitives.

    This path is used for unknown/low-confidence files and Layout C where
    fixed slot geometry can drift.
    """
    from alarm_app.bdt.image_assigner import assign_manifest_images
    from alarm_app.bdt.ooxml_reader import OOXMLPackage
    from alarm_app.bdt.section_parser import build_workbook_manifest

    with OOXMLPackage(file_path) as pkg:
        workbook = pkg.read_workbook_xml()
        sheet_names = [s.attrib.get("name", "") for s in workbook.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet")]
        target_sheet = bdt_sheet_name or _resolve_bdt_sheet_name(sheet_names)
        if not target_sheet:
            return [], len(pkg.list_media_files()), "LAYOUT_PHOTO_6", 6, "low", "structural", []

        ws_path = pkg.resolve_worksheet_xml_path(target_sheet)
        if not ws_path:
            return [], len(pkg.list_media_files()), "LAYOUT_PHOTO_6", 6, "low", "structural", []

        shared_strings = pkg.parse_shared_strings()
        cell_map, merged_ranges, style_ids = pkg.parse_worksheet_cells(ws_path, shared_strings=shared_strings)
        manifest = build_workbook_manifest(
            sheet_name=target_sheet,
            cell_map=cell_map,
            merged_ranges=merged_ranges,
            style_ids=style_ids,
            family_guess=family_guess,
            family_confidence=family_confidence,
            parser_mode="structural",
        )

        anchor_dicts: list[dict] = []
        drawing_paths = pkg.get_worksheet_drawing_paths(ws_path)
        for drawing_path in drawing_paths:
            for anchor in pkg.extract_two_cell_anchors(drawing_path):
                # Skip banner/logo anchors left of photo region.
                if anchor.from_col < 11:
                    continue
                anchor_dicts.append(
                    {
                        "sheet_name": target_sheet,
                        "from_row": anchor.from_row,
                        "from_col": anchor.from_col,
                        "to_row": anchor.to_row,
                        "to_col": anchor.to_col,
                        "r_id": anchor.r_id,
                        "media_path": anchor.media_path,
                        "drawing_path": drawing_path,
                    }
                )

        assign_manifest_images(manifest, anchor_dicts)

        max_anchor_col = max((a["from_col"] for a in anchor_dicts), default=-1)
        photo_layout_id, required_photo_count = _select_photo_layout(len(anchor_dicts), max_anchor_col)

        slots: list[PhotoSlot] = []
        categories_found: list[str] = []
        seen_image_keys: set[str] = set()
        for section in manifest.sections:
            if not section.images:
                slots.append(
                    PhotoSlot(
                        label=section.header_text or section.section_id,
                        image_data=None,
                        image_ext="",
                        category=section.category,
                    )
                )
                continue

            for image in section.images:
                image_data = None
                image_ext = ""
                if image.media_path:
                    try:
                        image_data = pkg.read_media(image.media_path)
                        image_ext = image.media_path.rsplit(".", 1)[-1].lower()
                    except Exception:
                        image_data = None
                        image_ext = ""

                # Deduplicate repeated embedded media that may be anchored
                # multiple times across overlapping/variant sections.
                image_key = ""
                if image.media_path:
                    image_key = f"{section.section_id}|path:{image.media_path}"
                    if image_data:
                        image_key = f"{image_key}|bytes:{hashlib.sha256(image_data).hexdigest()}"
                elif image_data:
                    image_key = f"{section.section_id}|bytes:{hashlib.sha256(image_data).hexdigest()}"
                if image_key and image_key in seen_image_keys:
                    continue
                if image_key:
                    seen_image_keys.add(image_key)

                label = section.header_text or section.section_id
                slots.append(
                    PhotoSlot(
                        label=label,
                        image_data=image_data,
                        image_ext=image_ext,
                        category=section.category,
                    )
                )
                if image_data and section.category and section.category not in categories_found:
                    categories_found.append(section.category)

        filled_slots = sum(1 for s in slots if s.image_data)
        if filled_slots == 0:
            mapping_confidence = "low"
        elif manifest.orphan_images:
            mapping_confidence = "medium"
        else:
            mapping_confidence = "high"

        return (
            slots,
            len(pkg.list_media_files()),
            photo_layout_id,
            required_photo_count,
            mapping_confidence,
            "structural",
            categories_found,
        )


def _extract_photo_slots(
    file_path: str,
    *,
    family_guess: str = "",
    family_confidence: str = "",
    bdt_sheet_name: str | None = None,
) -> tuple[list[PhotoSlot], int, str, int, str, str, list[str]]:
    """Extract labelled photo slots using structural parsing only."""
    try:
        return _extract_photo_slots_structural(
            file_path,
            family_guess=family_guess,
            family_confidence=family_confidence,
            bdt_sheet_name=bdt_sheet_name,
        )
    except Exception:
        logger.warning("Structural photo extraction failed", exc_info=True)
        return [], 0, "LAYOUT_PHOTO_6", 6, "low", "structural", []


def _extract_photo_slots_layout(file_path: str) -> tuple[list[PhotoSlot], int, str, int, str, str, list[str]]:
    """Extract labelled photo slots from a BDT xlsx file.

    Returns (list_of_PhotoSlot, total_media_count, photo_layout_id, required_count,
             mapping_confidence, detection_mode, categories_found).
    """
    import zipfile
    import xml.etree.ElementTree as ET

    ns_xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"
    ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    try:
        zf = zipfile.ZipFile(file_path)
    except Exception:
        logger.warning("Failed to open xlsx as zip, using conservative 6-photo fallback")
        return [], 0, "LAYOUT_PHOTO_6", 6, "low", "fallback", []

    try:
        # Total media count (for backwards compat)
        media_files = [n for n in zf.namelist() if n.startswith("xl/media/")]
        total_media = len(media_files)

        # Find the drawing XML that belongs to the BDT sheet only.
        # Each worksheet references its own drawing via its .rels file.
        # We must NOT parse drawings from other sheets (e.g. Power Alarm)
        # because their images (alarm emails, BTS screenshots) would
        # contaminate the BDT photo grid.
        all_names = zf.namelist()

        # Step 1: Find which sheet XML file is the BDT sheet
        ns_ss = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        bdt_sheet_file = None
        try:
            wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
            wb_rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            rid_to_sheet = {}
            for rel in wb_rels_xml:
                rid_to_sheet[rel.get("Id", "")] = rel.get("Target", "")
            bdt_sheet_name = _resolve_bdt_sheet_name(
                [s.get("name", "") for s in wb_xml.findall(f".//{{{ns_ss}}}sheet")])
            if bdt_sheet_name:
                ns_r_wb = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
                for s in wb_xml.findall(f".//{{{ns_ss}}}sheet"):
                    if s.get("name") == bdt_sheet_name:
                        rid = s.get(f"{{{ns_r_wb}}}id", "")
                        target = rid_to_sheet.get(rid, "")
                        if target:
                            bdt_sheet_file = "xl/" + target if not target.startswith("xl/") else target
                        break
        except Exception:
            pass

        # Step 2: Find drawing path from BDT sheet's rels
        drawing_paths = []
        if bdt_sheet_file:
            sheet_basename = bdt_sheet_file.rsplit("/", 1)[-1]
            sheet_rels = f"xl/worksheets/_rels/{sheet_basename}.rels"
            if sheet_rels in all_names:
                try:
                    sr_xml = ET.fromstring(zf.read(sheet_rels))
                    for rel in sr_xml:
                        target = rel.get("Target", "")
                        if "drawing" in target.lower():
                            # Resolve relative path
                            dp = "xl/drawings/" + target.rsplit("/", 1)[-1]
                            if dp in all_names:
                                drawing_paths.append(dp)
                except Exception:
                    pass

        # Fallback: if we couldn't determine the BDT sheet's drawing,
        # use only drawing1.xml (the first sheet's drawing)
        if not drawing_paths:
            fallback = "xl/drawings/drawing1.xml"
            if fallback in all_names:
                drawing_paths = [fallback]

        if not drawing_paths:
            logger.warning("No drawing path found for BDT sheet, using conservative 6-photo fallback")
            return [], total_media, "LAYOUT_PHOTO_6", 6, "low", "fallback", []

        # Build rId → media zip path map from BDT drawing rels only
        rid_to_path: dict[str, str] = {}
        for dp in drawing_paths:
            dp_basename = dp.rsplit("/", 1)[-1]
            rels_path = f"xl/drawings/_rels/{dp_basename}.rels"
            if rels_path in all_names:
                rels_xml = ET.fromstring(zf.read(rels_path))
                for rel in rels_xml:
                    rid = rel.get("Id", "")
                    target = rel.get("Target", "")
                    if target.startswith("../media/"):
                        rid_to_path[rid] = "xl/media/" + target.split("/")[-1]

        # Parse BDT drawing only — extract twoCellAnchor positions + rIds
        all_slot_images: dict[int, tuple[str, str, int]] = {}
        valid_anchor_count = 0
        max_anchor_col = -1

        for dp in drawing_paths:
            drawing_xml = ET.fromstring(zf.read(dp))
            for anchor in drawing_xml.findall(f"{{{ns_xdr}}}twoCellAnchor"):
                frm = anchor.find(f"{{{ns_xdr}}}from")
                if frm is None:
                    continue
                from_col = int(frm.find(f"{{{ns_xdr}}}col").text)
                from_row = int(frm.find(f"{{{ns_xdr}}}row").text)

                # Skip non-photo anchors (logo at col 0, etc.)
                if from_col < 11:
                    continue

                # Find embedded image rId
                blip = anchor.find(f".//{{{ns_a}}}blip")
                if blip is None:
                    continue
                rid = blip.get(f"{{{ns_r}}}embed", "")
                if not rid or rid not in rid_to_path:
                    continue

                slot_idx = _anchor_to_slot(
                    from_row,
                    from_col,
                    _PHOTO_LAYOUTS["LAYOUT_PHOTO_16"]["band_ranges"],
                    _PHOTO_LAYOUTS["LAYOUT_PHOTO_16"]["col_groups"],
                )
                if slot_idx is None:
                    continue  # Non-slot image, don't count it

                # Only count anchors that map to valid photo slots
                valid_anchor_count += 1
                if from_col > max_anchor_col:
                    max_anchor_col = from_col
                # On duplicate, keep the anchor whose row is closest to the
                # slot's expected label row so the right image wins.
                full_slot_defs = _PHOTO_LAYOUTS["LAYOUT_PHOTO_16"]["slot_defs"]
                if slot_idx in all_slot_images:
                    expected_row = full_slot_defs[slot_idx][0] if slot_idx < len(full_slot_defs) else 0
                    prev_row = all_slot_images[slot_idx][2]
                    if abs(from_row - expected_row) >= abs(prev_row - expected_row):
                        continue  # existing anchor is closer, keep it

                all_slot_images[slot_idx] = (rid, rid_to_path[rid], from_row)

        photo_layout_id, required_photo_count = _select_photo_layout(
            valid_anchor_count,
            max_anchor_col,
        )
        layout = _PHOTO_LAYOUTS[photo_layout_id]
        slot_defs = layout["slot_defs"]
        band_categories = layout["band_categories"]
        target_cols_per_band = len(layout["col_groups"])

        # Remap from full 4-col index to selected layout index when needed.
        slot_images: dict[int, tuple[str, str, int]] = {}
        for idx4, meta in all_slot_images.items():
            band = idx4 // 4
            col_in_band = idx4 % 4
            if col_in_band >= target_cols_per_band:
                continue
            mapped_idx = band * target_cols_per_band + col_in_band
            if mapped_idx >= len(slot_defs):
                continue
            slot_images[mapped_idx] = meta

        # Read labels from the worksheet and build PhotoSlot list
        # Re-open with openpyxl just for label reading
        wb = load_workbook(file_path, data_only=True)
        sheet_name = _resolve_bdt_sheet_name(list(wb.sheetnames))
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else None

        slots: list[PhotoSlot] = []
        for idx, (label_row, label_col) in enumerate(slot_defs):
            # Read label from worksheet
            label = ""
            if ws is not None:
                raw = ws.cell(row=label_row, column=label_col).value
                label = _safe_str(raw) if raw else ""
            if not label:
                label = f"Slot {idx + 1}"

            img_data = None
            img_ext = ""
            if idx in slot_images:
                _, media_path, _ = slot_images[idx]
                try:
                    img_data = zf.read(media_path)
                    img_ext = media_path.rsplit(".", 1)[-1].lower()
                except Exception:
                    pass

            slots.append(PhotoSlot(
                label=label,
                image_data=img_data,
                image_ext=img_ext,
                category=band_categories.get(idx // target_cols_per_band, "other"),
            ))

        wb.close()
        
        # Extract category metadata (FR-003, FR-004)
        categories_found: list[str] = []
        for slot in slots:
            if slot.image_data and slot.category:
                cat = slot.category.lower()
                if cat not in categories_found:
                    categories_found.append(cat)
        
        # Determine mapping confidence based on anchor-to-slot mapping success
        filled_slots = sum(1 for s in slots if s.image_data)
        mapping_confidence = "high" if filled_slots >= required_photo_count else "medium"
        if filled_slots == 0:
            mapping_confidence = "low"
        
        detection_mode = "normal"  # Could be "fallback" if we used non-standard mapping
        
        return slots, total_media, photo_layout_id, required_photo_count, mapping_confidence, detection_mode, categories_found

    finally:
        zf.close()
